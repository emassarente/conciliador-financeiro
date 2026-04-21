# =============================================================================
# DASHBOARD - INTERFACE VISUAL DA CONCILIAÇÃO FINANCEIRA
# Tecnologia: Streamlit
#
# Funcionalidades:
# - Upload dos arquivos (Razão e Extrato)
# - Execução da conciliação com barra de progresso
# - Tabela colorida com os resultados:
#     🟢 VERDE  = CONCILIADO
#     🟡 AMARELO = MATCH_COMBINADO
#     🔵 AZUL   = MATCH_PROVAVEL
#     🔴 VERMELHO = NAO_CONCILIADO
# - Métricas (KPIs) no topo
# - Tabela principal com separação visual Razão | Extrato
# - Detalhe expansível por linha (clique para ver como foi conciliado)
# - Seção de conciliação MANUAL com checkboxes
# - Filtros por status
# - Download do resultado em Excel
# =============================================================================

import sys
import os
import logging
import pandas as pd
import streamlit as st
from pathlib import Path
from io import BytesIO
import datetime

# Adiciona o diretório raiz ao path para que os módulos sejam encontrados
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from import_.razao_parser import RazaoParser
from import_.extrato_parser import ExtratoParser
from engine.conciliacao_engine import ConciliacaoEngine

# Configura logging para o Streamlit
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURAÇÃO DA PÁGINA
# =============================================================================
st.set_page_config(
    page_title="Conciliação Financeira",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# CSS PERSONALIZADO - CORES E ESTILOS DA TABELA
# =============================================================================
st.markdown("""
<style>
    /* Fundo suave para o app */
    .main { background-color: #f8f9fa; }

    /* Cabeçalho principal */
    .titulo-principal {
        font-size: 2.2rem;
        font-weight: 800;
        color: #1a1a2e;
        margin-bottom: 0;
    }
    .subtitulo {
        font-size: 1rem;
        color: #6c757d;
        margin-top: 0;
    }

    /* Cards de métricas */
    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        text-align: center;
        border-top: 4px solid;
    }
    .metric-card-verde  { border-top-color: #28a745; }
    .metric-card-amarelo{ border-top-color: #ffc107; }
    .metric-card-azul   { border-top-color: #007bff; }
    .metric-card-vermelho{ border-top-color: #dc3545; }
    .metric-card-cinza  { border-top-color: #6c757d; }
    .metric-numero { font-size: 2.4rem; font-weight: 700; color: #1a1a2e; }
    .metric-label  { font-size: 0.85rem; color: #6c757d; text-transform: uppercase; letter-spacing: 1px; }

    /* Badges de status */
    .badge-conciliado   { background:#d4edda; color:#155724; padding:3px 10px; border-radius:20px; font-size:0.82rem; font-weight:600; }
    .badge-combinado    { background:#fff3cd; color:#856404; padding:3px 10px; border-radius:20px; font-size:0.82rem; font-weight:600; }
    .badge-provavel     { background:#cce5ff; color:#004085; padding:3px 10px; border-radius:20px; font-size:0.82rem; font-weight:600; }
    .badge-nao          { background:#f8d7da; color:#721c24; padding:3px 10px; border-radius:20px; font-size:0.82rem; font-weight:600; }

    /* Separador de seção */
    .secao-header {
        background: linear-gradient(90deg, #1a1a2e 0%, #16213e 100%);
        color: white;
        padding: 10px 18px;
        border-radius: 8px;
        font-weight: 600;
        font-size: 1rem;
        margin: 20px 0 10px 0;
    }

    /* Cabeçalho do lado Razão vs Extrato */
    .header-razao   { background:#fff3cd; border-radius:6px; padding:8px 14px; font-weight:700; color:#856404; }
    .header-extrato { background:#f8d7da; border-radius:6px; padding:8px 14px; font-weight:700; color:#721c24; }

    /* Badge manual */
    .badge-manual { background:#ffe0b2; color:#e65100; padding:3px 10px; border-radius:20px; font-size:0.82rem; font-weight:600; }

    /* Separador central entre Razão e Extrato na tabela */
    .separador-central {
        border-left: 3px solid #dee2e6;
        padding-left: 12px;
        margin-left: 4px;
    }

    /* Detalhe expansível */
    .detalhe-box {
        background: #f0f4ff;
        border-left: 4px solid #007bff;
        border-radius: 6px;
        padding: 12px 18px;
        margin: 4px 0 8px 0;
        font-size: 0.9rem;
        color: #1a1a2e;
    }
    .detalhe-box-manual {
        background: #fff8f0;
        border-left: 4px solid #ff9800;
    }

    /* Linha de seleção manual */
    .manual-row-razao   { background:#fff3e0 !important; }
    .manual-row-extrato { background:#fce4ec !important; }

    /* Oculta índice da tabela */
    [data-testid="stDataFrame"] table { width: 100%; }

    /* Botão fantasma para expandir linha — apenas ícone, sem fundo ou borda */
    div[data-testid="stButton"].btn-expandir > button {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        color: #6c757d !important;
        font-size: 0.85rem !important;
        padding: 2px 6px !important;
        min-height: 0 !important;
        line-height: 1.2 !important;
        width: auto !important;
        cursor: pointer !important;
        transition: color 0.1s, transform 0.1s !important;
    }
    div[data-testid="stButton"].btn-expandir > button:hover {
        background: rgba(0,0,0,0.05) !important;
        color: #1a3a6b !important;
        transform: scale(1.15) !important;
        border-radius: 3px !important;
    }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# FUNÇÕES UTILITÁRIAS
# =============================================================================

def formatar_moeda(valor) -> str:
    """Formata um número como moeda brasileira."""
    if pd.isna(valor) or valor is None:
        return "-"
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(valor)


def formatar_data(data) -> str:
    """Formata uma data para DD/MM/AAAA."""
    if pd.isna(data) or data is None:
        return "-"
    try:
        return pd.Timestamp(data).strftime("%d/%m/%Y")
    except Exception:
        return str(data)


def colorir_linha(row) -> list:
    """
    Retorna o estilo CSS para cada célula de uma linha do DataFrame,
    baseado no status da conciliação.
    
    Cores:
    - CONCILIADO      → Verde claro
    - MATCH_COMBINADO → Amarelo claro
    - MATCH_PROVAVEL  → Azul claro
    - NAO_CONCILIADO  → Vermelho claro
    """
    status = row.get("Status", "")
    cores = {
        "✅ CONCILIADO":      "background-color: #d4edda; color: #155724;",
        "🟡 COMBINADO":       "background-color: #fff8e1; color: #856404;",
        "🔵 SIMILARIDADE":    "background-color: #e3f2fd; color: #004085;",
        "🔴 NÃO CONCILIADO":  "background-color: #fde8e8; color: #721c24;",
        "🟠 MANUAL":          "background-color: #fff3e0; color: #e65100;",
    }
    cor = cores.get(status, "")
    return [cor] * len(row)


def gerar_excel(df: pd.DataFrame) -> bytes:
    """Gera arquivo Excel com formatação de cores para download."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conciliação")

        # Aplica cores nas linhas
        from openpyxl.styles import PatternFill, Font
        ws = writer.sheets["Conciliação"]

        cores_excel = {
            "✅ CONCILIADO":     ("D4EDDA", "155724"),
            "🟡 COMBINADO":      ("FFF8E1", "856404"),
            "🔵 SIMILARIDADE":   ("E3F2FD", "004085"),
            "🔴 NÃO CONCILIADO": ("FDE8E8", "721C24"),
            "🟠 MANUAL":         ("FFF3E0", "E65100"),
        }

        # Cabeçalho em negrito com fundo escuro
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Linhas de dados
        col_status = None
        for i, col in enumerate(df.columns, 1):
            if col == "Status":
                col_status = i
                break

        if col_status:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                status_cell = ws.cell(row=row_idx, column=col_status)
                status_val = str(status_cell.value or "")
                if status_val in cores_excel:
                    bg, fg = cores_excel[status_val]
                    for cell in row:
                        cell.fill = PatternFill(
                            start_color=bg, end_color=bg, fill_type="solid"
                        )
                        cell.font = Font(color=fg)

        # Ajusta largura das colunas
        for col in ws.columns:
            max_width = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_width + 4, 50)

    return output.getvalue()


def preparar_df_exibicao(df_resultado: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara o DataFrame para exibição na tabela do Streamlit.
    Renomeia colunas, formata valores e traduz status.
    """
    STATUS_MAP = {
        "CONCILIADO":        "✅ CONCILIADO",
        "MATCH_COMBINADO":   "🟡 COMBINADO",
        "MATCH_PROVAVEL":    "🔵 SIMILARIDADE",
        "NAO_CONCILIADO":    "🔴 NÃO CONCILIADO",
        "MANUAL_CONCILIADO": "🟠 MANUAL",
    }

    df = df_resultado.copy()

    # Formata colunas
    df["data_razao"]    = df["data_razao"].apply(formatar_data)
    df["data_extrato"]  = df["data_extrato"].apply(formatar_data)
    df["valor_razao"]   = df["valor_razao"].apply(formatar_moeda)
    df["valor_extrato"] = df["valor_extrato"].apply(formatar_moeda)
    df["confidence"]    = df["confidence"].apply(
        lambda x: f"{x:.1f}%" if pd.notna(x) else "-"
    )

    # Traduz status
    df["status"] = df["status"].map(STATUS_MAP).fillna(df["status"])

    # Seleciona e renomeia colunas para exibição
    colunas_exibicao = {
        "data_razao":        "Data Razão",
        "historico_razao":   "Histórico (Razão)",
        "valor_razao":       "Valor Razão",
        "data_extrato":      "Data Extrato",
        "descricao_extrato": "Descrição (Extrato)",
        "valor_extrato":     "Valor Extrato",
        "status":            "Status",
        "tipo_match":        "Tipo Match",
        "confidence":        "Confiança",
        "observacoes":       "Observações",
    }

    colunas_disponiveis = {k: v for k, v in colunas_exibicao.items() if k in df.columns}
    df = df[list(colunas_disponiveis.keys())].rename(columns=colunas_disponiveis)

    return df


# =============================================================================
# SIDEBAR - CONFIGURAÇÕES
# =============================================================================

def renderizar_sidebar():
    """Renderiza a barra lateral com configurações da conciliação."""
    with st.sidebar:
        st.image("https://img.icons8.com/color/96/null/money-transfer.png", width=60)
        st.title("⚙️ Configurações")
        st.divider()

        st.subheader("🎚️ Parâmetros da Engine")

        score_minimo = st.slider(
            "Score mínimo similaridade (%)",
            min_value=60, max_value=95, value=80, step=5,
            help="Mínimo de similaridade de texto para aceitar um match por descrição"
        )

        usar_similaridade = st.checkbox(
            "Ativar match por similaridade",
            value=True,
            help="Desative para conciliação mais conservadora (apenas exato + combinado)"
        )

        st.divider()

        st.subheader("📅 Tolerâncias")
        tolerancia_dias = st.number_input(
            "Tolerância de datas (dias)",
            min_value=0, max_value=10, value=3,
            help="Diferença máxima de dias aceita entre data do Razão e data do Extrato"
        )

        st.divider()

        st.subheader("ℹ️ Legenda de Cores")
        st.markdown("""
        <div style='line-height:2.2;'>
        ✅ <b>Verde</b> → Conciliado (exato)<br>
        🟡 <b>Amarelo</b> → Combinação de lançamentos<br>
        🔵 <b>Azul</b> → Similaridade de texto<br>
        � <b>Laranja</b> → Conciliado manualmente<br>
        � <b>Vermelho</b> → Não conciliado
        </div>
        """, unsafe_allow_html=True)

        st.divider()
        st.caption("Sistema de Conciliação Financeira v1.0")
        st.caption("Desenvolvido com Python + Streamlit")

    return score_minimo, usar_similaridade, tolerancia_dias


# =============================================================================
# MÉTRICAS (KPIs)
# =============================================================================

def renderizar_metricas(metricas: dict):
    """Exibe os cards de métricas no topo da página."""
    c1, c2, c3, c4, c5, c6 = st.columns(6)

    with c1:
        st.markdown(f"""
        <div class="metric-card metric-card-cinza">
            <div class="metric-numero">{metricas.get('total', 0)}</div>
            <div class="metric-label">Total de Registros</div>
        </div>
        """, unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
        <div class="metric-card metric-card-verde">
            <div class="metric-numero">{metricas.get('conciliados', 0)}</div>
            <div class="metric-label">✅ Conciliados</div>
        </div>
        """, unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
        <div class="metric-card metric-card-amarelo">
            <div class="metric-numero">{metricas.get('combinados', 0)}</div>
            <div class="metric-label">🟡 Match Combinado</div>
        </div>
        """, unsafe_allow_html=True)

    with c4:
        st.markdown(f"""
        <div class="metric-card metric-card-azul">
            <div class="metric-numero">{metricas.get('provaveis', 0)}</div>
            <div class="metric-label">🔵 Match Similaridade</div>
        </div>
        """, unsafe_allow_html=True)

    with c5:
        st.markdown(f"""
        <div class="metric-card" style="border-top-color:#ff9800;background:white;
             border-radius:12px;padding:20px 24px;box-shadow:0 2px 8px rgba(0,0,0,0.08);text-align:center;">
            <div class="metric-numero">{metricas.get('manuais', 0)}</div>
            <div class="metric-label">🟠 Manual</div>
        </div>
        """, unsafe_allow_html=True)

    with c6:
        st.markdown(f"""
        <div class="metric-card metric-card-vermelho">
            <div class="metric-numero">{metricas.get('nao_conciliados', 0)}</div>
            <div class="metric-label">🔴 Não Conciliados</div>
        </div>
        """, unsafe_allow_html=True)

    # Barra de progresso de conciliação
    st.markdown("<br>", unsafe_allow_html=True)
    pct = metricas['pct_conciliacao']
    st.markdown(f"**Taxa de Conciliação: {pct:.1f}%**")
    st.progress(pct / 100)


# =============================================================================
# TABELA PRINCIPAL DE RESULTADOS — DOIS LADOS COM DETALHE EXPANSÍVEL
# =============================================================================

def _cor_status(status_raw: str) -> str:
    """Retorna o estilo CSS de fundo para um status bruto."""
    mapa = {
        "MATCH_APRENDIDO":   "background:#f3e5f5; color:#4a148c;",
        "CONCILIADO":        "background:#d4edda; color:#155724;",
        "MATCH_COMBINADO":   "background:#fff8e1; color:#856404;",
        "MATCH_PROVAVEL":    "background:#e3f2fd; color:#004085;",
        "NAO_CONCILIADO":    "background:#fde8e8; color:#721c24;",
        "MANUAL_CONCILIADO": "background:#fff3e0; color:#e65100;",
    }
    return mapa.get(status_raw, "")


def _badge_status(status_raw: str) -> str:
    """Retorna HTML de célula-badge colorida para uso dentro do grid CSS."""
    mapa = {
        "MATCH_APRENDIDO":   ("🧠 APRENDIDO",      "#f3e5f5", "#4a148c"),
        "CONCILIADO":        ("✅ CONCILIADO",     "#d4edda", "#155724"),
        "MATCH_COMBINADO":   ("🟡 COMBINADO",      "#fff3cd", "#856404"),
        "MATCH_PROVAVEL":    ("🔵 SIMILARIDADE",   "#cce5ff", "#004085"),
        "NAO_CONCILIADO":    ("🔴 NÃO CONCILIADO", "#f8d7da", "#721c24"),
        "MANUAL_CONCILIADO": ("🟠 MANUAL",          "#ffe0b2", "#e65100"),
    }
    label, bg, fg = mapa.get(status_raw, (status_raw, "#eee", "#333"))
    return (
        f"<div style='background:{bg};color:{fg};"
        f"padding:3px 8px;border-radius:4px;"
        f"font-size:0.78rem;font-weight:700;"
        f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
        f"display:flex;align-items:center'>{label}</div>"
    )


def _celula(texto: str, bg: str, fg: str, negrito: bool = False, max_chars: int = 28) -> str:
    """
    Retorna HTML de célula compacta com texto truncado e tooltip completo no hover.
    bg/fg são cores de fundo e texto da célula.
    """
    texto_str = str(texto) if texto else "—"
    truncado  = (texto_str[:max_chars] + "…") if len(texto_str) > max_chars else texto_str
    peso      = "font-weight:700;" if negrito else ""
    # Escapa aspas simples no título do tooltip
    tooltip   = texto_str.replace("'", "&#39;")
    return (
        f"<div title='{tooltip}' style='"
        f"background:{bg};color:{fg};{peso}"
        f"padding:3px 7px;border-radius:4px;"
        f"font-size:0.82rem;line-height:1.35;"
        f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
        f"cursor:default'>{truncado}</div>"
    )


def renderizar_tabela_principal(df_resultado: pd.DataFrame, filtro_status: list):
    """
    Tabela principal: RAZÃO (azul) | separador | EXTRATO (laranja).
    Linhas compactas, texto truncado com tooltip.
    Clique no badge de status expande o detalhe inline abaixo da linha.
    """
    st.markdown('<div class="secao-header">📋 Resultado da Conciliação</div>',
                unsafe_allow_html=True)

    # Paleta fixa dos dois lados
    COR_R_BG  = "#e8f0fe"   # azul claro — fundo Razão
    COR_R_FG  = "#1a3a6b"   # azul escuro — texto Razão
    COR_R_CAB = "#1a3a6b"   # cabeçalho Razão
    COR_E_BG  = "#fff3e0"   # laranja claro — fundo Extrato
    COR_E_FG  = "#7c3c00"   # laranja escuro — texto Extrato
    COR_E_CAB = "#b84c00"   # cabeçalho Extrato

    # Aplica filtro de status
    if filtro_status and "Todos" not in filtro_status:
        df_filtrado = df_resultado[df_resultado["status"].isin(filtro_status)].copy()
    else:
        df_filtrado = df_resultado.copy()

    if df_filtrado.empty:
        st.info("Nenhum registro encontrado com os filtros aplicados.")
        return

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style='display:grid;
        grid-template-columns:1.7fr 3fr 1.4fr 4px 1.7fr 3fr 1.4fr 1.7fr;
        gap:4px;margin-bottom:2px;'>
      <div style='background:{COR_R_CAB};color:white;padding:5px 8px;
           border-radius:5px 0 0 5px;font-size:0.78rem;font-weight:700'>
           📅 Data Razão</div>
      <div style='background:{COR_R_CAB};color:white;padding:5px 8px;
           font-size:0.78rem;font-weight:700'>📒 Histórico (Razão)</div>
      <div style='background:{COR_R_CAB};color:white;padding:5px 8px;
           font-size:0.78rem;font-weight:700'>💰 Valor Razão</div>
      <div style='background:#dee2e6;border-radius:2px'></div>
      <div style='background:{COR_E_CAB};color:white;padding:5px 8px;
           font-size:0.78rem;font-weight:700'>📅 Data Extrato</div>
      <div style='background:{COR_E_CAB};color:white;padding:5px 8px;
           font-size:0.78rem;font-weight:700'>🏦 Descrição (Extrato)</div>
      <div style='background:{COR_E_CAB};color:white;padding:5px 8px;
           font-size:0.78rem;font-weight:700'>💰 Valor Extrato</div>
      <div style='background:#37474f;color:white;padding:5px 8px;
           border-radius:0 5px 5px 0;font-size:0.78rem;font-weight:700'>
           📊 Status</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Injeção de CSS por key para estilizar o botão como fantasma ───────────
    # O seletor usa o atributo aria-label que o Streamlit coloca no botão
    st.markdown("""
    <style>
    /* Remove toda aparência dos botões de expandir linha */
    button[kind="secondary"][data-testid="baseButton-secondary"] {
        /* fallback — só afeta se não tiver classe específica */
    }
    /* Seletor via aria-label: funciona porque o Streamlit usa o label como aria-label */
    button[aria-label="▼"], button[aria-label="▲"] {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        color: #9e9e9e !important;
        font-size: 0.9rem !important;
        padding: 1px 5px !important;
        min-height: 0 !important;
        height: 26px !important;
        line-height: 1 !important;
        cursor: pointer !important;
        border-radius: 3px !important;
        transition: background 0.1s, color 0.1s !important;
    }
    button[aria-label="▼"]:hover, button[aria-label="▲"]:hover {
        background: rgba(26,58,107,0.1) !important;
        color: #1a3a6b !important;
    }
    button[aria-label="▲"] {
        color: #1a3a6b !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Linhas de dados ───────────────────────────────────────────────────────
    df_filtrado = df_filtrado.reset_index(drop=True)
    for i, row in df_filtrado.iterrows():
        status_raw = str(row.get("status", ""))
        chave_exp  = f"det_{i}_{status_raw}"
        expandido  = st.session_state.get(chave_exp, False)

        hist_r  = str(row.get("historico_razao",  "") or "—")
        desc_e  = str(row.get("descricao_extrato","") or "—")
        data_r  = formatar_data(row.get("data_razao"))
        data_e  = formatar_data(row.get("data_extrato"))
        val_r   = formatar_moeda(row.get("valor_razao"))
        val_e   = formatar_moeda(row.get("valor_extrato"))

        borda_exp = "outline:2px solid #1a3a6b;outline-offset:-2px;" if expandido else ""

        # ── Linha: colunas Streamlit com HTML dentro de cada célula ───────────
        # A última coluna contém o botão ▼/▲ (fantasma via CSS aria-label)
        c1, c2, c3, csep, c4, c5, c6, c7, c8 = st.columns(
            [1.7, 3.0, 1.4, 0.05, 1.7, 3.0, 1.4, 1.6, 0.45]
        )
        # Container externo com borda quando expandido
        st.markdown(
            f"<style>#row_wrap_{i}{{outline:{('2px solid #1a3a6b' if expandido else 'none')};"
            f"outline-offset:-2px;border-radius:4px}}</style>"
            f"<div id='row_wrap_{i}'></div>",
            unsafe_allow_html=True,
        )

        with c1:
            st.markdown(
                f"<div title='{data_r}' style='background:{COR_R_BG};color:{COR_R_FG};"
                f"padding:3px 6px;border-radius:4px;font-size:0.82rem;"
                f"white-space:nowrap;margin-bottom:1px'>{data_r}</div>",
                unsafe_allow_html=True)
        with c2:
            trunc = (hist_r[:30] + "…") if len(hist_r) > 30 else hist_r
            st.markdown(
                f"<div title='{hist_r.replace(chr(39), '')}' style='background:{COR_R_BG};"
                f"color:{COR_R_FG};padding:3px 6px;border-radius:4px;font-size:0.82rem;"
                f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
                f"margin-bottom:1px'>{trunc}</div>",
                unsafe_allow_html=True)
        with c3:
            st.markdown(
                f"<div style='background:{COR_R_BG};color:{COR_R_FG};font-weight:700;"
                f"padding:3px 6px;border-radius:4px;font-size:0.82rem;"
                f"white-space:nowrap;margin-bottom:1px'>{val_r}</div>",
                unsafe_allow_html=True)
        with csep:
            st.markdown(
                "<div style='background:#dee2e6;height:26px;border-radius:2px;"
                "margin-bottom:1px'></div>",
                unsafe_allow_html=True)
        with c4:
            st.markdown(
                f"<div title='{data_e}' style='background:{COR_E_BG};color:{COR_E_FG};"
                f"padding:3px 6px;border-radius:4px;font-size:0.82rem;"
                f"white-space:nowrap;margin-bottom:1px'>{data_e}</div>",
                unsafe_allow_html=True)
        with c5:
            trunc_e = (desc_e[:30] + "…") if len(desc_e) > 30 else desc_e
            st.markdown(
                f"<div title='{desc_e.replace(chr(39), '')}' style='background:{COR_E_BG};"
                f"color:{COR_E_FG};padding:3px 6px;border-radius:4px;font-size:0.82rem;"
                f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
                f"margin-bottom:1px'>{trunc_e}</div>",
                unsafe_allow_html=True)
        with c6:
            st.markdown(
                f"<div style='background:{COR_E_BG};color:{COR_E_FG};font-weight:700;"
                f"padding:3px 6px;border-radius:4px;font-size:0.82rem;"
                f"white-space:nowrap;margin-bottom:1px'>{val_e}</div>",
                unsafe_allow_html=True)
        with c7:
            st.markdown(_badge_status(status_raw), unsafe_allow_html=True)
        with c8:
            # Botão fantasma — aria-label é ▼ ou ▲, CSS remove toda aparência
            icone = "▲" if expandido else "▼"
            if st.button(icone, key=f"btn_{i}", help="Ver detalhes da conciliação"):
                st.session_state[chave_exp] = not expandido
                st.rerun()

        # ── Detalhe expansível ────────────────────────────────────────────────
        if expandido:
            obs   = str(row.get("observacoes", "") or "—")
            tipo  = str(row.get("tipo_match",  "") or "—")
            conf  = float(row.get("confidence", 0) or 0)
            doc_r = str(row.get("documento_razao",   "") or "—")
            doc_e = str(row.get("documento_extrato", "") or "—")
            conta = str(row.get("conta_razao", "") or "—")

            borda_cor = "#ff9800" if status_raw == "MANUAL_CONCILIADO" else "#1a3a6b"
            bg_det    = "#fff8f0" if status_raw == "MANUAL_CONCILIADO" else "#f0f4ff"

            st.markdown(f"""
            <div style='border-left:4px solid {borda_cor};background:{bg_det};
                 border-radius:0 6px 6px 0;padding:10px 18px;
                 margin:0 0 6px 0;font-size:0.85rem;color:#1a1a2e;line-height:1.9'>
              <b>🔍 Como foi conciliado</b><br>
              &nbsp;• <b>Tipo de Match:</b> {tipo} &nbsp;|&nbsp;
                <b>Confiança:</b> {conf:.1f}%<br>
              &nbsp;• <b>Documento Razão:</b> {doc_r} &nbsp;|&nbsp;
                <b>Documento Extrato:</b> {doc_e}<br>
              &nbsp;• <b>Conta:</b> {conta}<br>
              &nbsp;• <b>Observações:</b> {obs}
            </div>""", unsafe_allow_html=True)

    st.caption(f"Exibindo {len(df_filtrado)} de {len(df_resultado)} registros")


# =============================================================================
# SEÇÃO NÃO CONCILIADOS — CONCILIAÇÃO MANUAL COM CHECKBOXES
# =============================================================================

def renderizar_nao_conciliados(df_resultado: pd.DataFrame):
    """
    Exibe os lançamentos não conciliados em duas colunas com checkboxes.

    O usuário pode:
    - Marcar 1 ou mais lançamentos do RAZÃO (esquerda)
    - Marcar 1 ou mais lançamentos do EXTRATO (direita)
    - Clicar em "✅ Confirmar Conciliação Manual"
    - Os itens marcados são unidos e marcados como MANUAL_CONCILIADO (laranja)

    Também mostra o totalizador e a diferença entre os dois lados.
    """
    df_nc = df_resultado[df_resultado["status"] == "NAO_CONCILIADO"].copy()

    if df_nc.empty:
        st.success("🎉 Todos os lançamentos foram conciliados!")
        return

    st.markdown(
        '<div class="secao-header">🔴 Lançamentos Não Conciliados — Revisão Manual</div>',
        unsafe_allow_html=True
    )

    # Instrução de uso
    st.info(
        "✏️ **Como fazer conciliação manual:** "
        "Marque os checkboxes dos lançamentos que se correspondem "
        "(pode ser 1 do Razão + 1 do Extrato, ou múltiplos de cada lado), "
        "depois clique em **✅ Confirmar Conciliação Manual**."
    )

    # Separa os não conciliados por origem
    if "origem" in df_nc.columns:
        nc_razao   = df_nc[df_nc["origem"] == "RAZAO"].copy()
        nc_extrato = df_nc[df_nc["origem"] == "EXTRATO"].copy()
    else:
        nc_razao   = df_nc[df_nc["valor_razao"].notna()].copy()
        nc_extrato = df_nc[df_nc["valor_extrato"].notna()].copy()

    # Inicializa listas de seleção no session_state
    if "manual_sel_razao" not in st.session_state:
        st.session_state["manual_sel_razao"]   = []
    if "manual_sel_extrato" not in st.session_state:
        st.session_state["manual_sel_extrato"] = []

    col_razao, col_sep, col_extrato = st.columns([10, 0.3, 10])

    # ── Coluna RAZÃO ──────────────────────────────────────────────────────────
    with col_razao:
        st.markdown(
            f'<div class="header-razao">📒 RAZÃO CONTÁBIL — {len(nc_razao)} sem par</div>',
            unsafe_allow_html=True
        )
        if nc_razao.empty:
            st.success("Todos os lançamentos do Razão foram conciliados!")
        else:
            st.markdown("<br>", unsafe_allow_html=True)
            sel_razao = []
            for idx, row in nc_razao.iterrows():
                c_chk, c_data, c_hist, c_val = st.columns([0.5, 1.5, 4, 1.8])
                chave = f"chk_r_{idx}"
                marcado = c_chk.checkbox(
                    "", key=chave,
                    value=(idx in st.session_state["manual_sel_razao"])
                )
                if marcado:
                    sel_razao.append(idx)
                c_data.markdown(
                    f"<div style='background:#fff8e1;padding:4px 6px;border-radius:4px;"
                    f"font-size:0.85rem'>{formatar_data(row.get('data_razao'))}</div>",
                    unsafe_allow_html=True
                )
                c_hist.markdown(
                    f"<div style='background:#fff8e1;padding:4px 6px;border-radius:4px;"
                    f"font-size:0.85rem'>{row.get('historico_razao','') or '—'}</div>",
                    unsafe_allow_html=True
                )
                c_val.markdown(
                    f"<div style='background:#fff8e1;padding:4px 6px;border-radius:4px;"
                    f"font-size:0.85rem;font-weight:700'>{formatar_moeda(row.get('valor_razao'))}</div>",
                    unsafe_allow_html=True
                )
            st.session_state["manual_sel_razao"] = sel_razao

            total_razao = nc_razao["valor_razao"].dropna().sum()
            st.markdown(
                f"<b>Total sem conciliar: {formatar_moeda(total_razao)}</b>",
                unsafe_allow_html=True
            )
            if sel_razao:
                soma_sel_r = nc_razao.loc[
                    nc_razao.index.isin(sel_razao), "valor_razao"
                ].dropna().sum()
                st.markdown(
                    f"<span style='color:#856404'>☑️ Selecionados: "
                    f"<b>{len(sel_razao)}</b> lançamentos → "
                    f"<b>{formatar_moeda(soma_sel_r)}</b></span>",
                    unsafe_allow_html=True
                )

    # ── Separador visual central ───────────────────────────────────────────────
    with col_sep:
        st.markdown(
            "<div style='border-left:2px solid #dee2e6;min-height:300px;margin-top:40px'></div>",
            unsafe_allow_html=True
        )

    # ── Coluna EXTRATO ────────────────────────────────────────────────────────
    with col_extrato:
        st.markdown(
            f'<div class="header-extrato">🏦 EXTRATO BANCÁRIO — {len(nc_extrato)} sem par</div>',
            unsafe_allow_html=True
        )
        if nc_extrato.empty:
            st.success("Todos os lançamentos do Extrato foram conciliados!")
        else:
            st.markdown("<br>", unsafe_allow_html=True)
            sel_extrato = []
            for idx, row in nc_extrato.iterrows():
                c_chk, c_data, c_desc, c_val = st.columns([0.5, 1.5, 4, 1.8])
                chave = f"chk_e_{idx}"
                marcado = c_chk.checkbox(
                    "", key=chave,
                    value=(idx in st.session_state["manual_sel_extrato"])
                )
                if marcado:
                    sel_extrato.append(idx)
                c_data.markdown(
                    f"<div style='background:#fde8e8;padding:4px 6px;border-radius:4px;"
                    f"font-size:0.85rem'>{formatar_data(row.get('data_extrato'))}</div>",
                    unsafe_allow_html=True
                )
                c_desc.markdown(
                    f"<div style='background:#fde8e8;padding:4px 6px;border-radius:4px;"
                    f"font-size:0.85rem'>{row.get('descricao_extrato','') or '—'}</div>",
                    unsafe_allow_html=True
                )
                c_val.markdown(
                    f"<div style='background:#fde8e8;padding:4px 6px;border-radius:4px;"
                    f"font-size:0.85rem;font-weight:700'>{formatar_moeda(row.get('valor_extrato'))}</div>",
                    unsafe_allow_html=True
                )
            st.session_state["manual_sel_extrato"] = sel_extrato

            total_extrato = nc_extrato["valor_extrato"].dropna().sum()
            st.markdown(
                f"<b>Total sem conciliar: {formatar_moeda(total_extrato)}</b>",
                unsafe_allow_html=True
            )
            if sel_extrato:
                soma_sel_e = nc_extrato.loc[
                    nc_extrato.index.isin(sel_extrato), "valor_extrato"
                ].dropna().sum()
                st.markdown(
                    f"<span style='color:#721c24'>☑️ Selecionados: "
                    f"<b>{len(sel_extrato)}</b> lançamentos → "
                    f"<b>{formatar_moeda(soma_sel_e)}</b></span>",
                    unsafe_allow_html=True
                )

    # ── Totalizador geral e diferença ──────────────────────────────────────────
    st.divider()
    total_r = nc_razao["valor_razao"].dropna().sum()   if not nc_razao.empty   else 0.0
    total_e = nc_extrato["valor_extrato"].dropna().sum() if not nc_extrato.empty else 0.0
    diferenca = total_r - total_e
    cor_dif = "🟢" if abs(diferenca) < 0.01 else "🔴"
    st.markdown(
        f"**{cor_dif} Diferença total entre Razão e Extrato não conciliados: "
        f"{formatar_moeda(diferenca)}**"
    )

    # ── Botão de confirmar conciliação manual ─────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    sel_r = st.session_state.get("manual_sel_razao", [])
    sel_e = st.session_state.get("manual_sel_extrato", [])

    btn_disabled = (len(sel_r) == 0 and len(sel_e) == 0)

    col_btn, col_aviso = st.columns([3, 9])
    with col_btn:
        confirmar = st.button(
            "✅ Confirmar Conciliação Manual",
            type="primary",
            use_container_width=True,
            disabled=btn_disabled,
        )
    with col_aviso:
        if btn_disabled:
            st.warning("Marque pelo menos um lançamento em cada lado para conciliar.")
        elif len(sel_r) == 0:
            st.warning("⚠️ Nenhum lançamento do Razão selecionado.")
        elif len(sel_e) == 0:
            st.warning("⚠️ Nenhum lançamento do Extrato selecionado.")
        else:
            soma_r = nc_razao.loc[nc_razao.index.isin(sel_r), "valor_razao"].dropna().sum()
            soma_e = nc_extrato.loc[nc_extrato.index.isin(sel_e), "valor_extrato"].dropna().sum()
            dif = soma_r - soma_e
            if abs(dif) > 0.01:
                st.warning(
                    f"⚠️ Valores não batem: Razão {formatar_moeda(soma_r)} vs "
                    f"Extrato {formatar_moeda(soma_e)} "
                    f"(diferença {formatar_moeda(dif)}). "
                    f"Você ainda pode confirmar se for intencional."
                )
            else:
                st.success(
                    f"✔️ {len(sel_r)} Razão + {len(sel_e)} Extrato selecionados "
                    f"| Valores batem: {formatar_moeda(soma_r)}"
                )

    # ── Processa a conciliação manual quando confirmado ────────────────────────
    if confirmar and (sel_r or sel_e):
        df_atual = st.session_state["df_resultado"].copy()

        # Monta descrições resumidas para a observação
        descs_r = []
        for idx in sel_r:
            row = df_atual.loc[idx]
            descs_r.append(
                f"{formatar_data(row.get('data_razao'))} "
                f"{str(row.get('historico_razao',''))[:30]} "
                f"{formatar_moeda(row.get('valor_razao'))}"
            )
        descs_e = []
        for idx in sel_e:
            row = df_atual.loc[idx]
            descs_e.append(
                f"{formatar_data(row.get('data_extrato'))} "
                f"{str(row.get('descricao_extrato',''))[:30]} "
                f"{formatar_moeda(row.get('valor_extrato'))}"
            )

        obs_manual = (
            f"MANUAL: Razão [{' | '.join(descs_r)}] "
            f"↔ Extrato [{' | '.join(descs_e)}]"
        )

        # Quando há múltiplos de um lado, o primeiro recebe o par completo
        # e os demais ficam como complemento (mesmo grupo manual)
        primeiro_r = sel_r[0]   if sel_r   else None
        primeiro_e = sel_e[0]   if sel_e   else None

        # Atualiza todos os lançamentos do Razão selecionados
        for i_r, idx in enumerate(sel_r):
            df_atual.at[idx, "status"]      = "MANUAL_CONCILIADO"
            df_atual.at[idx, "tipo_match"]  = "MANUAL"
            df_atual.at[idx, "confidence"]  = 100.0
            df_atual.at[idx, "observacoes"] = obs_manual
            # Para os extras, aponta para o par do extrato do primeiro
            if primeiro_e is not None:
                df_atual.at[idx, "data_extrato"]      = df_atual.at[primeiro_e, "data_extrato"]
                df_atual.at[idx, "descricao_extrato"] = df_atual.at[primeiro_e, "descricao_extrato"]
                df_atual.at[idx, "valor_extrato"]     = df_atual.at[primeiro_e, "valor_extrato"]
                df_atual.at[idx, "documento_extrato"] = df_atual.at[primeiro_e, "documento_extrato"] \
                    if "documento_extrato" in df_atual.columns else ""

        # Atualiza todos os lançamentos do Extrato selecionados
        for i_e, idx in enumerate(sel_e):
            df_atual.at[idx, "status"]      = "MANUAL_CONCILIADO"
            df_atual.at[idx, "tipo_match"]  = "MANUAL"
            df_atual.at[idx, "confidence"]  = 100.0
            df_atual.at[idx, "observacoes"] = obs_manual
            if primeiro_r is not None:
                df_atual.at[idx, "data_razao"]      = df_atual.at[primeiro_r, "data_razao"]
                df_atual.at[idx, "historico_razao"] = df_atual.at[primeiro_r, "historico_razao"]
                df_atual.at[idx, "valor_razao"]     = df_atual.at[primeiro_r, "valor_razao"]
                df_atual.at[idx, "documento_razao"] = df_atual.at[primeiro_r, "documento_razao"] \
                    if "documento_razao" in df_atual.columns else ""

        # Persiste e limpa seleção
        st.session_state["df_resultado"]     = df_atual
        st.session_state["manual_sel_razao"]   = []
        st.session_state["manual_sel_extrato"] = []

        # Salva padrões manuais no banco para aprendizado ML
        try:
            from database.db_manager import DatabaseManager
            db = DatabaseManager()
            # Registra cada par razão↔extrato selecionado manualmente
            for idx_r in sel_r:
                hist = str(df_atual.at[idx_r, "historico_razao"] or "").strip()
                for idx_e in sel_e:
                    desc = str(df_atual.at[idx_e, "descricao_extrato"] or "").strip()
                    if hist and desc:
                        db.registrar_padrao(
                            historico_razao   = hist,
                            descricao_extrato = desc,
                            tipo_match        = "MANUAL",
                            confirmado_por    = "MANUAL",
                            data_ocorrencia   = df_atual.at[idx_r, "data_razao"],
                        )
        except Exception:
            pass  # Não bloqueia o fluxo se o banco falhar

        st.success(
            f"🟠 Conciliação manual aplicada! "
            f"{len(sel_r)} lançamento(s) do Razão + {len(sel_e)} do Extrato "
            f"marcados como **MANUAL_CONCILIADO**."
        )
        st.rerun()


# =============================================================================
# PÁGINA PRINCIPAL
# =============================================================================

def main():
    """Função principal que renderiza toda a interface do dashboard."""

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    st.markdown(
        '<p class="titulo-principal">💰 Conciliação Financeira Automatizada</p>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<p class="subtitulo">Sistema de conciliação inteligente com múltiplos níveis de match</p>',
        unsafe_allow_html=True
    )
    st.divider()

    # ── Sidebar com configurações ─────────────────────────────────────────────
    score_minimo, usar_similaridade, tolerancia_dias = renderizar_sidebar()

    # ── Upload de Arquivos ────────────────────────────────────────────────────
    st.markdown('<div class="secao-header">📁 Importar Arquivos</div>',
                unsafe_allow_html=True)

    col_upload1, col_upload2 = st.columns(2)

    with col_upload1:
        st.markdown("**📒 Razão Contábil** (1 ou mais arquivos Excel/CSV exportados do Domínio)")
        arquivos_razao = st.file_uploader(
            "Arquivos do Razão",
            type=["xlsx", "xls", "csv"],
            key="razao",
            accept_multiple_files=True,
            label_visibility="collapsed"
        )
        if arquivos_razao:
            for f in arquivos_razao:
                st.success(f"✅ {f.name} ({f.size / 1024:.1f} KB)")
        conta_filtro = st.text_input(
            "🔎 Filtrar por conta (opcional)",
            placeholder="Ex: 1.1.10.200.01 ou ITAU ou deixe vazio para todas",
            help="Filtra os lançamentos do Razão por código ou nome da conta contábil."
        )

    with col_upload2:
        st.markdown("**🏦 Extrato Bancário** (1 ou mais arquivos Excel, CSV ou PDF — ex: 12 meses)")
        arquivos_extrato = st.file_uploader(
            "Arquivos do Extrato",
            type=["xlsx", "xls", "csv", "pdf"],
            key="extrato",
            accept_multiple_files=True,
            label_visibility="collapsed"
        )
        if arquivos_extrato:
            for f in arquivos_extrato:
                st.success(f"✅ {f.name} ({f.size / 1024:.1f} KB)")

    # ── Botão de Execução ─────────────────────────────────────────────────────
    tem_razao   = len(arquivos_razao)   > 0 if arquivos_razao   else False
    tem_extrato = len(arquivos_extrato) > 0 if arquivos_extrato else False

    st.markdown("<br>", unsafe_allow_html=True)
    col_btn, col_info = st.columns([2, 8])

    with col_btn:
        executar = st.button(
            "🚀 Executar Conciliação",
            type="primary",
            use_container_width=True,
            disabled=(not tem_razao or not tem_extrato)
        )

    with col_info:
        if not tem_razao or not tem_extrato:
            st.info("⬆️ Faça upload de pelo menos 1 arquivo de cada lado para habilitar a conciliação.")
        else:
            st.info(
                f"📒 {len(arquivos_razao)} arquivo(s) de Razão  ·  "
                f"🏦 {len(arquivos_extrato)} arquivo(s) de Extrato prontos para conciliar."
            )

    # ── Execução da Conciliação ───────────────────────────────────────────────
    if executar and tem_razao and tem_extrato:
        st.markdown('<div class="secao-header">⚙️ Processando...</div>',
                    unsafe_allow_html=True)

        barra = st.progress(0, text="Iniciando...")

        try:
            import tempfile

            # ── Etapa 1: Lendo todos os arquivos do Razão ────────────────────
            barra.progress(5, text=f"📒 Lendo {len(arquivos_razao)} arquivo(s) do Razão...")
            filtro = conta_filtro.strip() if conta_filtro.strip() else None
            dfs_razao  = []
            meta_razao = {}
            erros_razao = []

            for i, arq in enumerate(arquivos_razao):
                try:
                    tmp = os.path.join(tempfile.gettempdir(), arq.name)
                    with open(tmp, "wb") as f:
                        f.write(arq.read())
                    parser = RazaoParser(conta_filtro=filtro)
                    df_r = parser.carregar(tmp)
                    os.unlink(tmp)
                    if not meta_razao:
                        meta_razao = parser.obter_metadados()
                    dfs_razao.append(df_r)
                    barra.progress(
                        5 + int(20 * (i + 1) / len(arquivos_razao)),
                        text=f"📒 {arq.name}: {len(df_r)} lançamentos"
                    )
                except Exception as e:
                    erros_razao.append(f"{arq.name}: {e}")

            if erros_razao:
                for err in erros_razao:
                    st.warning(f"⚠️ Razão — {err}")

            if not dfs_razao:
                raise ValueError("Nenhum arquivo do Razão foi carregado com sucesso.")

            df_razao = pd.concat(dfs_razao, ignore_index=True).drop_duplicates()
            df_razao["used"] = False
            barra.progress(25, text=f"📒 Razão: {len(df_razao)} lançamentos no total ({len(arquivos_razao)} arquivo(s))")

            # ── Etapa 2: Lendo todos os arquivos do Extrato ──────────────────
            barra.progress(30, text=f"🏦 Lendo {len(arquivos_extrato)} arquivo(s) do Extrato...")
            dfs_extrato  = []
            nomes_extrato = []
            erros_extrato = []

            for i, arq in enumerate(arquivos_extrato):
                try:
                    tmp = os.path.join(tempfile.gettempdir(), arq.name)
                    with open(tmp, "wb") as f:
                        f.write(arq.read())
                    parser_e = ExtratoParser()
                    df_e = parser_e.carregar(tmp)
                    os.unlink(tmp)
                    dfs_extrato.append(df_e)
                    nomes_extrato.append(arq.name)
                    barra.progress(
                        30 + int(25 * (i + 1) / len(arquivos_extrato)),
                        text=f"🏦 {arq.name}: {len(df_e)} lançamentos"
                    )
                except Exception as e:
                    erros_extrato.append(f"{arq.name}: {e}")

            if erros_extrato:
                for err in erros_extrato:
                    st.warning(f"⚠️ Extrato — {err}")

            if not dfs_extrato:
                raise ValueError("Nenhum arquivo do Extrato foi carregado com sucesso.")

            df_extrato = pd.concat(dfs_extrato, ignore_index=True).drop_duplicates()
            df_extrato["used"] = False
            barra.progress(57, text=f"🏦 Extrato: {len(df_extrato)} lançamentos no total ({len(arquivos_extrato)} arquivo(s))")

            # ── Etapa 3: Banco + ML antes da conciliação ─────────────────────
            barra.progress(60, text="💾 Conectando banco de dados...")
            try:
                import sys as _sys
                _sys.path.insert(0, str(Path(__file__).parent.parent))
                from database.db_manager import DatabaseManager
                db = DatabaseManager()
                db.importar_razao(
                    df_razao,
                    arquivo=" | ".join(a.name for a in arquivos_razao),
                    empresa=meta_razao.get("empresa", ""),
                    cnpj=meta_razao.get("cnpj", ""),
                )
                for nome, df_e in zip(nomes_extrato, dfs_extrato):
                    db.importar_extrato(df_e, arquivo=nome)
            except Exception as db_err:
                db = None
                st.warning(f"⚠️ Banco de dados: {db_err}")

            # Etapa 4: Conciliação (com aprendizado ML se banco disponível)
            total_padroes = db.total_padroes() if db else 0
            msg_ml = f" | 🧠 {total_padroes} padrões aprendidos" if total_padroes >= 5 else " | 🧠 aguardando histórico (mín. 5)"
            barra.progress(70, text=f"� Executando conciliação{msg_ml}...")
            engine = ConciliacaoEngine(
                score_minimo_similaridade=score_minimo,
                usar_similaridade=usar_similaridade,
                db_manager=db,
            )
            df_resultado = engine.conciliar(df_razao, df_extrato)
            barra.progress(93, text="✅ Conciliação finalizada!")

            # Salva resultado no estado da sessão
            st.session_state["df_resultado"] = df_resultado
            st.session_state["executado"] = True

            # Salva padrões aprendidos no banco
            barra.progress(96, text="🧠 Salvando padrões aprendidos...")
            if db is not None:
                try:
                    db.registrar_padroes_batch(df_resultado, confirmado_por="AUTO")
                    st.session_state["db_stats"] = db.estatisticas()
                except Exception as ml_err:
                    st.warning(f"⚠️ Aprendizado: {ml_err}")

            barra.progress(100, text="✅ Concluído!")
            st.success(f"🎉 Conciliação realizada! {len(df_resultado)} registros processados.")

        except Exception as e:
            barra.empty()
            st.error(f"❌ Erro durante a conciliação: {str(e)}")
            st.exception(e)
            return

    # ── Exibição dos Resultados ───────────────────────────────────────────────
    if st.session_state.get("executado") and "df_resultado" in st.session_state:
        df_resultado = st.session_state["df_resultado"]

        if df_resultado is None or df_resultado.empty:
            st.warning("⚠️ Nenhum resultado gerado. Verifique os arquivos enviados.")
            return

        # Inicializa engine para métricas
        engine_metricas = ConciliacaoEngine()
        metricas = engine_metricas.obter_metricas(df_resultado)

        # ── KPIs ──────────────────────────────────────────────────────────────
        st.markdown('<div class="secao-header">📊 Métricas da Conciliação</div>',
                    unsafe_allow_html=True)
        renderizar_metricas(metricas)
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Filtros ───────────────────────────────────────────────────────────
        col_f1, col_f2 = st.columns([3, 7])
        with col_f1:
            filtro_status = st.multiselect(
                "🔍 Filtrar por Status",
                options=["Todos", "MATCH_APRENDIDO", "CONCILIADO", "MATCH_COMBINADO",
                         "MATCH_PROVAVEL", "MANUAL_CONCILIADO", "NAO_CONCILIADO"],
                default=["Todos"],
                label_visibility="visible"
            )

        # ── Tabela Principal ──────────────────────────────────────────────────
        renderizar_tabela_principal(df_resultado, filtro_status)

        # ── Não Conciliados ───────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        renderizar_nao_conciliados(df_resultado)

        # ── Download do Resultado ─────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="secao-header">💾 Exportar Resultados</div>',
                    unsafe_allow_html=True)

        col_d1, col_d2, col_d3 = st.columns([2, 2, 6])

        df_exibicao_export = preparar_df_exibicao(df_resultado)

        with col_d1:
            excel_bytes = gerar_excel(df_exibicao_export)
            st.download_button(
                label="📥 Baixar Excel",
                data=excel_bytes,
                file_name=f"conciliacao_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col_d2:
            csv_bytes = df_exibicao_export.to_csv(index=False, sep=";").encode("utf-8-sig")
            st.download_button(
                label="📄 Baixar CSV",
                data=csv_bytes,
                file_name=f"conciliacao_{datetime.date.today().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # ── Botão Limpar ──────────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Limpar e Iniciar Nova Conciliação"):
            for key in ["df_resultado", "executado"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()




# =============================================================================
# INICIALIZAÇÃO DO ESTADO DA SESSÃO
# =============================================================================
if "executado" not in st.session_state:
    st.session_state["executado"] = False

# =============================================================================
# PONTO DE ENTRADA
# =============================================================================
if __name__ == "__main__":
    main()
