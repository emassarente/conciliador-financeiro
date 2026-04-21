# =============================================================================
# DASHBOARD - CONCILIAÇÃO FINANCEIRA MULTI-EMPRESA
# Stack: Streamlit + Plotly + pandas + SQLite
# =============================================================================

import sys
import os
import logging
import html
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from pathlib import Path
from io import BytesIO
import datetime

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from import_.razao_parser import RazaoParser
from import_.extrato_parser import ExtratoParser
from engine.conciliacao_engine import ConciliacaoEngine
from database.db_manager import DatabaseManager

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
SIDEBAR_MANUAL_SLOT = None

# =============================================================================
# CONFIGURAÇÃO DA PÁGINA
# =============================================================================
st.set_page_config(
    page_title="Conciliação Financeira",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# CSS
# =============================================================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main { background-color: #f0f2f6; }
    [data-testid="stSidebar"] { background: #0f172a; }
    [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] .stRadio label { color: #cbd5e1 !important; }
    [data-testid="stSidebar"] hr { border-color: #1e293b; }

    /* KPI cards */
    .kpi-card {
        background: white;
        border-radius: 16px;
        padding: 22px 20px 18px 20px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.08);
        border-left: 5px solid;
        display: flex;
        flex-direction: column;
        gap: 4px;
    }
    .kpi-card.verde  { border-left-color: #10b981; }
    .kpi-card.azul   { border-left-color: #3b82f6; }
    .kpi-card.amarelo{ border-left-color: #f59e0b; }
    .kpi-card.vermelho{ border-left-color: #ef4444; }
    .kpi-card.roxo   { border-left-color: #8b5cf6; }
    .kpi-card.cinza  { border-left-color: #94a3b8; }
    .kpi-num  { font-size: 2rem; font-weight: 800; color: #0f172a; line-height: 1; }
    .kpi-lbl  { font-size: 0.78rem; font-weight: 600; color: #64748b; text-transform: uppercase; letter-spacing: .8px; }
    .kpi-sub  { font-size: 0.82rem; color: #94a3b8; }

    /* Page title */
    .page-title { font-size: 1.7rem; font-weight: 800; color: #0f172a; margin: 0; }
    .page-sub   { font-size: 0.9rem; color: #64748b; margin: 0; }

    /* Empresa badge na sidebar */
    .empresa-badge {
        background: #1e3a5f;
        border-radius: 10px;
        padding: 10px 14px;
        margin: 8px 0;
        border-left: 3px solid #3b82f6;
    }
    .empresa-badge-nome { font-size: 0.95rem; font-weight: 700; color: #f1f5f9 !important; }
    .empresa-badge-cnpj { font-size: 0.75rem; color: #94a3b8 !important; }

    /* Section header */
    .section-header {
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
        color: white !important;
        padding: 10px 18px;
        border-radius: 10px;
        font-weight: 700;
        font-size: 0.95rem;
        margin: 18px 0 10px 0;
    }

    /* Status badges */
    .badge { padding: 3px 11px; border-radius: 20px; font-size: 0.78rem; font-weight: 700; display:inline-block; }
    .badge-green  { background:#d1fae5; color:#065f46; }
    .badge-yellow { background:#fef3c7; color:#92400e; }
    .badge-blue   { background:#dbeafe; color:#1e40af; }
    .badge-red    { background:#fee2e2; color:#991b1b; }
    .badge-orange { background:#ffedd5; color:#9a3412; }
    .badge-purple { background:#ede9fe; color:#5b21b6; }

    /* Table */
    .stDataFrame { border-radius: 10px !important; overflow: hidden; }

    /* Empresa card */
    .emp-card {
        background: white;
        border-radius: 12px;
        padding: 18px 20px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.08);
        border-top: 4px solid #3b82f6;
        margin-bottom: 12px;
    }
    .emp-nome { font-size: 1rem; font-weight: 700; color: #0f172a; }
    .emp-cnpj { font-size: 0.8rem; color: #64748b; }

    /* Nav pills */
    div[data-testid="stRadio"] > div { flex-direction: column; gap: 4px; }
    div[data-testid="stRadio"] label {
        background: #1e293b;
        border-radius: 8px;
        padding: 8px 14px !important;
        cursor: pointer;
        font-size: 0.88rem !important;
        transition: background .15s;
    }
    div[data-testid="stRadio"] label:hover { background: #334155; }
    div[data-testid="stRadio"] label[data-baseweb="radio"] input:checked + div + div { font-weight: 700; }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# HELPERS
# =============================================================================
def fmt_moeda(v) -> str:
    if pd.isna(v) or v is None:
        return "-"
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)

def fmt_data(d) -> str:
    if pd.isna(d) or d is None:
        return "-"
    try:
        return pd.Timestamp(d).strftime("%d/%m/%Y")
    except Exception:
        return str(d)

def fmt_pct(v) -> str:
    if v is None or pd.isna(v):
        return "-"
    return f"{float(v):.1f}%"

MESES_PT = {
    "01":"Janeiro","02":"Fevereiro","03":"Março","04":"Abril",
    "05":"Maio","06":"Junho","07":"Julho","08":"Agosto",
    "09":"Setembro","10":"Outubro","11":"Novembro","12":"Dezembro"
}

STATUS_MAP = {
    "CONCILIADO":        "✅ CONCILIADO",
    "MATCH_COMBINADO":   "🟡 COMBINADO",
    "MATCH_PROVAVEL":    "🔵 SIMILARIDADE",
    "MATCH_APRENDIDO":   "🟣 APRENDIDO",
    "NAO_CONCILIADO":    "🔴 NÃO CONCILIADO",
    "MANUAL_CONCILIADO": "🟠 MANUAL",
}

STATUS_CORES = {
    "✅ CONCILIADO":      ("#d1fae5","#065f46"),
    "🟡 COMBINADO":       ("#fef3c7","#92400e"),
    "🔵 SIMILARIDADE":    ("#dbeafe","#1e40af"),
    "🟣 APRENDIDO":       ("#ede9fe","#5b21b6"),
    "🔴 NÃO CONCILIADO":  ("#fee2e2","#991b1b"),
    "🟠 MANUAL":          ("#ffedd5","#9a3412"),
}

def _badge_status(status: str) -> str:
    mapa = {
        "✅ CONCILIADO":     ("badge badge-green",  "✅ CONCILIADO"),
        "🟡 COMBINADO":      ("badge badge-yellow", "🟡 COMBINADO"),
        "🔵 SIMILARIDADE":   ("badge badge-blue",   "🔵 SIMILARIDADE"),
        "🟣 APRENDIDO":      ("badge badge-purple", "🟣 APRENDIDO"),
        "🔴 NÃO CONCILIADO": ("badge badge-red",    "🔴 NÃO CONCILIADO"),
        "🟠 MANUAL":         ("badge badge-orange", "🟠 MANUAL"),
    }
    cls, lbl = mapa.get(status, ("badge badge-red", status))
    return f'<span class="{cls}">{lbl}</span>'

@st.cache_resource
def get_db():
    return DatabaseManager()

def colorir_linha(row):
    s = row.get("Status", "")
    bg, fg = STATUS_CORES.get(s, ("",""))
    if bg:
        return [f"background-color:{bg};color:{fg}"] * len(row)
    return [""] * len(row)

def gerar_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conciliação")
        from openpyxl.styles import PatternFill, Font
        ws = writer.sheets["Conciliação"]
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        col_s = next((i for i,c in enumerate(df.columns,1) if c=="Status"), None)
        if col_s:
            cores_xl = {
                "✅ CONCILIADO":     ("D1FAE5","065F46"),
                "🟡 COMBINADO":      ("FEF3C7","92400E"),
                "🔵 SIMILARIDADE":   ("DBEAFE","1E40AF"),
                "🟣 APRENDIDO":      ("EDE9FE","5B21B6"),
                "🔴 NÃO CONCILIADO": ("FEE2E2","991B1B"),
                "🟠 MANUAL":         ("FFEDD5","9A3412"),
            }
            for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
                val = str(ws.cell(ri, col_s).value or "")
                if val in cores_xl:
                    bg, fg = cores_xl[val]
                    for cell in row:
                        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                        cell.font = Font(color=fg)
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(w+4, 50)
    return output.getvalue()

def preparar_df_exibicao(df_resultado: pd.DataFrame) -> pd.DataFrame:
    df = df_resultado.copy()
    df["data_razao"]    = df["data_razao"].apply(fmt_data)
    df["data_extrato"]  = df["data_extrato"].apply(fmt_data)
    df["valor_razao"]   = df["valor_razao"].apply(fmt_moeda)
    df["valor_extrato"] = df["valor_extrato"].apply(fmt_moeda)
    df["confidence"]    = df["confidence"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")
    df["status"]        = df["status"].map(STATUS_MAP).fillna(df["status"])
    cols = {
        "data_razao":"Data Razão","historico_razao":"Histórico (Razão)",
        "valor_razao":"Valor Razão","data_extrato":"Data Extrato",
        "descricao_extrato":"Descrição (Extrato)","valor_extrato":"Valor Extrato",
        "status":"Status","tipo_match":"Tipo Match","confidence":"Confiança","observacoes":"Obs.",
    }
    disp = {k:v for k,v in cols.items() if k in df.columns}
    return df[list(disp.keys())].rename(columns=disp)


# =============================================================================
# SIDEBAR
# =============================================================================
def renderizar_sidebar():
    global SIDEBAR_MANUAL_SLOT
    db = get_db()
    
    # Filtra clientes por contabilidade (exceto IGP que vê todos)
    usuario = st.session_state.get("usuario_logado", {})
    if usuario.get("perfil") == "igp":
        clientes = db.listar_clientes()
    else:
        empresas_permitidas = st.session_state.get("empresas_permitidas", [])
        if empresas_permitidas:
            clientes = empresas_permitidas
        else:
            clientes = db.listar_clientes()
    
    escritorio = db.obter_escritorio()
    nome_escritorio = escritorio["nome"] if escritorio else "IA Conciliação"

    with st.sidebar:
        # ── Cabeçalho do escritório
        st.markdown(f"""
        <div style="text-align:center; padding: 12px 0 4px 0;">
            <div style="font-size:1.4rem; font-weight:800; color:#f1f5f9; letter-spacing:-0.5px;">
                {nome_escritorio}
            </div>
            <div style="font-size:0.72rem; color:#64748b; font-weight:500; letter-spacing:1px;
                        text-transform:uppercase; margin-top:2px;">
                Sistema de Conciliação
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.divider()

        # ── Seleção de cliente
        st.markdown("<div style='font-size:0.72rem;font-weight:600;color:#94a3b8;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;'>Cliente Ativo</div>", unsafe_allow_html=True)
        if not clientes:
            st.warning("Nenhum cliente cadastrado.")
            cliente_id = None
            cliente_ativo = None
        else:
            opcoes = {c["nome"]: c["id"] for c in clientes}
            sel_nome = st.session_state.get("cliente_nome_sel", list(opcoes.keys())[0])
            if sel_nome not in opcoes:
                sel_nome = list(opcoes.keys())[0]

            nome_sel = st.selectbox(
                "Selecione o cliente",
                options=list(opcoes.keys()),
                index=list(opcoes.keys()).index(sel_nome),
                key="cliente_select_sidebar",
                label_visibility="collapsed",
            )
            cliente_id = opcoes[nome_sel]
            cliente_ativo = next((c for c in clientes if c["id"] == cliente_id), None)
            st.session_state["cliente_nome_sel"] = nome_sel
            st.session_state["cliente_id_ativo"] = cliente_id

            if cliente_ativo:
                st.markdown(f"""
                <div class="empresa-badge">
                    <div class="empresa-badge-nome">{cliente_ativo['nome']}</div>
                    <div class="empresa-badge-cnpj">CNPJ: {cliente_ativo.get('cnpj','—')}</div>
                    {('<div class="empresa-badge-cnpj">Resp.: '+cliente_ativo['responsavel']+'</div>') if cliente_ativo.get('responsavel') else ''}
                </div>
                """, unsafe_allow_html=True)

        SIDEBAR_MANUAL_SLOT = st.empty()
        st.divider()

        # ── Navegação
        st.markdown("<div style='font-size:0.72rem;font-weight:600;color:#94a3b8;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;'>Menu</div>", unsafe_allow_html=True)
        pagina = st.radio(
            "nav",
            ["🏠  Visão Geral", "📂  Conciliação", "📊  Análise Mensal", "👥  Clientes", "⚙️  Escritório"],
            label_visibility="collapsed",
            key="nav_radio",
        )

        st.divider()

        # ── Parâmetros
        with st.expander("⚙️ Parâmetros"):
            score_min = st.slider("Similaridade mínima (%)", 60, 95, 80, 5)
            usar_sim  = st.checkbox("Match por similaridade", True)
            tol_dias  = st.number_input("Tolerância datas (dias)", 0, 10, 3)

        st.divider()
        st.caption("v2.0 · Python + Streamlit")

    # Retrocompatibilidade: retorna empresa_id também como cliente_id
    return pagina, cliente_id, cliente_ativo, score_min, usar_sim, tol_dias


# =============================================================================
# PÁGINA: VISÃO GERAL
# =============================================================================
def pagina_visao_geral(empresa_id, empresa_ativa):
    db = get_db()
    nome_emp = empresa_ativa["nome"] if empresa_ativa else "Todos os Clientes"

    st.markdown(f'<p class="page-title">🏠 Visão Geral</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="page-sub">{nome_emp}</p>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    stats = db.estatisticas(cliente_id=empresa_id)
    df_base = db.consultar_lancamentos_emparelhados(cliente_id=empresa_id)
    df_mensal = db.dashboard_mensal(cliente_id=empresa_id)
    df_status = db.dashboard_por_status(cliente_id=empresa_id)
    df_banco  = db.dashboard_por_banco(cliente_id=empresa_id)

    total_r = stats["total_razao"]
    total_e = stats["total_extrato"]
    total_c = stats["total_conciliacoes"]
    taxa = (total_c / total_r * 100) if total_r > 0 else 0
    pendentes = total_r - total_c

    if not df_base.empty:
        df_base = df_base.copy()
        df_base["_data_ref"] = (
            df_base["data_razao"].fillna(df_base["data_extrato"])
            if "data_razao" in df_base.columns and "data_extrato" in df_base.columns
            else df_base["data_razao"]
            if "data_razao" in df_base.columns
            else df_base["data_extrato"]
        )

        def _classifica_visao(row):
            s = row.get("status", "")
            vr = pd.to_numeric(row.get("valor_razao", 0), errors="coerce")
            ve = pd.to_numeric(row.get("valor_extrato", 0), errors="coerce")
            valores_batem = (
                pd.notna(vr) and pd.notna(ve) and round(abs(abs(float(vr)) - abs(float(ve))), 2) == 0
            )
            if s == "CONCILIADO":
                return "exato"
            if s == "MATCH_APRENDIDO" and valores_batem:
                return "exato"
            if s == "MATCH_APRENDIDO":
                return "checar"
            if s == "MATCH_PROVAVEL":
                return "checar"
            if s in {"MATCH_COMBINADO", "MATCH_APRENDIDO", "MANUAL_CONCILIADO", "MANUAL_DIVERGENTE"}:
                return "cond"
            return "nao"

        df_base["_classe"] = df_base.apply(_classifica_visao, axis=1)
        n_exato = int((df_base["_classe"] == "exato").sum())
        n_cond = int((df_base["_classe"] == "cond").sum())
        n_checar = int((df_base["_classe"] == "checar").sum())
        n_nao = int((df_base["_classe"] == "nao").sum())

        total_c = n_exato + n_cond
        pendentes = n_nao
        taxa = (total_c / total_r * 100) if total_r > 0 else 0

        df_visao_mes = (
            df_base.assign(mes=df_base["_data_ref"].astype(str).str[:7])
            .groupby("mes", dropna=True)
            .agg(
                conciliados=("status", lambda s: 0),
                checar=("status", lambda s: 0),
                pendentes=("status", lambda s: 0),
            )
            .reset_index()
        )
        if not df_visao_mes.empty:
            for idx, row in df_visao_mes.iterrows():
                mes = row["mes"]
                fatia = df_base[df_base["_data_ref"].astype(str).str.startswith(mes)]
                df_visao_mes.at[idx, "conciliados"] = int(fatia["_classe"].isin(["exato", "cond"]).sum())
                df_visao_mes.at[idx, "checar"] = int((fatia["_classe"] == "checar").sum())
                df_visao_mes.at[idx, "pendentes"] = int((fatia["_classe"] == "nao").sum())
            df_visao_mes["mes_label"] = df_visao_mes["mes"].apply(
                lambda m: f"{MESES_PT.get(m[5:7], m[5:7])}/{m[:4]}" if m and len(m) >= 7 else m
            )
            df_mensal = df_visao_mes

        df_status = pd.DataFrame(
            [
                {"status": "CONCILIADO", "total": n_exato},
                {"status": "CONDICIONAL", "total": n_cond},
                {"status": "CHECAR", "total": n_checar},
                {"status": "NAO_CONCILIADO", "total": n_nao},
            ]
        )
        df_status = df_status[df_status["total"] > 0].reset_index(drop=True)

    # ── KPIs
    k1, k2, k3, k4, k5 = st.columns(5)
    kpis = [
        (k1, "cinza",   total_r,          "Lançamentos Razão",   ""),
        (k2, "azul",    total_e,          "Lançamentos Extrato", ""),
        (k3, "verde",   total_c,          "Conciliados",         ""),
        (k4, "vermelho",pendentes,        "Pendentes",           ""),
        (k5, "roxo",    f"{taxa:.1f}%",   "Taxa de Conciliação", ""),
    ]
    for col, cor, num, lbl, sub in kpis:
        with col:
            st.markdown(f"""
            <div class="kpi-card {cor}">
                <div class="kpi-lbl">{lbl}</div>
                <div class="kpi-num">{num}</div>
                {'<div class="kpi-sub">'+sub+'</div>' if sub else ''}
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Gráficos
    col_a, col_b = st.columns([2, 1])

    with col_a:
        if not df_mensal.empty:
            df_mensal["mes_label"] = df_mensal["mes"].apply(
                lambda m: f"{MESES_PT.get(m[5:7], m[5:7])}/{m[:4]}" if m else m
            )
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=df_mensal["mes_label"], y=df_mensal["conciliados"],
                name="Conciliados", marker_color="#10b981",
                text=df_mensal["conciliados"], textposition="outside",
            ))
            if "checar" in df_mensal.columns:
                fig.add_trace(go.Bar(
                    x=df_mensal["mes_label"], y=df_mensal["checar"],
                    name="Checar", marker_color="#fbbf24",
                    text=df_mensal["checar"], textposition="outside",
                ))
            fig.add_trace(go.Bar(
                x=df_mensal["mes_label"], y=df_mensal["pendentes"],
                name="Pendentes", marker_color="#f87171",
                text=df_mensal["pendentes"], textposition="outside",
            ))
            fig.update_layout(
                barmode="stack",
                title="Lançamentos por Mês",
                title_font_size=14,
                title_font_color="#0f172a",
                plot_bgcolor="white",
                paper_bgcolor="white",
                legend=dict(orientation="h", y=1.12),
                margin=dict(t=50, b=30, l=0, r=0),
                height=340,
                font_family="Inter",
            )
            fig.update_xaxes(showgrid=False)
            fig.update_yaxes(gridcolor="#f1f5f9")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Nenhum dado disponível. Importe arquivos para visualizar.")

    with col_b:
        if not df_status.empty:
            labels_map = {
                "CONCILIADO":"Conciliado","MATCH_COMBINADO":"Combinado",
                "MATCH_PROVAVEL":"Similaridade","MATCH_APRENDIDO":"Aprendido",
                "NAO_CONCILIADO":"Não Conciliado","MANUAL_CONCILIADO":"Manual",
                "CONDICIONAL":"Com condição","CHECAR":"Checar",
            }
            cores_pie = {
                "Conciliado":"#10b981","Combinado":"#f59e0b","Similaridade":"#3b82f6",
                "Aprendido":"#8b5cf6","Não Conciliado":"#ef4444","Manual":"#f97316",
                "Com condição":"#f59e0b","Checar":"#fbbf24",
            }
            df_s = df_status.copy()
            df_s["label"] = df_s["status"].map(labels_map).fillna(df_s["status"])
            df_s = df_s[df_s["total"] > 0]
            if not df_s.empty:
                fig2 = px.pie(
                    df_s, values="total", names="label",
                    color="label", color_discrete_map=cores_pie,
                    title="Distribuição por Status",
                    hole=0.45,
                )
                fig2.update_traces(textinfo="percent+label", textfont_size=12)
                fig2.update_layout(
                    showlegend=False,
                    margin=dict(t=50, b=10, l=0, r=0),
                    height=340,
                    paper_bgcolor="white",
                    font_family="Inter",
                )
                st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Sem conciliações registradas ainda.")

    # ── Evolução de valor conciliado
    if not df_mensal.empty and "valor_conciliado" in df_mensal.columns:
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(
            x=df_mensal["mes_label"],
            y=df_mensal["valor_total"].abs(),
            mode="lines+markers",
            name="Valor Total",
            line=dict(color="#94a3b8", width=2, dash="dot"),
            marker=dict(size=6),
        ))
        fig3.add_trace(go.Scatter(
            x=df_mensal["mes_label"],
            y=df_mensal["valor_conciliado"].abs(),
            mode="lines+markers",
            name="Valor Conciliado",
            line=dict(color="#3b82f6", width=2.5),
            marker=dict(size=7),
            fill="tozeroy",
            fillcolor="rgba(59,130,246,0.08)",
        ))
        fig3.update_layout(
            title="Evolução de Valor Conciliado (R$)",
            title_font_size=14,
            plot_bgcolor="white",
            paper_bgcolor="white",
            legend=dict(orientation="h", y=1.12),
            margin=dict(t=50, b=30, l=0, r=0),
            height=300,
            font_family="Inter",
        )
        fig3.update_yaxes(gridcolor="#f1f5f9", tickformat=",.0f", tickprefix="R$ ")
        fig3.update_xaxes(showgrid=False)
        st.plotly_chart(fig3, use_container_width=True)

    # ── Por banco
    if not df_banco.empty and df_banco["total_lancamentos"].sum() > 0:
        st.markdown('<div class="section-header">🏦 Extratos por Banco</div>', unsafe_allow_html=True)
        df_b = df_banco[df_banco["total_lancamentos"] > 0].copy()
        fig4 = px.bar(
            df_b, x="banco", y="total_lancamentos",
            color_discrete_sequence=["#3b82f6"],
            text="total_lancamentos",
            title="",
        )
        fig4.update_traces(textposition="outside")
        fig4.update_layout(
            plot_bgcolor="white", paper_bgcolor="white",
            margin=dict(t=10, b=20, l=0, r=0),
            height=260, font_family="Inter",
            xaxis_title="", yaxis_title="Lançamentos",
        )
        fig4.update_xaxes(showgrid=False)
        fig4.update_yaxes(gridcolor="#f1f5f9")
        st.plotly_chart(fig4, use_container_width=True)


# =============================================================================
# PÁGINA: ANÁLISE MENSAL — PAINEL BI
# =============================================================================
def _layout_chart(title="", height=320, legend_h=True):
    """Retorna dict padrão de layout para gráficos."""
    leg = dict(orientation="h", y=1.1, x=0) if legend_h else dict()
    return dict(
        title=dict(text=title, font=dict(size=14, color="#1e293b"), x=0),
        plot_bgcolor="white", paper_bgcolor="white",
        margin=dict(t=50, b=20, l=0, r=10),
        height=height,
        font=dict(family="Inter, sans-serif", size=12),
        legend=leg,
        hovermode="x unified",
    )


def pagina_analise_mensal(empresa_id, empresa_ativa):
    db = get_db()
    nome_emp = empresa_ativa["nome"] if empresa_ativa else "Todos os Clientes"

    st.markdown('<p class="page-title">📊 Análise Mensal</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="page-sub">{nome_emp}</p>', unsafe_allow_html=True)

    df_mensal = db.dashboard_mensal(cliente_id=empresa_id)
    df_status  = db.dashboard_por_status(cliente_id=empresa_id)
    df_banco   = db.dashboard_por_banco(cliente_id=empresa_id)

    if df_mensal.empty:
        st.info("Nenhum dado disponível ainda. Execute uma conciliação na aba **📂 Conciliação**.")
        return

    # ── Prepara dados
    df_mensal["mes_label"] = df_mensal["mes"].apply(
        lambda m: f"{MESES_PT.get(m[5:7], m[5:7])}/{m[:4]}" if m else m
    )
    df_mensal["taxa_conc"] = (
        df_mensal["conciliados"] / df_mensal["total_razao"].replace(0, 1) * 100
    ).round(1)
    df_mensal["valor_total_abs"]      = df_mensal["valor_total"].abs()
    df_mensal["valor_conciliado_abs"] = df_mensal["valor_conciliado"].abs()

    # ── Filtro de ano + mês
    anos = sorted({m[:4] for m in df_mensal["mes"] if m}, reverse=True)
    st.markdown("<br>", unsafe_allow_html=True)
    col_f1, col_f2, col_f3 = st.columns([1, 1, 2])
    with col_f1:
        ano_sel = st.selectbox("Ano", ["Todos"] + anos, key="bi_ano")
    with col_f2:
        meta_pct = st.number_input("Meta conciliação (%)", 50, 100, 90, 5, key="bi_meta")

    df_fil = df_mensal.copy()
    if ano_sel != "Todos":
        df_fil = df_fil[df_fil["mes"].str.startswith(ano_sel)]

    if df_fil.empty:
        st.warning("Nenhum dado para o período selecionado.")
        return

    # ═══════════════════════════════════════════════════════
    # LINHA 1 — KPI GAUGES (4 indicadores circulares)
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### 🎯 Indicadores Globais do Período")

    total_r   = int(df_fil["total_razao"].sum())
    total_c   = int(df_fil["conciliados"].sum())
    total_p   = int(df_fil["pendentes"].sum())
    taxa_geral = round(total_c / total_r * 100, 1) if total_r else 0
    val_tot   = df_fil["valor_total_abs"].sum()
    val_conc  = df_fil["valor_conciliado_abs"].sum()
    taxa_val  = round(val_conc / val_tot * 100, 1) if val_tot else 0
    melhor_mes_row = df_fil.loc[df_fil["taxa_conc"].idxmax()]
    melhor_mes = melhor_mes_row["mes_label"]

    def make_gauge(value, title, color, max_val=100, suffix="%"):
        bar_color = "#10b981" if value >= meta_pct else ("#f59e0b" if value >= meta_pct * 0.7 else "#ef4444")
        fig = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=value,
            title=dict(text=title, font=dict(size=13, color="#475569")),
            number=dict(suffix=suffix, font=dict(size=28, color="#1e293b")),
            delta=dict(reference=meta_pct, valueformat=".1f",
                       increasing=dict(color="#10b981"),
                       decreasing=dict(color="#ef4444")),
            gauge=dict(
                axis=dict(range=[0, max_val], tickwidth=1, tickcolor="#cbd5e1",
                          tickfont=dict(size=10)),
                bar=dict(color=bar_color, thickness=0.25),
                bgcolor="white",
                borderwidth=0,
                steps=[
                    dict(range=[0, meta_pct * 0.7], color="#fef2f2"),
                    dict(range=[meta_pct * 0.7, meta_pct], color="#fef9c3"),
                    dict(range=[meta_pct, max_val], color="#f0fdf4"),
                ],
                threshold=dict(
                    line=dict(color="#3b82f6", width=3),
                    thickness=0.85,
                    value=meta_pct,
                ),
            ),
        ))
        fig.update_layout(
            paper_bgcolor="white",
            margin=dict(t=30, b=10, l=10, r=10),
            height=200,
            font=dict(family="Inter, sans-serif"),
        )
        return fig

    g1, g2, g3, g4 = st.columns(4)
    with g1:
        st.plotly_chart(make_gauge(taxa_geral, "Taxa Conciliação", "#10b981"), use_container_width=True, key="g1")
    with g2:
        st.plotly_chart(make_gauge(taxa_val, "Taxa por Valor (R$)", "#3b82f6"), use_container_width=True, key="g2")
    with g3:
        max_mes = df_fil["taxa_conc"].max()
        st.plotly_chart(make_gauge(max_mes, "Melhor Mês", "#8b5cf6"), use_container_width=True, key="g3")
    with g4:
        min_mes = df_fil["taxa_conc"].min()
        st.plotly_chart(make_gauge(min_mes, "Pior Mês", "#f59e0b"), use_container_width=True, key="g4")

    # Métricas rápidas abaixo dos gauges
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Lançamentos", f"{total_r:,}")
    m2.metric("✅ Conciliados", f"{total_c:,}", delta=f"{taxa_geral:.1f}%")
    m3.metric("⏳ Pendentes", f"{total_p:,}", delta=f"-{100-taxa_geral:.1f}%", delta_color="inverse")
    m4.metric("💰 Valor Total", fmt_moeda(val_tot))
    m5.metric("💚 Valor Conciliado", fmt_moeda(val_conc))

    # ═══════════════════════════════════════════════════════
    # LINHA 2 — Rosca de status + Evolução taxa (área)
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    col_donut, col_area = st.columns([1, 2])

    with col_donut:
        st.markdown("#### 🍩 Status de Conciliação")
        STATUS_LABELS = {
            "CONCILIADO":        "Conciliado",
            "MATCH_COMBINADO":   "Match Combinado",
            "MATCH_SIMILARIDADE":"Match Similaridade",
            "MATCH_PROVAVEL":    "Match Provável",
            "MANUAL_CONCILIADO": "Manual",
            "NAO_CONCILIADO":    "Não Conciliado",
        }
        STATUS_CORES = {
            "CONCILIADO":        "#10b981",
            "MATCH_COMBINADO":   "#3b82f6",
            "MATCH_SIMILARIDADE":"#8b5cf6",
            "MATCH_PROVAVEL":    "#06b6d4",
            "MANUAL_CONCILIADO": "#f59e0b",
            "NAO_CONCILIADO":    "#ef4444",
        }
        if not df_status.empty:
            labels = [STATUS_LABELS.get(s, s) for s in df_status["status"]]
            cores  = [STATUS_CORES.get(s, "#94a3b8") for s in df_status["status"]]
            fig_donut = go.Figure(go.Pie(
                labels=labels,
                values=df_status["total"],
                hole=0.55,
                marker=dict(colors=cores, line=dict(color="white", width=2)),
                textinfo="percent",
                hovertemplate="<b>%{label}</b><br>%{value} registros<br>%{percent}<extra></extra>",
                pull=[0.04 if s == "NAO_CONCILIADO" else 0 for s in df_status["status"]],
            ))
            total_geral_donut = int(df_status["total"].sum())
            fig_donut.add_annotation(
                text=f"<b>{total_geral_donut}</b><br><span style='font-size:11px'>total</span>",
                x=0.5, y=0.5, font_size=20, showarrow=False,
                font=dict(color="#1e293b"),
            )
            fig_donut.update_layout(
                paper_bgcolor="white",
                margin=dict(t=10, b=10, l=0, r=0),
                height=300,
                showlegend=True,
                legend=dict(orientation="v", x=1.0, y=0.5, font=dict(size=11)),
                font=dict(family="Inter, sans-serif"),
            )
            st.plotly_chart(fig_donut, use_container_width=True, key="donut_status")
        else:
            st.info("Execute uma conciliação para ver o status.")

    with col_area:
        st.markdown("#### 📈 Evolução da Taxa de Conciliação")
        fig_area = go.Figure()
        # Área preenchida para taxa
        fig_area.add_trace(go.Scatter(
            x=df_fil["mes_label"], y=df_fil["taxa_conc"],
            mode="lines+markers+text",
            name="Taxa (%)",
            line=dict(color="#3b82f6", width=3),
            marker=dict(size=9, color="#3b82f6",
                        line=dict(color="white", width=2)),
            fill="tozeroy",
            fillcolor="rgba(59,130,246,0.12)",
            text=df_fil["taxa_conc"].apply(lambda v: f"{v:.0f}%"),
            textposition="top center",
            textfont=dict(size=11, color="#1e40af"),
            hovertemplate="<b>%{x}</b><br>Taxa: %{y:.1f}%<extra></extra>",
        ))
        # Linha de meta
        fig_area.add_hline(
            y=meta_pct,
            line=dict(color="#f59e0b", width=2, dash="dot"),
            annotation_text=f"Meta {meta_pct}%",
            annotation_position="top right",
            annotation_font=dict(color="#b45309", size=12),
        )
        # Zonas coloridas de fundo
        fig_area.add_hrect(y0=0, y1=meta_pct * 0.7,
                           fillcolor="rgba(239,68,68,0.04)", layer="below", line_width=0)
        fig_area.add_hrect(y0=meta_pct * 0.7, y1=meta_pct,
                           fillcolor="rgba(245,158,11,0.04)", layer="below", line_width=0)
        fig_area.add_hrect(y0=meta_pct, y1=101,
                           fillcolor="rgba(16,185,129,0.05)", layer="below", line_width=0)

        fig_area.update_layout(
            **_layout_chart(height=300, legend_h=False),
            yaxis=dict(range=[0, 105], ticksuffix="%", gridcolor="#f1f5f9"),
            xaxis=dict(showgrid=False),
        )
        st.plotly_chart(fig_area, use_container_width=True, key="area_taxa")

    # ═══════════════════════════════════════════════════════
    # LINHA 3 — Barras agrupadas mês a mês + Valor R$
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    col_bar, col_val = st.columns(2)

    with col_bar:
        st.markdown("#### 📊 Volume por Mês")
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            x=df_fil["mes_label"], y=df_fil["conciliados"],
            name="Conciliados", marker_color="#10b981",
            hovertemplate="<b>%{x}</b><br>Conciliados: %{y}<extra></extra>",
        ))
        fig_bar.add_trace(go.Bar(
            x=df_fil["mes_label"], y=df_fil["pendentes"],
            name="Pendentes", marker_color="#ef4444",
            hovertemplate="<b>%{x}</b><br>Pendentes: %{y}<extra></extra>",
        ))
        fig_bar.update_layout(
            **_layout_chart(height=300),
            barmode="stack",
            yaxis=dict(gridcolor="#f1f5f9"),
            xaxis=dict(showgrid=False),
        )
        st.plotly_chart(fig_bar, use_container_width=True, key="bar_vol")

    with col_val:
        st.markdown("#### 💰 Valor Conciliado vs Total (R$)")
        fig_val = go.Figure()
        fig_val.add_trace(go.Bar(
            x=df_fil["mes_label"], y=df_fil["valor_total_abs"],
            name="Total", marker_color="#e2e8f0",
            hovertemplate="<b>%{x}</b><br>Total: R$ %{y:,.2f}<extra></extra>",
        ))
        fig_val.add_trace(go.Bar(
            x=df_fil["mes_label"], y=df_fil["valor_conciliado_abs"],
            name="Conciliado", marker_color="#3b82f6",
            hovertemplate="<b>%{x}</b><br>Conciliado: R$ %{y:,.2f}<extra></extra>",
        ))
        fig_val.update_layout(
            **_layout_chart(height=300),
            barmode="overlay",
            yaxis=dict(gridcolor="#f1f5f9", tickformat=",.0f"),
            xaxis=dict(showgrid=False),
        )
        st.plotly_chart(fig_val, use_container_width=True, key="bar_val")

    # ═══════════════════════════════════════════════════════
    # LINHA 4 — Rosca bancos + Heatmap mensal
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    col_banco, col_heat = st.columns([1, 2])

    with col_banco:
        st.markdown("#### 🏦 Extrato por Banco")
        if not df_banco.empty:
            fig_banco = go.Figure(go.Pie(
                labels=df_banco["banco"],
                values=df_banco["total_lancamentos"],
                hole=0.5,
                marker=dict(
                    colors=["#3b82f6","#8b5cf6","#10b981","#f59e0b","#ef4444","#06b6d4"],
                    line=dict(color="white", width=2),
                ),
                textinfo="percent+label",
                hovertemplate="<b>%{label}</b><br>%{value} lançamentos<extra></extra>",
            ))
            fig_banco.update_layout(
                paper_bgcolor="white",
                margin=dict(t=10, b=10, l=0, r=0),
                height=300,
                showlegend=False,
                font=dict(family="Inter, sans-serif"),
            )
            st.plotly_chart(fig_banco, use_container_width=True, key="donut_banco")
        else:
            st.info("Sem dados de extrato bancário.")

    with col_heat:
        st.markdown("#### 🗓️ Heatmap — Taxa de Conciliação por Mês/Ano")
        if len(df_mensal) >= 2:
            import numpy as np
            df_heat = df_mensal[["mes", "taxa_conc"]].copy()
            df_heat["ano"] = df_heat["mes"].str[:4]
            df_heat["mes_num"] = df_heat["mes"].str[5:7]
            df_heat["mes_nome"] = df_heat["mes_num"].map(MESES_PT)

            pivot = df_heat.pivot_table(
                index="ano", columns="mes_nome",
                values="taxa_conc", aggfunc="mean"
            )
            ordem_meses = [MESES_PT[f"{i:02d}"] for i in range(1,13) if MESES_PT.get(f"{i:02d}") in pivot.columns]
            pivot = pivot.reindex(columns=ordem_meses)

            fig_heat = go.Figure(go.Heatmap(
                z=pivot.values,
                x=pivot.columns.tolist(),
                y=pivot.index.tolist(),
                colorscale=[
                    [0.0,  "#fef2f2"],
                    [0.4,  "#fca5a5"],
                    [0.6,  "#fde68a"],
                    [0.75, "#86efac"],
                    [1.0,  "#10b981"],
                ],
                zmin=0, zmax=100,
                text=[[f"{v:.0f}%" if not np.isnan(v) else "" for v in row] for row in pivot.values],
                texttemplate="%{text}",
                textfont=dict(size=12),
                hoverongaps=False,
                hovertemplate="<b>%{y} — %{x}</b><br>Taxa: %{z:.1f}%<extra></extra>",
                colorbar=dict(
                    title="Taxa %",
                    ticksuffix="%",
                    thickness=12,
                    len=0.8,
                ),
            ))
            fig_heat.update_layout(
                paper_bgcolor="white",
                plot_bgcolor="white",
                margin=dict(t=20, b=20, l=40, r=20),
                height=300 if len(pivot) <= 2 else 250 * len(pivot),
                font=dict(family="Inter, sans-serif", size=12),
                xaxis=dict(side="top"),
            )
            st.plotly_chart(fig_heat, use_container_width=True, key="heatmap")
        else:
            st.info("São necessários dados de pelo menos 2 meses para o heatmap.")

    # ═══════════════════════════════════════════════════════
    # LINHA 5 — Tabela detalhada expandível
    # ═══════════════════════════════════════════════════════
    st.markdown("---")
    with st.expander("📋 Tabela Detalhada por Mês", expanded=False):
        df_show = df_fil[[
            "mes_label","total_razao","conciliados","pendentes",
            "taxa_conc","valor_total_abs","valor_conciliado_abs"
        ]].copy()
        df_show.columns = ["Mês","Total","Conciliados","Pendentes","Taxa (%)","Valor Total","Valor Conciliado"]
        df_show["Valor Total"]      = df_show["Valor Total"].apply(fmt_moeda)
        df_show["Valor Conciliado"] = df_show["Valor Conciliado"].apply(fmt_moeda)
        df_show["Taxa (%)"]         = df_show["Taxa (%)"].apply(lambda v: f"{v:.1f}%")
        st.dataframe(
            df_show,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Taxa (%)": st.column_config.TextColumn("Taxa (%)"),
                "Conciliados": st.column_config.NumberColumn("✅ Conciliados"),
                "Pendentes":   st.column_config.NumberColumn("⏳ Pendentes"),
            },
        )


# =============================================================================
# PÁGINA: CLIENTES
# =============================================================================
def pagina_clientes():
    db = get_db()
    escritorio = db.obter_escritorio()
    nome_escritorio = escritorio["nome"] if escritorio else "IA Conciliação"

    st.markdown('<p class="page-title">👥 Clientes</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="page-sub">Empresas atendidas por {nome_escritorio}</p>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    tab_lista, tab_novo, tab_editar, tab_reset = st.tabs([
        "📋 Clientes Cadastrados", "➕ Novo Cliente", "✏️ Editar / Desativar", "🧪 Reset para Apresentação"
    ])

    with tab_lista:
        # Filtra clientes por contabilidade (exceto IGP)
        usuario = st.session_state.get("usuario_logado", {})
        if usuario.get("perfil") == "igp":
            clientes = db.listar_clientes(apenas_ativos=False)
        else:
            empresas_permitidas = st.session_state.get("empresas_permitidas", [])
            clientes = empresas_permitidas if empresas_permitidas else db.listar_clientes(apenas_ativos=False)
        
        if not clientes:
            st.info("Nenhum cliente cadastrado ainda. Use a aba **Novo Cliente** para começar.")
        else:
            for cli in clientes:
                stats = db.estatisticas(cliente_id=cli["id"])
                ativo_badge = "🟢 Ativo" if cli.get("ativo", 1) else "🔴 Inativo"
                col_info, col_stats = st.columns([2, 2])
                with col_info:
                    st.markdown(f"""
                    <div class="emp-card">
                        <div class="emp-nome">{cli['nome']}</div>
                        <div class="emp-cnpj">CNPJ: {cli.get('cnpj','—')} · Código: {cli.get('codigo','—')}</div>
                        {'<div class="emp-cnpj">Responsável: '+cli['responsavel']+'</div>' if cli.get('responsavel') else ''}
                        <div style="margin-top:6px;font-size:0.8rem;color:#64748b;">{ativo_badge} · ID: {cli['id']} · Desde: {cli.get('criado_em','—')[:10]}</div>
                    </div>
                    """, unsafe_allow_html=True)
                with col_stats:
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Razão", stats["total_razao"])
                    c2.metric("Extrato", stats["total_extrato"])
                    c3.metric("Conciliações", stats["total_conciliacoes"])
                st.divider()

    with tab_novo:
        st.subheader("Cadastrar Novo Cliente")
        with st.form("form_novo_cliente", clear_on_submit=True):
            nome = st.text_input("Razão Social *", placeholder="Ex: AFRIKA Consultoria e Tecnologia Ltda")
            col1, col2 = st.columns(2)
            with col1:
                cnpj = st.text_input("CNPJ", placeholder="00.000.000/0000-00")
            with col2:
                codigo = st.text_input("Código interno", placeholder="Ex: CLI001")
            responsavel = st.text_input("Responsável / Contato", placeholder="Nome do responsável na empresa")
            submitted = st.form_submit_button("✅ Cadastrar Cliente", type="primary", use_container_width=True)
            if submitted:
                if not nome.strip():
                    st.error("Razão Social é obrigatória.")
                else:
                    cid = db.criar_cliente(nome, cnpj, codigo, responsavel)
                    if cid > 0:
                        st.success(f"Cliente **{nome}** cadastrado com sucesso.")
                        st.rerun()
                    else:
                        st.error("CNPJ já cadastrado ou erro ao salvar.")

    with tab_editar:
        clientes = db.listar_clientes(apenas_ativos=False)
        if not clientes:
            st.info("Nenhum cliente cadastrado.")
        else:
            opcoes = {f"{c['nome']} (ID: {c['id']})": c for c in clientes}
            sel = st.selectbox("Selecione o cliente", list(opcoes.keys()))
            cli = opcoes[sel]

            with st.form("form_editar_cliente"):
                nome2  = st.text_input("Razão Social", value=cli["nome"])
                c1, c2 = st.columns(2)
                with c1:
                    cnpj2  = st.text_input("CNPJ", value=cli.get("cnpj",""))
                with c2:
                    cod2   = st.text_input("Código", value=cli.get("codigo",""))
                resp2 = st.text_input("Responsável", value=cli.get("responsavel",""))

                col_s, col_d = st.columns(2)
                with col_s:
                    salvar = st.form_submit_button("💾 Salvar Alterações", use_container_width=True)
                with col_d:
                    toggle = st.form_submit_button(
                        "🔴 Desativar" if cli.get("ativo", 1) else "🟢 Reativar",
                        use_container_width=True
                    )

                if salvar:
                    db.atualizar_cliente(cli["id"], nome2, cnpj2, cod2, resp2)
                    st.success("Cliente atualizado.")
                    st.rerun()
                if toggle:
                    if cli.get("ativo", 1):
                        db.desativar_cliente(cli["id"])
                        st.warning(f"Cliente **{cli['nome']}** desativado.")
                    else:
                        db.reativar_cliente(cli["id"])
                        st.success(f"Cliente **{cli['nome']}** reativado.")
                    st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    with tab_reset:
        st.markdown("""
        <div style="background:#fef9c3;border:1px solid #f59e0b;border-radius:10px;
                    padding:14px 18px;margin-bottom:20px;font-size:0.9rem;color:#78350f;">
            <b>⚠️ Área exclusiva para testes e apresentações.</b><br>
            Use para limpar dados de uma empresa e reimportá-los ao vivo na frente do cliente.
        </div>
        """, unsafe_allow_html=True)

        clientes_reset = db.listar_clientes(apenas_ativos=False)
        if not clientes_reset:
            st.info("Nenhum cliente cadastrado ainda.")
        else:
            opcoes_r = {f"{c['nome']} (ID: {c['id']})": c for c in clientes_reset}
            sel_r = st.selectbox("Selecione a empresa para resetar",
                                 list(opcoes_r.keys()), key="reset_sel_cli")
            cli_r = opcoes_r[sel_r]

            # Mostra estatísticas atuais
            stats_r = db.estatisticas(cliente_id=cli_r["id"])
            cr1, cr2, cr3 = st.columns(3)
            cr1.metric("Lançamentos Razão",    stats_r["total_razao"])
            cr2.metric("Lançamentos Extrato",  stats_r["total_extrato"])
            cr3.metric("Conciliações",         stats_r["total_conciliacoes"])

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("**Escolha o modo de reset:**")

            modo = st.radio(
                "modo_reset",
                [
                    "🗂️ Limpar apenas os dados  (mantém empresa cadastrada — reimportar na apresentação)",
                    "💥 Apagar tudo  (remove empresa + todos os dados — fluxo 100% do zero)",
                ],
                key="reset_modo",
                label_visibility="collapsed",
            )

            # Confirmação com nome digitado para evitar acidentes
            st.markdown("<br>", unsafe_allow_html=True)
            confirmacao = st.text_input(
                f"Digite o nome da empresa para confirmar: **{cli_r['nome']}**",
                key="reset_confirmacao",
                placeholder=cli_r["nome"],
            )
            nome_bate = confirmacao.strip().upper() == cli_r["nome"].strip().upper()

            col_exec, _ = st.columns([1, 3])
            with col_exec:
                executar = st.button(
                    "🔴 Executar Reset",
                    type="primary",
                    use_container_width=True,
                    key="btn_reset_exec",
                    disabled=not nome_bate,
                )

            if not nome_bate and confirmacao:
                st.caption("⚠️ Nome não confere. Digite exatamente como aparece acima.")

            if executar and nome_bate:
                cid_reset = cli_r["id"]
                apagar_empresa = "Apagar tudo" in modo
                try:
                    db.limpar_dados_cliente(cid_reset)
                    if apagar_empresa:
                        with db._conn() as conn:
                            conn.execute("DELETE FROM clientes WHERE id=?", (cid_reset,))
                        # Se era o cliente ativo, volta ao hub
                        if st.session_state.get("cliente_id_ativo") == cid_reset:
                            st.session_state.pop("cliente_id_ativo", None)
                            st.session_state.pop("cliente_nome_sel", None)
                            st.session_state["hub_ativo"] = True
                        st.success(f"✅ Empresa **{cli_r['nome']}** e todos os dados foram removidos. "
                                   f"Ao importar novamente, ela será recriada automaticamente.")
                    else:
                        st.success(f"✅ Dados de **{cli_r['nome']}** apagados. "
                                   f"A empresa permanece cadastrada. Basta reimportar na apresentação!")
                    st.session_state["reset_confirmacao"] = ""
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao executar reset: {e}")


# =============================================================================
# PÁGINA: ESCRITÓRIO
# =============================================================================
def pagina_escritorio():
    db = get_db()
    escritorio = db.obter_escritorio() or {}

    st.markdown('<p class="page-title">⚙️ Escritório</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-sub">Dados do escritório de contabilidade (operador do sistema)</p>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    col_form, col_card = st.columns([2, 1])

    with col_form:
        st.markdown('<div class="section-header">🏢 Dados do Escritório</div>', unsafe_allow_html=True)
        with st.form("form_escritorio"):
            nome_e = st.text_input("Nome / Razão Social *",
                                   value=escritorio.get("nome", ""),
                                   placeholder="Ex: IA Conciliação")
            col1, col2 = st.columns(2)
            with col1:
                cnpj_e = st.text_input("CNPJ",
                                       value=escritorio.get("cnpj", ""),
                                       placeholder="00.000.000/0000-00")
            with col2:
                resp_e = st.text_input("Responsável",
                                       value=escritorio.get("responsavel", ""),
                                       placeholder="Nome do sócio/contador")
            col3, col4 = st.columns(2)
            with col3:
                tel_e = st.text_input("Telefone",
                                      value=escritorio.get("telefone", ""),
                                      placeholder="(00) 00000-0000")
            with col4:
                email_e = st.text_input("E-mail",
                                        value=escritorio.get("email", ""),
                                        placeholder="contato@escritorio.com.br")
            salvar = st.form_submit_button("💾 Salvar", type="primary", use_container_width=True)
            if salvar:
                if not nome_e.strip():
                    st.error("Nome do escritório é obrigatório.")
                else:
                    db.salvar_escritorio(nome_e, cnpj_e, resp_e, tel_e, email_e)
                    st.success("✅ Dados do escritório salvos com sucesso.")
                    st.rerun()

    with col_card:
        if escritorio:
            st.markdown('<div class="section-header">📍 Cadastro Atual</div>', unsafe_allow_html=True)
            st.markdown(f"""
            <div class="emp-card" style="border-top-color:#10b981;">
                <div class="emp-nome">{escritorio.get('nome', '—')}</div>
                <div class="emp-cnpj">CNPJ: {escritorio.get('cnpj', '—')}</div>
                <div class="emp-cnpj">Resp.: {escritorio.get('responsavel', '—')}</div>
                <div class="emp-cnpj">☎️ {escritorio.get('telefone', '—')}</div>
                <div class="emp-cnpj">✉️ {escritorio.get('email', '—')}</div>
            </div>
            """, unsafe_allow_html=True)

            # Totais globais do sistema
            st.markdown('<div class="section-header">📊 Resumo Global</div>', unsafe_allow_html=True)
            clientes = db.listar_clientes()
            stats_global = db.estatisticas()  # sem filtro = todos
            c1, c2 = st.columns(2)
            c1.metric("Clientes ativos", len(clientes))
            c2.metric("Total conciliações", stats_global["total_conciliacoes"])
            st.metric("Lançamentos Razão", stats_global["total_razao"])
            st.metric("Lançamentos Extrato", stats_global["total_extrato"])
        else:
            st.info("Preencha e salve os dados do escritório ao lado.")


# =============================================================================
# HELPERS: LANÇAMENTOS POR PERÍODO
# =============================================================================

def _carregar_lancamentos_banco(db, empresa_id: int) -> pd.DataFrame:
    """
    Carrega todos os lançamentos (conciliados e não-conciliados) do banco.
    Retorna DataFrame com colunas: status, data_razao, historico_razao,
    valor_razao, data_extrato, descricao_extrato, valor_extrato.
    """
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = db.consultar_lancamentos_emparelhados(cliente_id=empresa_id)
        if df is None or (hasattr(df, "empty") and df.empty):
            return pd.DataFrame()
        # Garante colunas mínimas
        for col in ["data_razao", "historico_razao", "valor_razao",
                    "data_extrato", "descricao_extrato", "valor_extrato", "status"]:
            if col not in df.columns:
                df[col] = None
        return df.reset_index(drop=True)
    except Exception as e:
        logger.warning(f"_carregar_lancamentos_banco: {e}")
        return pd.DataFrame()


def _renderizar_lancamentos_por_mes(db, empresa_id: int, df_base: pd.DataFrame):
    """
    Exibe lançamentos em 3 boxes por mês:
      Box 1 🟢 Conciliados 100% — linhas verdes + totalizador
      Box 2 🟡 Conciliados c/ condição — linhas amarelas
      Box 3 🔴 Não conciliados — checkboxes ambos lados + acumulador + vincular
    """
    STATUS_EXATO  = {"CONCILIADO"}
    STATUS_COND   = {"MATCH_COMBINADO", "MATCH_APRENDIDO", "MATCH_PROVAVEL",
                     "MANUAL_CONCILIADO", "MANUAL_DIVERGENTE"}
    STATUS_NAO    = {"NAO_CONCILIADO"}

    data_ref = (
        df_base["data_razao"].fillna(df_base["data_extrato"])
        if "data_razao" in df_base.columns and "data_extrato" in df_base.columns
        else df_base["data_razao"]
        if "data_razao" in df_base.columns
        else df_base["data_extrato"]
    )

    # ── Filtros
    st.markdown('<div class="section-header">🔍 Filtros</div>', unsafe_allow_html=True)
    fa, fb, fc = st.columns([2, 2, 6])

    datas_str = data_ref.dropna().astype(str)
    meses_raw = sorted(set(d[:7] for d in datas_str if len(d) >= 7))
    meses_lbl = ["Todos os meses"] + [
        f"{MESES_PT.get(m[5:7], m[5:7])}/{m[:4]}" for m in meses_raw
    ]
    meses_map = {"Todos os meses": None}
    meses_map.update({f"{MESES_PT.get(m[5:7], m[5:7])}/{m[:4]}": m for m in meses_raw})

    with fa:
        sel_mes_lbl = st.selectbox("📅 Mês", meses_lbl, key="pm_mes")
        sel_mes = meses_map.get(sel_mes_lbl)
    with fb:
        mostrar_por = st.selectbox("👁️ Visualização", ["Por Mês", "Tudo junto"], key="pm_view")
    with fc:
        busca = st.text_input("🔍 Buscar", key="pm_busca",
                              placeholder="Histórico, descrição, valor...")

    # ── Aplica filtros
    df_view = df_base.copy()
    df_view["_data_ref"] = (
        df_view["data_razao"].fillna(df_view["data_extrato"])
        if "data_razao" in df_view.columns and "data_extrato" in df_view.columns
        else df_view["data_razao"]
        if "data_razao" in df_view.columns
        else df_view["data_extrato"]
    )
    if sel_mes:
        df_view = df_view[df_view["_data_ref"].fillna("").astype(str).str.startswith(sel_mes)]
    if busca:
        cols_b = [c for c in ["historico_razao", "descricao_extrato"]
                  if c in df_view.columns]
        if cols_b:
            mask = pd.Series(False, index=df_view.index)
            for c in cols_b:
                mask |= df_view[c].fillna("").astype(str).str.contains(busca, case=False)
            df_view = df_view[mask]

    if df_view.empty:
        st.info("Nenhum lançamento encontrado com os filtros selecionados.")
        return

    # Classificação — leva em conta divergência de valor
    TOLERANCIA_VALOR_COND = 0.05   # 5%: acima disso vai para Box 3 mesmo com status COND

    def _classifica(row):
        s = row.get("status", "")
        vr = pd.to_numeric(row.get("valor_razao", 0), errors="coerce")
        ve = pd.to_numeric(row.get("valor_extrato", 0), errors="coerce")
        valores_batem = (
            pd.notna(vr) and pd.notna(ve) and round(abs(abs(float(vr)) - abs(float(ve))), 2) == 0
        )
        if s in STATUS_EXATO:
            return "exato"
        if s == "MATCH_APRENDIDO" and valores_batem:
            return "exato"
        if s == "MATCH_APRENDIDO":
            return "checar"
        # MATCH_PROVAVEL sempre vai para Checar — precisa de confirmação manual
        if s == "MATCH_PROVAVEL":
            return "checar"
        if s in STATUS_COND:
            return "cond"
        return "nao"

    df_view = df_view.copy()
    df_view["_classe"] = df_view.apply(_classifica, axis=1)

    n_exato  = (df_view["_classe"] == "exato").sum()
    n_cond   = (df_view["_classe"] == "cond").sum()
    n_checar = (df_view["_classe"] == "checar").sum()
    n_nao    = (df_view["_classe"] == "nao").sum()

    # Resumo topo
    st.markdown(
        f"<small>📋 <b>{len(df_view):,}</b> registros — "
        f"🟢 {n_exato:,} conciliados · "
        f"🟡 {n_cond:,} com condição · "
        f"⚠️ {n_checar:,} para checar · "
        f"🔴 {n_nao:,} pendentes</small>",
        unsafe_allow_html=True
    )
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Helper: formata df para exibição
    def _formatar(df_t: pd.DataFrame) -> pd.DataFrame:
        def _condicao(row):
            status = str(row.get("status", "") or "")
            tipo = str(row.get("tipo_match", "") or "")
            obs = str(row.get("observacoes", "") or "")
            if status == "MATCH_APRENDIDO":
                return "Aprendido"
            if status == "MATCH_COMBINADO":
                return "Combinado"
            if status == "MATCH_PROVAVEL":
                return "Provável"
            if status == "MANUAL_DIVERGENTE":
                return obs or "Manual com divergência"
            if status == "MANUAL_CONCILIADO":
                return obs or "Manual"
            if tipo:
                return tipo
            return "—"

        nomes = {
            "data_razao":        "Data Razão",
            "historico_razao":   "Histórico Razão",
            "valor_razao":       "Valor Razão (R$)",
            "data_extrato":      "Data Extrato",
            "descricao_extrato": "Descrição Extrato",
            "valor_extrato":     "Valor Extrato (R$)",
            "condicao":          "Condição",
            "status":            "Status",
        }
        STATUS_LABEL = {
            "CONCILIADO":        "✅ Conciliado",
            "MATCH_COMBINADO":   "🟡 Combinado",
            "MATCH_APRENDIDO":   "🧠 Aprendido",
            "MATCH_PROVAVEL":    "🔵 Provável",
            "NAO_CONCILIADO":    "🔴 Pendente",
            "MANUAL_CONCILIADO": "🟠 Manual",
            "MANUAL_DIVERGENTE": "⚠️ Divergente",
        }
        df_aux = df_t.copy()
        df_aux["condicao"] = df_aux.apply(_condicao, axis=1)
        cols = [c for c in nomes if c in df_aux.columns]
        df_e = df_aux[cols].rename(columns=nomes).copy()
        for cd in ["Data Razão", "Data Extrato"]:
            if cd in df_e.columns:
                df_e[cd] = df_e[cd].apply(fmt_data)
        for cv in ["Valor Razão (R$)", "Valor Extrato (R$)"]:
            if cv in df_e.columns:
                df_e[cv] = df_e[cv].apply(
                    lambda x: fmt_moeda(x) if pd.notna(x) and str(x) not in ("", "nan") else "—"
                )
        if "Status" in df_e.columns:
            df_e["Status"] = df_e["Status"].map(lambda s: STATUS_LABEL.get(s, s))
        return df_e

    # ── Renderização por mês
    meses_iter = (
        sorted(set(df_view["_data_ref"].fillna("").astype(str).str[:7].tolist()))
        if (mostrar_por == "Por Mês" and not sel_mes)
        else [sel_mes or ""]
    )

    for mes_key in meses_iter:
        if mostrar_por == "Por Mês" and not sel_mes:
            if not mes_key or len(mes_key) < 7:
                continue
            df_mes = df_view[df_view["_data_ref"].fillna("").astype(str).str.startswith(mes_key)]
            lbl = f"{MESES_PT.get(mes_key[5:7], mes_key[5:7])}/{mes_key[:4]}"
            n_p = (df_mes["_classe"] == "nao").sum()
            st.markdown(
                f"<h4 style='margin:16px 0 6px 0;color:#f1f5f9'>📅 {lbl} "
                f"<span style='font-size:0.8rem;color:#94a3b8'>"
                f"{len(df_mes):,} lançamentos</span></h4>",
                unsafe_allow_html=True
            )
        else:
            df_mes = df_view
            lbl = "Todos"

        df_exato  = df_mes[df_mes["_classe"] == "exato"]
        df_cond   = df_mes[df_mes["_classe"] == "cond"]
        df_checar = df_mes[df_mes["_classe"] == "checar"]
        df_nao    = df_mes[df_mes["_classe"] == "nao"]

        sfx = mes_key.replace("-", "") if mes_key else "all"

        # ══ BOX 1: CONCILIADOS 100% ══════════════════════════════════════
        with st.expander(
            f"🟢 Conciliados 100% — {len(df_exato):,} lançamentos",
            expanded=False
        ):
            if df_exato.empty:
                st.caption("Nenhum lançamento conciliado com exatidão neste período.")
            else:
                df_ex = _formatar(df_exato)
                # Totalizador
                v_razao = pd.to_numeric(
                    df_exato["valor_razao"].fillna(0), errors="coerce"
                ).abs().sum().round(2)
                v_ext = pd.to_numeric(
                    df_exato["valor_extrato"].fillna(0), errors="coerce"
                ).abs().sum().round(2)
                dif = round(abs(v_razao - v_ext), 2)
                c1, c2, c3 = st.columns(3)
                c1.metric("∑ Razão", fmt_moeda(v_razao))
                c2.metric("∑ Extrato", fmt_moeda(v_ext))
                c3.metric("Diferença", fmt_moeda(dif),
                          delta="✅ OK" if dif < 0.01 else f"⚠️ {fmt_moeda(dif)}",
                          delta_color="off" if dif < 0.01 else "inverse")
                st.markdown(
                    "<style>.row-verde tbody tr{background-color:#14532d22!important}</style>",
                    unsafe_allow_html=True
                )
                st.dataframe(df_ex, use_container_width=True, hide_index=True,
                             height=min(400, 40 + len(df_ex) * 36))

        # ══ BOX 2: CONCILIADOS COM CONDIÇÃO ══════════════════════════════
        with st.expander(
            f"🟡 Conciliados com condição — {len(df_cond):,} para conferência",
            expanded=(len(df_cond) > 0 and len(df_nao) == 0)
        ):
            if df_cond.empty:
                st.caption("Sem lançamentos com conciliação condicional.")
            else:
                v_razao_c = pd.to_numeric(
                    df_cond["valor_razao"].fillna(0), errors="coerce"
                ).abs().sum()
                v_ext_c = pd.to_numeric(
                    df_cond["valor_extrato"].fillna(0), errors="coerce"
                ).abs().sum()
                c1, c2 = st.columns(2)
                c1.metric("∑ Razão", fmt_moeda(v_razao_c))
                c2.metric("∑ Extrato", fmt_moeda(v_ext_c))
                df_cd = _formatar(df_cond)
                st.dataframe(df_cd, use_container_width=True, hide_index=True,
                             height=min(400, 40 + len(df_cd) * 36))

        # ══ BOX 3: CHECAR — PROVÁVEIS COM DIVERGÊNCIA ═══════════════════
        with st.expander(
            f"⚠️ Checar — {len(df_checar):,} prováveis com valores divergentes",
            expanded=(len(df_checar) > 0)
        ):
            if df_checar.empty:
                st.caption("Nenhum item para checar neste período.")
            else:
                st.info(
                    "🔍 Estes lançamentos têm **correspondência provável** mas os valores divergem "
                    "mais de 5%. Confira e concilie manualmente se correto."
                )
                _box_nao_conciliados(db, empresa_id, df_checar, df_mes, f"chk_{sfx}", fixar_extrato=True)

        # ══ BOX 4: NÃO CONCILIADOS ═══════════════════════════════════════
        with st.expander(
            f"🔴 Não conciliados — {len(df_nao):,} pendentes",
            expanded=(len(df_nao) > 0)
        ):
            if df_nao.empty:
                st.success("✅ Todos os lançamentos estão conciliados neste período!")
            else:
                _box_nao_conciliados(db, empresa_id, df_nao, df_mes, sfx)
                # ── Exportar não conciliados
                st.markdown("<br>", unsafe_allow_html=True)
                cols_exp_nao = [c for c in [
                    "data_razao","historico_razao","valor_razao",
                    "data_extrato","descricao_extrato","valor_extrato","status"
                ] if c in df_nao.columns]
                df_nao_exp = df_nao[cols_exp_nao].copy()
                excel_nao = gerar_excel(df_nao_exp)
                st.download_button(
                    "📋 Exportar Não Conciliados",
                    data=excel_nao,
                    file_name=f"nao_conciliados_{empresa_id}_{mes_key or 'todos'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"btn_exp_nao_{sfx}",
                    use_container_width=True,
                )

    # ── Exportar Excel
    st.markdown("<br>", unsafe_allow_html=True)
    col_exp, _ = st.columns([2, 8])
    with col_exp:
        excel_bytes = gerar_excel(df_view.drop(columns=["_classe", "_data_ref"], errors="ignore"))
        st.download_button(
            "⬇️ Exportar Excel",
            data=excel_bytes,
            file_name=f"conciliacao_{empresa_id}_{sel_mes or 'todos'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def _box_nao_conciliados(db, empresa_id: int,
                          df_nao: pd.DataFrame,
                          df_todos: pd.DataFrame,
                          sfx: str,
                          fixar_extrato: bool = False):
    """
    Box de conciliação manual.
    fixar_extrato=True → lado Extrato mostra só as correspondências sugeridas
                         que já vêm embutidas nas linhas de df_nao (Box Checar).
    fixar_extrato=False → lado Extrato mostra todos os itens NAO_CONCILIADO do mês.
    """
    ck_r_prefix = f"ck_r_{sfx}_"
    ck_e_prefix = f"ck_e_{sfx}_"
    key_just = f"just_{sfx}"
    top_slot = st.container()

    # ── Prepara DataFrames ordenados por data
    df_r = (df_nao[df_nao["valor_razao"].notna()]
            .copy()
            .sort_values("data_razao", ascending=True)
            .reset_index(drop=True))

    if fixar_extrato and not df_r.empty:
        subset_r = [c for c in ["razao_id", "data_razao", "valor_razao", "historico_razao"] if c in df_r.columns]
        if subset_r:
            df_r = df_r.drop_duplicates(subset=subset_r).reset_index(drop=True)

    if fixar_extrato:
        # Usa só os itens de extrato já vinculados nas linhas do df_nao
        has_ext = (
            df_nao["descricao_extrato"].notna()
            if "descricao_extrato" in df_nao.columns
            else pd.Series(False, index=df_nao.index)
        )
        df_e = (
            df_nao[has_ext][[c for c in [
                "extrato_id", "data_extrato", "descricao_extrato", "valor_extrato", "status"
            ] if c in df_nao.columns]]
            .drop_duplicates()
            .sort_values("data_extrato", ascending=True)
            .reset_index(drop=True)
        )
    else:
        df_e_all = df_todos[df_todos["status"] == "NAO_CONCILIADO"].copy()
        if "descricao_extrato" in df_e_all.columns:
            df_e = (df_e_all[df_e_all["descricao_extrato"].notna()]
                    .sort_values("data_extrato", ascending=True)
                    .reset_index(drop=True))
        else:
            df_e = pd.DataFrame()

    def _stable_row_key(row: pd.Series, side: str) -> str:
        if side == "r":
            row_id = row.get("razao_id")
            if pd.notna(row_id) and str(row_id).strip():
                return str(int(row_id))
            return "|".join([
                str(row.get("data_razao", "") or ""),
                str(row.get("valor_razao", "") or ""),
                str(row.get("historico_razao", "") or ""),
            ])
        row_id = row.get("extrato_id")
        if pd.notna(row_id) and str(row_id).strip():
            return str(int(row_id))
        return "|".join([
            str(row.get("data_extrato", "") or ""),
            str(row.get("valor_extrato", "") or ""),
            str(row.get("descricao_extrato", "") or ""),
        ])

    # ── Checkboxes
    col_r, col_e = st.columns(2)
    sels_r, sels_e = [], []

    with col_r:
        st.markdown("**📒 Razão — itens pendentes**")
        if df_r.empty:
            st.caption("Sem itens do Razão pendentes.")
        else:
            for i, row in df_r.iterrows():
                row_key = _stable_row_key(row, "r")
                lbl = f"{fmt_data(row.get('data_razao',''))}  |  {fmt_moeda(row.get('valor_razao',0))}  |  {str(row.get('historico_razao',''))[:45]}"
                if st.checkbox(lbl, key=f"{ck_r_prefix}{row_key}"):
                    sels_r.append(i)

    with col_e:
        st.markdown("**🏦 Extrato — itens pendentes**")
        if df_e.empty:
            st.caption("Sem itens do Extrato pendentes.")
        else:
            for i, row in df_e.iterrows():
                row_key = _stable_row_key(row, "e")
                lbl = f"{fmt_data(row.get('data_extrato',''))}  |  {fmt_moeda(row.get('valor_extrato',0))}  |  {str(row.get('descricao_extrato',''))[:45]}"
                if st.checkbox(lbl, key=f"{ck_e_prefix}{row_key}"):
                    sels_e.append(i)

    total_r = sum(
        abs(float(df_r.loc[i, "valor_razao"]))
        for i in sels_r
        if i < len(df_r) and pd.notna(df_r.loc[i, "valor_razao"])
    )
    total_e = sum(
        abs(float(df_e.loc[i, "valor_extrato"]))
        for i in sels_e
        if i < len(df_e) and pd.notna(df_e.loc[i, "valor_extrato"])
    )
    dif = abs(total_r - total_e)
    batem = dif < 0.01

    with top_slot:
        if sels_r or sels_e:
            status_badge = "✅ Valores batem" if batem else f"⚠️ Divergência de {fmt_moeda(dif)}"
            bg_badge = "#dcfce7" if batem else "#fef2f2"
            fg_badge = "#166534" if batem else "#991b1b"

            st.markdown(
                f"""
                <div style="
                    background: linear-gradient(135deg, #0f766e 0%, #15803d 55%, #166534 100%);
                    border-radius: 22px;
                    padding: 28px 26px 22px 26px;
                    box-shadow: 0 20px 45px rgba(21, 128, 61, 0.18);
                    border: 1px solid rgba(255,255,255,0.10);
                    margin-bottom: 18px;
                ">
                    <div style="display:flex; justify-content:space-between; gap:16px; align-items:flex-start; margin-bottom:18px;">
                        <div>
                            <div style="color:#f0fdf4; font-size:21px; font-weight:800; margin-bottom:8px;">Conciliação Manual Assistida</div>
                            <div style="color:rgba(240,253,244,0.88); font-size:14px;">Revise os lançamentos selecionados, compare os totais e finalize a vinculação aqui mesmo.</div>
                        </div>
                        <div style="background:{bg_badge}; color:{fg_badge}; padding:10px 16px; border-radius:999px; font-size:14px; font-weight:800; white-space:nowrap;">{status_badge}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            m1, m2, m3 = st.columns(3)
            m1.metric("Total Razão", fmt_moeda(total_r), f"{len(sels_r)} item(ns)")
            m2.metric("Total Extrato", fmt_moeda(total_e), f"{len(sels_e)} item(ns)")
            m3.metric("Diferença", fmt_moeda(dif), "OK" if batem else "Revisar")

            d1, d2 = st.columns(2)
            with d1:
                st.markdown("**Razão selecionado(s)**")
                for i in sels_r:
                    row = df_r.iloc[i]
                    st.markdown(
                        f"- {fmt_data(row.get('data_razao',''))} | {fmt_moeda(row.get('valor_razao',0))} | {str(row.get('historico_razao',''))[:70]}"
                    )
            with d2:
                st.markdown("**Extrato selecionado(s)**")
                for i in sels_e:
                    row = df_e.iloc[i]
                    st.markdown(
                        f"- {fmt_data(row.get('data_extrato',''))} | {fmt_moeda(row.get('valor_extrato',0))} | {str(row.get('descricao_extrato',''))[:70]}"
                    )

            justificativa = ""
            if not batem and sels_r and sels_e:
                justificativa = st.text_input(
                    "Justificativa obrigatória",
                    key=key_just,
                    placeholder="Ex: juros do mês anterior, tarifa, arredondamento..."
                )

            if sels_r and sels_e and st.button("🔗 Executar Conciliação", key=f"btn_{sfx}", type="primary", use_container_width=True):
                if not batem and not justificativa.strip():
                    st.error("Preencha a justificativa para continuar.")
                    return

                razao_ids = []
                extrato_ids = []

                for i in sels_r:
                    row_r = df_r.iloc[i]
                    rid = row_r.get("razao_id")
                    if pd.notna(rid) and int(rid) > 0:
                        razao_ids.append(int(rid))
                        continue
                    data_r = str(row_r.get("data_razao", "") or "")
                    hist_r = str(row_r.get("historico_razao", "") or "")
                    valor_r = float(row_r.get("valor_razao", 0) or 0)
                    with db._conn() as _c:
                        found = _c.execute(
                            "SELECT id FROM razao WHERE cliente_id=? AND data_lancamento=? AND ABS(valor-?)<0.01 AND historico=? LIMIT 1",
                            (empresa_id, data_r, valor_r, hist_r)
                        ).fetchone() or _c.execute(
                            "SELECT id FROM razao WHERE cliente_id=? AND data_lancamento=? AND ABS(valor-?)<0.01 LIMIT 1",
                            (empresa_id, data_r, valor_r)
                        ).fetchone()
                    if found:
                        razao_ids.append(int(found["id"]))

                for i in sels_e:
                    row_e = df_e.iloc[i]
                    eid = row_e.get("extrato_id")
                    if pd.notna(eid) and int(eid) > 0:
                        extrato_ids.append(int(eid))
                        continue
                    data_e = str(row_e.get("data_extrato", "") or "")
                    desc_e = str(row_e.get("descricao_extrato", "") or "")
                    valor_e = float(row_e.get("valor_extrato", 0) or 0)
                    with db._conn() as _c:
                        found = _c.execute(
                            "SELECT id FROM extrato WHERE cliente_id=? AND data_lancamento=? AND ABS(valor-?)<0.01 AND descricao=? LIMIT 1",
                            (empresa_id, data_e, valor_e, desc_e)
                        ).fetchone() or _c.execute(
                            "SELECT id FROM extrato WHERE cliente_id=? AND data_lancamento=? AND ABS(valor-?)<0.01 LIMIT 1",
                            (empresa_id, data_e, valor_e)
                        ).fetchone()
                    if found:
                        extrato_ids.append(int(found["id"]))

                if not razao_ids or not extrato_ids:
                    st.error(f"Não foi possível resolver os IDs. Razão: {razao_ids} | Extrato: {extrato_ids}")
                    return

                try:
                    for rid in razao_ids:
                        for eid in extrato_ids:
                            db.registrar_conciliacao_manual(
                                cliente_id=empresa_id,
                                razao_id=rid,
                                extrato_id=eid,
                                valor_razao=total_r / max(len(razao_ids), 1),
                                valor_extrato=total_e / max(len(extrato_ids), 1),
                                justificativa=justificativa if not batem else "",
                            )
                except Exception as exc:
                    st.error(f"Erro ao gravar conciliação: {exc}")
                    return

                for k in list(st.session_state.keys()):
                    if str(k).startswith(ck_r_prefix) or str(k).startswith(ck_e_prefix):
                        st.session_state.pop(k, None)
                st.session_state.pop(key_just, None)
                st.session_state.pop("resultado_df", None)
                st.session_state.pop("resultado_empresa", None)
                st.rerun()



# =============================================================================
# PÁGINA: CONCILIAÇÃO
# =============================================================================
def pagina_conciliacao(empresa_id, empresa_ativa, score_min, usar_sim, tol_dias):
    db = get_db()
    nome_emp = empresa_ativa["nome"] if empresa_ativa else None

    st.markdown('<p class="page-title">📂 Conciliação</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="page-sub">{nome_emp or "Selecione um cliente na sidebar"}</p>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    if not empresa_ativa:
        st.warning("⚠️ Selecione um cliente na sidebar para continuar.")
        return

    # ══════════════════════════════════════════════════════════════════════════
    # PAINEL DE CONSULTA — tabela simples igual ao Excel
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="section-header">🗂️ Resultado da Conciliação</div>', unsafe_allow_html=True)

    stats = db.estatisticas(cliente_id=empresa_id)
    total_r = stats.get("total_razao", 0)
    total_e = stats.get("total_extrato", 0)
    total_c = stats.get("total_conciliacoes", 0)
    pendentes = stats.get("pendentes_razao", max(total_r - total_c, 0))

    mc1, mc2, mc3, mc4 = st.columns(4)
    for col, cor, num, lbl in [
        (mc1, "azul",    total_r,    "📒 Razão na base"),
        (mc2, "amarelo", total_e,    "🏦 Extrato na base"),
        (mc3, "verde",   total_c,    "✅ Conciliações"),
        (mc4, "vermelho",pendentes,  "⏳ Pendentes"),
    ]:
        col.markdown(
            f'<div class="kpi-card {cor}"><div class="kpi-lbl">{lbl}</div>'
            f'<div class="kpi-num">{num:,}</div></div>',
            unsafe_allow_html=True
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # CARREGA DADOS DO BANCO (independe de session_state)
    # ══════════════════════════════════════════════════════════════════════════
    df_banco = _carregar_lancamentos_banco(db, empresa_id)

    # Usa banco como fonte primária. resultado_df só entra se o banco estiver vazio
    # (logo após upload antes de salvar) e a empresa for a mesma.
    df_sess = st.session_state.get("resultado_df")
    if (df_banco is None or df_banco.empty) and df_sess is not None and not df_sess.empty:
        emp_sess = st.session_state.get("resultado_empresa")
        df_base = df_sess if emp_sess == empresa_id else pd.DataFrame()
    else:
        df_base = df_banco

    if df_base is None or df_base.empty:
        st.info("ℹ️ Nenhum dado encontrado. Importe os arquivos de Razão e Extrato abaixo.")
    else:
        _renderizar_lancamentos_por_mes(db, empresa_id, df_base)

    # ── Limpar base
    st.markdown("<br>", unsafe_allow_html=True)
    col_lx, _ = st.columns([2, 8])
    with col_lx:
        if st.button("🗑️ Limpar base de dados deste cliente", key="btn_limpar_base"):
            st.session_state["confirmar_limpeza"] = True

    if st.session_state.get("confirmar_limpeza"):
        st.warning(
            f"⚠️ Tem certeza? Todos os lançamentos e conciliações de **{nome_emp}** "
            "serão apagados permanentemente."
        )
        c_ok, c_cancel, _ = st.columns([1, 1, 8])
        if c_ok.button("✅ Sim, limpar", key="btn_confirma_ok", type="primary"):
            resultado = db.limpar_dados_cliente(cliente_id=empresa_id)
            st.session_state.pop("confirmar_limpeza", None)
            st.session_state.pop("resultado_df", None)
            st.session_state.pop("resultado_empresa", None)
            st.success(
                f"✅ Base limpa! Removidos: "
                f"{resultado.get('razao', 0)} razão · "
                f"{resultado.get('extrato', 0)} extrato · "
                f"{resultado.get('conciliacoes', 0)} conciliações"
            )
            st.rerun()
        if c_cancel.button("❌ Cancelar", key="btn_confirma_cancel"):
            st.session_state.pop("confirmar_limpeza", None)
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Upload
    st.markdown('<div class="section-header">📁 Upload de Arquivos</div>', unsafe_allow_html=True)
    col_r, col_e = st.columns(2)
    with col_r:
        st.markdown("**📒 Razão Contábil**")
        arquivos_razao = st.file_uploader(
            "Arraste os arquivos do Razão",
            type=["xlsx","xls","csv","pdf"],
            accept_multiple_files=True,
            key="upload_razao",
            help="Aceita múltiplos arquivos. Ex: 1 razão anual ou 12 mensais.",
        )
    with col_e:
        st.markdown("**🏦 Extrato Bancário**")
        arquivos_extrato = st.file_uploader(
            "Arraste os arquivos do Extrato",
            type=["xlsx","xls","csv","pdf"],
            accept_multiple_files=True,
            key="upload_extrato",
            help="Aceita múltiplos extratos de qualquer banco suportado.",
        )

    # Botão executar
    pode_executar = bool(arquivos_razao) and bool(arquivos_extrato)
    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        executar = st.button(
            "🚀 Executar Conciliação",
            type="primary",
            disabled=not pode_executar,
            use_container_width=True,
        )
    with col_info:
        if not pode_executar:
            st.caption("Carregue pelo menos 1 arquivo de Razão e 1 de Extrato.")
        else:
            st.caption(f"✅ {len(arquivos_razao)} razão · {len(arquivos_extrato)} extrato(s) prontos")

    if executar and pode_executar:
        _executar_conciliacao(
            arquivos_razao, arquivos_extrato,
            empresa_id, empresa_ativa,
            score_min, usar_sim, tol_dias, db
        )

    # Resultado em session_state
    if "resultado_df" in st.session_state and "resultado_empresa" in st.session_state:
        if st.session_state["resultado_empresa"] == empresa_id:
            _renderizar_resultado(db, empresa_id)


def _executar_conciliacao(arqs_razao, arqs_extrato, empresa_id, empresa_ativa,
                           score_min, usar_sim, tol_dias, db):
    import tempfile

    progress = st.progress(0, "Iniciando...")

    df_razao_total   = pd.DataFrame()
    df_extrato_total = pd.DataFrame()
    erros = []

    # Metadados detectados nos arquivos de Razão
    empresa_nome_detectada = empresa_ativa["nome"] if empresa_ativa else ""
    empresa_cnpj_detectada = empresa_ativa.get("cnpj", "") if empresa_ativa else ""

    total_arqs = len(arqs_razao) + len(arqs_extrato)
    step = 0

    # ── Lê Razão
    for arq in arqs_razao:
        step += 1
        progress.progress(int(step / total_arqs * 40), f"Lendo razão: {arq.name}")
        try:
            sufixo = Path(arq.name).suffix.lower()
            with tempfile.NamedTemporaryFile(delete=False, suffix=sufixo) as tmp:
                tmp.write(arq.read())
                tmp_path = tmp.name
            parser = RazaoParser()
            df = parser.carregar(tmp_path)
            if df.empty:
                erros.append(f"⚠️ {arq.name}: nenhum lançamento encontrado.")
            else:
                df_razao_total = pd.concat([df_razao_total, df], ignore_index=True)
                # Captura metadados da empresa do primeiro arquivo com dados
                meta = parser.obter_metadados()
                if meta.get("empresa") and not empresa_nome_detectada:
                    empresa_nome_detectada = meta["empresa"]
                if meta.get("cnpj") and not empresa_cnpj_detectada:
                    empresa_cnpj_detectada = meta["cnpj"]
        except Exception as e:
            erros.append(f"❌ {arq.name}: {e}")

    # ── Auto-detecta / cria empresa a partir dos metadados do Razão
    #    Tem prioridade sobre a seleção da sidebar quando o arquivo informa CNPJ
    empresa_id_efetivo = empresa_id
    if empresa_cnpj_detectada or empresa_nome_detectada:
        try:
            id_detectado = db.get_or_create_cliente(
                nome=empresa_nome_detectada,
                cnpj=empresa_cnpj_detectada,
            )
            if id_detectado and id_detectado > 0:
                if id_detectado != empresa_id:
                    # Empresa diferente da que estava selecionada na sidebar
                    cli = db.obter_cliente(id_detectado)
                    nome_cli = cli["nome"] if cli else empresa_nome_detectada
                    st.info(
                        f"🏢 Empresa detectada automaticamente no arquivo: **{nome_cli}** "
                        f"(CNPJ: {empresa_cnpj_detectada or '—'}). "
                        f"Os dados serão vinculados a esta empresa."
                    )
                    # Atualiza a empresa ativa na sessão para refletir na sidebar
                    st.session_state["cliente_id_ativo"] = id_detectado
                    if cli:
                        st.session_state["cliente_nome_sel"] = cli["nome"]
                empresa_id_efetivo = id_detectado
        except Exception as e:
            logger.warning(f"get_or_create_cliente: {e}")

    # ── Lê Extrato
    extrato_infos = []
    for arq in arqs_extrato:
        step += 1
        progress.progress(int(step / total_arqs * 40 + 40), f"Lendo extrato: {arq.name}")
        try:
            sufixo = Path(arq.name).suffix.lower()
            with tempfile.NamedTemporaryFile(delete=False, suffix=sufixo) as tmp:
                tmp.write(arq.read())
                tmp_path = tmp.name
            parser = ExtratoParser()
            df = parser.carregar(tmp_path, nome_original=arq.name)
            if df.empty:
                erros.append(f"⚠️ {arq.name}: nenhum lançamento.")
            else:
                extrato_infos.append({
                    "df": df, "nome": arq.name,
                    "banco":   parser.banco   or Path(arq.name).stem.upper().split()[0],
                    "agencia": parser.agencia or "",
                    "conta":   parser.conta   or "",
                })
                df_extrato_total = pd.concat([df_extrato_total, df], ignore_index=True)
        except Exception as e:
            erros.append(f"❌ {arq.name}: {e}")

    for e in erros:
        st.warning(e)

    if df_razao_total.empty or df_extrato_total.empty:
        st.error("Não foi possível carregar os dados. Verifique os arquivos.")
        progress.empty()
        return

    # Deduplicação
    df_razao_total   = df_razao_total.drop_duplicates()
    df_extrato_total = df_extrato_total.drop_duplicates()

    # ── Conciliação
    progress.progress(85, "Executando conciliação...")
    try:
        engine = ConciliacaoEngine(
            score_minimo_similaridade=float(score_min),
            usar_similaridade=usar_sim,
            db_manager=db,
        )
        df_resultado = engine.conciliar(df_razao_total, df_extrato_total)
    except Exception as e:
        st.error(f"Erro na engine: {e}")
        progress.empty()
        return

    # ── Salva no banco (usando empresa_id_efetivo = detectada ou da sidebar)
    progress.progress(90, "Salvando no banco...")
    db.importar_razao(
        df_razao_total, ",".join(a.name for a in arqs_razao),
        empresa=empresa_nome_detectada,
        cnpj=empresa_cnpj_detectada,
        cliente_id=empresa_id_efetivo,
    )
    for info in extrato_infos:
        db.importar_extrato(
            info["df"], info["nome"],
            cliente_id=empresa_id_efetivo,
            banco=info["banco"],
            agencia=info["agencia"],
            conta=info["conta"],
        )

    # ── Persiste conciliações no banco
    progress.progress(95, "Salvando conciliações...")
    try:
        db.salvar_conciliacoes(df_resultado, cliente_id=empresa_id_efetivo)
    except Exception as e:
        logger.warning(f"salvar_conciliacoes: {e}")

    # ── Salva padrões ML (empresa) + padrões globais (inteligência compartilhada)
    try:
        db.registrar_padroes_batch(df_resultado)
    except Exception:
        pass
    try:
        db.registrar_padroes_globais_batch(df_resultado)
    except Exception:
        pass

    progress.progress(100, "Concluído!")
    progress.empty()

    st.session_state["resultado_df"]      = df_resultado
    st.session_state["resultado_empresa"] = empresa_id_efetivo
    st.session_state["cliente_id_ativo"]  = empresa_id_efetivo
    st.session_state["hub_ativo"]         = False
    st.success(f"✅ Conciliação finalizada — {len(df_resultado)} registros processados.")
    st.rerun()


def _badge_status(status_raw: str) -> str:
    """Retorna HTML de célula-badge colorida para uso dentro do grid CSS."""
    mapa = {
        "MATCH_APRENDIDO":   ("🧠 APRENDIDO",      "#f3e5f5", "#4a148c"),
        "CONCILIADO":        ("✅ CONCILIADO",      "#d4edda", "#155724"),
        "MATCH_COMBINADO":   ("🟡 COMBINADO",       "#fff3cd", "#856404"),
        "MATCH_PROVAVEL":    ("🔵 SIMILARIDADE",    "#cce5ff", "#004085"),
        "NAO_CONCILIADO":    ("🔴 NÃO CONCILIADO",  "#f8d7da", "#721c24"),
        "MANUAL_CONCILIADO": ("🟠 MANUAL",           "#ffe0b2", "#e65100"),
        "MANUAL_DIVERGENTE": ("⚠️ DIVERGENTE",       "#fef3c7", "#92400e"),
    }
    label, bg, fg = mapa.get(status_raw, (status_raw, "#eee", "#333"))
    return (
        f"<div style='background:{bg};color:{fg};"
        f"padding:3px 8px;border-radius:4px;"
        f"font-size:0.78rem;font-weight:700;"
        f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
        f"display:flex;align-items:center'>{label}</div>"
    )


def _celula(texto: str, bg: str, fg: str, negrito: bool = False, max_chars: int = 28) -> str:
    """Retorna HTML de célula compacta com texto truncado e tooltip completo no hover."""
    texto_str = str(texto) if texto else "—"
    truncado  = (texto_str[:max_chars] + "…") if len(texto_str) > max_chars else texto_str
    peso      = "font-weight:700;" if negrito else ""
    tooltip   = texto_str.replace("'", "&#39;")
    return (
        f"<div title='{tooltip}' style='"
        f"background:{bg};color:{fg};{peso}"
        f"padding:3px 7px;border-radius:4px;"
        f"font-size:0.82rem;line-height:1.35;"
        f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
        f"cursor:default'>{truncado}</div>"
    )


def _tooltip_match(row) -> str:
    """Monta texto de detalhe do match para exibir como expander."""
    status = str(row.get("status", ""))
    tipo   = str(row.get("tipo_match", "") or "—")
    conf   = row.get("confidence", None)
    obs    = str(row.get("observacoes", "") or "")
    dif_v  = row.get("diferenca_valor", None)
    dif_d  = row.get("diferenca_dias", None)

    conf_str = f"{conf:.1f}%" if conf is not None and str(conf) not in ("", "nan") else "—"
    dif_v_str = fmt_moeda(abs(float(dif_v))) if dif_v is not None and str(dif_v) not in ("", "nan") else "R$ 0,00"
    dif_d_str = f"{int(dif_d)} dias" if dif_d is not None and str(dif_d) not in ("", "nan") else "0 dias"

    MOTIVOS = {
        "EXACT":        "Valor e data idênticos",
        "COMBINATION":  "Soma de múltiplos lançamentos bateu",
        "SIMILARITY":   "Descrição textual similar",
        "LEARNING":     "Par aprendido por histórico anterior",
        "MANUAL":       "Confirmado manualmente pelo usuário",
    }
    motivo = MOTIVOS.get(tipo.upper(), tipo)

    if status == "NAO_CONCILIADO":
        return f"❌ **Não conciliado** — nenhum par encontrado\n\n" \
               f"- Valor Razão: {fmt_moeda(row.get('valor_razao'))}\n" \
               f"- Data Razão: {fmt_data(row.get('data_razao'))}\n" \
               f"- Histórico: {str(row.get('historico_razao',''))[:80]}"
    return (
        f"✅ **{motivo}**\n\n"
        f"- Tipo de match: `{tipo}`\n"
        f"- Confiança: **{conf_str}**\n"
        f"- Diferença de valor: {dif_v_str}\n"
        f"- Diferença de data: {dif_d_str}\n"
        + (f"- Obs: {obs}" if obs and obs != "nan" else "")
    )


def _renderizar_resultado(db, empresa_id):
    df_resultado = st.session_state["resultado_df"]

    # ── KPIs rápidos
    status_counts = df_resultado["status"].value_counts()
    conciliados  = (status_counts.get("CONCILIADO", 0)
                    + status_counts.get("MATCH_APRENDIDO", 0)
                    + status_counts.get("MATCH_COMBINADO", 0)
                    + status_counts.get("MATCH_PROVAVEL", 0))
    nao_conc     = status_counts.get("NAO_CONCILIADO", 0)
    manual       = status_counts.get("MANUAL_CONCILIADO", 0)
    divergente   = status_counts.get("MANUAL_DIVERGENTE", 0)
    total        = len(df_resultado)
    total_conc   = conciliados + manual + divergente

    # Taxa de clareza = média ponderada de confidence dos conciliados
    df_conc_rows = df_resultado[df_resultado["status"] != "NAO_CONCILIADO"]
    if not df_conc_rows.empty and "confidence" in df_conc_rows.columns:
        conf_vals = pd.to_numeric(df_conc_rows["confidence"], errors="coerce").dropna()
        taxa_clareza = conf_vals.mean() if not conf_vals.empty else 0.0
    else:
        taxa_clareza = (total_conc / total * 100) if total > 0 else 0.0

    st.markdown('<div class="section-header">📊 Resultado da Conciliação</div>', unsafe_allow_html=True)

    k1,k2,k3,k4,k5,k6 = st.columns(6)
    for col, cor, num, lbl in [
        (k1,"cinza",   total,                "Total"),
        (k2,"verde",   total_conc,           "✅ Conciliados"),
        (k3,"amarelo", manual,               "🟡 Manual"),
        (k4,"azul",    divergente,           "� Divergentes"),
        (k5,"vermelho",nao_conc,             "🔴 Pendentes"),
        (k6,"roxo",    f"{taxa_clareza:.1f}%","🎯 Clareza"),
    ]:
        with col:
            st.markdown(f"""
            <div class="kpi-card {cor}">
                <div class="kpi-lbl">{lbl}</div>
                <div class="kpi-num">{num}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Filtros
    col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1, 1, 2])
    with col_f1:
        opcoes_status = ["Todos"] + [STATUS_MAP.get(s, s) for s in df_resultado["status"].unique()]
        filtro_status = st.selectbox("Filtrar por status", opcoes_status, key="filtro_status")
    with col_f2:
        datas = df_resultado["data_razao"].dropna()
        data_ini = st.date_input("De", value=pd.Timestamp(datas.min()).date() if not datas.empty else datetime.date.today(), key="filtro_data_ini")
    with col_f3:
        data_fim = st.date_input("Até", value=pd.Timestamp(datas.max()).date() if not datas.empty else datetime.date.today(), key="filtro_data_fim")
    with col_f4:
        busca = st.text_input("🔍 Buscar descrição / histórico", key="filtro_busca")

    # Aplica filtros
    df_filtrado = df_resultado.copy()
    if filtro_status != "Todos":
        status_rev = {v: k for k, v in STATUS_MAP.items()}
        s_raw = status_rev.get(filtro_status, filtro_status)
        df_filtrado = df_filtrado[df_filtrado["status"] == s_raw]
    df_filtrado = df_filtrado[
        (pd.to_datetime(df_filtrado["data_razao"], errors="coerce") >= pd.Timestamp(data_ini)) &
        (pd.to_datetime(df_filtrado["data_razao"], errors="coerce") <= pd.Timestamp(data_fim))
    ]
    if busca:
        mask = (
            df_filtrado["historico_razao"].str.contains(busca, case=False, na=False) |
            df_filtrado["descricao_extrato"].str.contains(busca, case=False, na=False)
        )
        df_filtrado = df_filtrado[mask]

    # ════════════════════════════════════════════════════════════════
    # ABA 1: CONCILIADOS — tabela com hover de detalhes
    # ABA 2: NÃO CONCILIADOS — checkboxes + painel de seleção
    # ════════════════════════════════════════════════════════════════
    aba_conc, aba_nao, aba_div = st.tabs([
        f"✅ Conciliados ({total_conc + divergente})",
        f"🔴 Pendentes ({nao_conc})",
        f"⬇️ Exportar",
    ])

    # ─────────────────────────────────────────────────────────────
    # ABA: CONCILIADOS — layout original azul/laranja
    # ─────────────────────────────────────────────────────────────
    with aba_conc:
        COR_R_BG  = "#e8f0fe"
        COR_R_FG  = "#1a3a6b"
        COR_R_CAB = "#1a3a6b"
        COR_E_BG  = "#fff3e0"
        COR_E_FG  = "#7c3c00"
        COR_E_CAB = "#b84c00"

        df_conc = df_filtrado[df_filtrado["status"] != "NAO_CONCILIADO"].copy().reset_index(drop=True)
        if df_conc.empty:
            st.info("Nenhum item conciliado no filtro atual.")
        else:
            # Cabeçalho fixo azul/laranja
            st.markdown(f"""
            <div style='display:grid;
                grid-template-columns:0.45fr 1.7fr 3fr 1.4fr 4px 1.7fr 3fr 1.4fr 1.7fr;
                gap:4px;margin-bottom:2px;'>
              <div style='background:#37474f;color:white;padding:5px 4px;
                   border-radius:5px 0 0 5px;font-size:0.78rem;font-weight:700'></div>
              <div style='background:{COR_R_CAB};color:white;padding:5px 8px;
                   font-size:0.78rem;font-weight:700'>
                   📅 Data Razão</div>
              <div style='background:{COR_R_CAB};color:white;padding:5px 8px;
                   font-size:0.78rem;font-weight:700'>📒 Histórico (Razão)</div>
              <div style='background:{COR_R_CAB};color:white;padding:5px 8px;
                   font-size:0.78rem;font-weight:700'>💰 Valor Razão</div>
              <div style='background:#dee2e6;border-radius:2px'></div>
              <div style='background:{COR_E_CAB};color:white;padding:5px 8px;
                   font-size:0.78rem;font-weight:700'>📅 Data Extrato</div>
              <div style='background:{COR_E_CAB};color:white;padding:5px 8px;
                   font-size:0.78rem;font-weight:700'>🏦 Descrição (Extrato)</div>
              <div style='background:{COR_E_CAB};color:white;padding:5px 8px;
                   font-size:0.78rem;font-weight:700'>💰 Valor Extrato</div>
              <div style='background:#37474f;color:white;padding:5px 8px;
                   border-radius:0 5px 5px 0;font-size:0.78rem;font-weight:700'>📊 Status</div>
            </div>
            <style>
            button[aria-label="▼"], button[aria-label="▲"] {{
                background: transparent !important; border: none !important;
                box-shadow: none !important; color: #9e9e9e !important;
                font-size: 0.9rem !important; padding: 1px 5px !important;
                min-height: 0 !important; height: 26px !important;
                cursor: pointer !important; border-radius: 3px !important;
            }}
            button[aria-label="▼"]:hover, button[aria-label="▲"]:hover {{
                background: rgba(26,58,107,0.1) !important; color: #1a3a6b !important;
            }}
            button[aria-label="▲"] {{ color: #1a3a6b !important; }}
            </style>
            """, unsafe_allow_html=True)

            st.caption(f"{len(df_conc)} registros — clique em ▼ para ver detalhes do match")

            for i, row in df_conc.iterrows():
                status_raw = str(row.get("status", ""))
                chave_exp  = f"det_conc_{i}_{status_raw}"
                expandido  = st.session_state.get(chave_exp, False)

                hist_r = str(row.get("historico_razao",  "") or "—")
                desc_e = str(row.get("descricao_extrato","") or "—")
                data_r = fmt_data(row.get("data_razao"))
                data_e = fmt_data(row.get("data_extrato"))
                val_r  = fmt_moeda(row.get("valor_razao"))
                val_e  = fmt_moeda(row.get("valor_extrato"))

                c0, c1, c2, c3, csep, c4, c5, c6, c7 = st.columns(
                    [0.45, 1.7, 3.0, 1.4, 0.05, 1.7, 3.0, 1.4, 1.6]
                )
                with c0:
                    icone = "▲" if expandido else "▼"
                    if st.button(icone, key=f"btn_conc_{i}", help="Ver detalhes do match"):
                        st.session_state[chave_exp] = not expandido
                        st.rerun()
                with c1:
                    trunc = (data_r[:16] + "…") if len(data_r) > 16 else data_r
                    st.markdown(f"<div title='{data_r}' style='background:{COR_R_BG};color:{COR_R_FG};padding:3px 6px;border-radius:4px;font-size:0.82rem;white-space:nowrap;margin-bottom:1px'>{trunc}</div>", unsafe_allow_html=True)
                with c2:
                    trunc = (hist_r[:30] + "…") if len(hist_r) > 30 else hist_r
                    st.markdown(f"<div title='{hist_r.replace(chr(39),'')}' style='background:{COR_R_BG};color:{COR_R_FG};padding:3px 6px;border-radius:4px;font-size:0.82rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-bottom:1px'>{trunc}</div>", unsafe_allow_html=True)
                with c3:
                    st.markdown(f"<div style='background:{COR_R_BG};color:{COR_R_FG};font-weight:700;padding:3px 6px;border-radius:4px;font-size:0.82rem;white-space:nowrap;margin-bottom:1px'>{val_r}</div>", unsafe_allow_html=True)
                with csep:
                    st.markdown("<div style='background:#dee2e6;height:26px;border-radius:2px;margin-bottom:1px'></div>", unsafe_allow_html=True)
                with c4:
                    trunc = (data_e[:16] + "…") if len(data_e) > 16 else data_e
                    st.markdown(f"<div title='{data_e}' style='background:{COR_E_BG};color:{COR_E_FG};padding:3px 6px;border-radius:4px;font-size:0.82rem;white-space:nowrap;margin-bottom:1px'>{trunc}</div>", unsafe_allow_html=True)
                with c5:
                    trunc_e = (desc_e[:30] + "…") if len(desc_e) > 30 else desc_e
                    st.markdown(f"<div title='{desc_e.replace(chr(39),'')}' style='background:{COR_E_BG};color:{COR_E_FG};padding:3px 6px;border-radius:4px;font-size:0.82rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-bottom:1px'>{trunc_e}</div>", unsafe_allow_html=True)
                with c6:
                    st.markdown(f"<div style='background:{COR_E_BG};color:{COR_E_FG};font-weight:700;padding:3px 6px;border-radius:4px;font-size:0.82rem;white-space:nowrap;margin-bottom:1px'>{val_e}</div>", unsafe_allow_html=True)
                with c7:
                    st.markdown(_badge_status(status_raw), unsafe_allow_html=True)

                if expandido:
                    conf  = float(row.get("confidence", 0) or 0)
                    tipo  = str(row.get("tipo_match",  "") or "—")
                    obs   = str(row.get("observacoes", "") or "—")
                    doc_r = str(row.get("documento_razao",   "") or "—")
                    doc_e = str(row.get("documento_extrato", "") or "—")
                    conta = str(row.get("conta_razao", "") or "—")
                    dif_v = float(row.get("diferenca_valor", 0) or 0)
                    dif_d_raw = pd.to_numeric(row.get("diferenca_dias", 0), errors="coerce")
                    dif_d_txt = f"{int(dif_d_raw)} dias" if pd.notna(dif_d_raw) else "—"

                    borda_cor = "#ff9800" if "MANUAL" in status_raw else "#1a3a6b"
                    bg_det    = "#fff8f0" if "MANUAL" in status_raw else "#f0f4ff"

                    st.markdown(f"""
                    <div style='border-left:4px solid {borda_cor};background:{bg_det};
                         border-radius:0 6px 6px 0;padding:10px 18px;
                         margin:0 0 6px 0;font-size:0.85rem;color:#1a1a2e;line-height:1.9'>
                      <b>🔍 Como foi conciliado</b><br>
                      &nbsp;• <b>Tipo de Match:</b> {tipo} &nbsp;|&nbsp;
                        <b>Confiança:</b> {conf:.1f}%<br>
                      &nbsp;• <b>Dif. Valor:</b> {fmt_moeda(abs(dif_v))} &nbsp;|&nbsp;
                        <b>Dif. Dias:</b> {dif_d_txt}<br>
                      &nbsp;• <b>Documento Razão:</b> {doc_r} &nbsp;|&nbsp;
                        <b>Documento Extrato:</b> {doc_e}<br>
                      &nbsp;• <b>Conta:</b> {conta}<br>
                      &nbsp;• <b>Observações:</b> {obs}
                    </div>""", unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────
    # ABA: NÃO CONCILIADOS — checkboxes + painel de seleção
    # ─────────────────────────────────────────────────────────────
    with aba_nao:
        df_nao = df_filtrado[df_filtrado["status"] == "NAO_CONCILIADO"].copy().reset_index()
        if df_nao.empty:
            st.success("🎉 Todos os lançamentos estão conciliados!")
        else:
            # Inicializa estado de seleção
            if "sel_razao_idxs" not in st.session_state:
                st.session_state["sel_razao_idxs"] = []
            if "sel_ext_idxs" not in st.session_state:
                st.session_state["sel_ext_idxs"] = []

            # ── Painel superior flutuante com selecionados
            sel_r = st.session_state["sel_razao_idxs"]
            sel_e = st.session_state["sel_ext_idxs"]

            if sel_r or sel_e:
                rows_r = df_nao[df_nao["index"].isin(sel_r)]
                # Extrato: carrega do banco por ID
                if sel_e:
                    _df_ext_sel = db.consultar_extrato(cliente_id=empresa_id)
                    if not _df_ext_sel.empty:
                        rows_e = _df_ext_sel[_df_ext_sel["id"].isin(sel_e)]
                    else:
                        rows_e = pd.DataFrame()
                else:
                    rows_e = pd.DataFrame()
                soma_r = rows_r["valor_razao"].apply(lambda x: float(x) if str(x) not in ("","nan") else 0).sum()
                soma_e = rows_e["valor"].apply(lambda x: float(x) if str(x) not in ("","nan") else 0).sum() if not rows_e.empty else 0.0
                dif    = abs(soma_r) - abs(soma_e)
                valores_batem = abs(dif) < 0.02

                cor_painel = "#f0fdf4" if valores_batem else "#fef9c3"
                borda_painel = "#10b981" if valores_batem else "#f59e0b"

                st.markdown(
                    f"<div style='background:{cor_painel};border:2px solid {borda_painel};"
                    f"border-radius:10px;padding:16px 20px;margin-bottom:16px'>",
                    unsafe_allow_html=True
                )
                st.markdown("### 📋 Lançamentos Selecionados para Conciliar")
                pc1, pc2, pc3 = st.columns(3)

                with pc1:
                    st.markdown("**📒 Razão selecionados**")
                    for _, r in rows_r.iterrows():
                        st.markdown(f"- {fmt_data(r.get('data_razao'))} · {str(r.get('historico_razao',''))[:40]} · **{fmt_moeda(r.get('valor_razao'))}**")
                    st.markdown(f"**Soma Razão: {fmt_moeda(soma_r)}**")

                with pc2:
                    st.markdown("**🏦 Extrato selecionados**")
                    if not rows_e.empty:
                        for _, r in rows_e.iterrows():
                            data_e = r.get('data_lancamento') or r.get('data_extrato') or ''
                            desc_e = r.get('descricao') or r.get('descricao_extrato') or ''
                            val_e  = r.get('valor') or r.get('valor_extrato') or 0
                            st.markdown(f"- {fmt_data(data_e)} · {str(desc_e)[:40]} · **{fmt_moeda(val_e)}**")
                    st.markdown(f"**Soma Extrato: {fmt_moeda(soma_e)}**")

                with pc3:
                    if valores_batem:
                        st.success(f"✅ Valores batem!\nDiferença: {fmt_moeda(abs(dif))}")
                    else:
                        st.warning(f"⚠️ Diferença: **{fmt_moeda(abs(dif))}**\nOs valores não são iguais.")

                st.markdown("</div>", unsafe_allow_html=True)

                # Botão confirmar
                col_btn1, col_btn2 = st.columns([1, 4])
                with col_btn1:
                    confirmar = st.button("✅ Confirmar Conciliação", type="primary",
                                          key="btn_confirmar_manual", use_container_width=True)
                with col_btn2:
                    limpar = st.button("🗑️ Limpar seleção", key="btn_limpar_sel", use_container_width=False)
                    if limpar:
                        st.session_state["sel_razao_idxs"] = []
                        st.session_state["sel_ext_idxs"]   = []
                        st.rerun()

                if confirmar:
                    if not sel_r or not sel_e:
                        st.error("Selecione ao menos 1 item do Razão e 1 do Extrato.")
                    elif not valores_batem:
                        # Pede justificativa
                        st.session_state["_pendente_divergente"] = True
                        st.session_state["_pendente_sel_r"] = sel_r
                        st.session_state["_pendente_sel_e"] = sel_e
                        st.rerun()
                    else:
                        # Confirma direto
                        _aplicar_conciliacao_manual(db, empresa_id, sel_r, sel_e, df_nao, justificativa=None)

                # Formulário de justificativa para divergência
                if st.session_state.get("_pendente_divergente"):
                    st.markdown(
                        "<div style='background:#fef3c7;border:2px solid #f59e0b;"
                        "border-radius:10px;padding:16px 20px;margin-top:8px'>",
                        unsafe_allow_html=True
                    )
                    st.warning("⚠️ **Os valores não batem.** Deseja conciliar mesmo assim?")
                    just = st.text_area(
                        "Justificativa obrigatória *",
                        placeholder="Ex: Diferença de taxa bancária, estorno parcial, desconto negociado...",
                        key="input_justificativa"
                    )
                    cj1, cj2 = st.columns(2)
                    with cj1:
                        if st.button("✅ Confirmar mesmo assim", type="primary", key="btn_conf_div"):
                            if not just.strip():
                                st.error("A justificativa é obrigatória para conciliação com divergência.")
                            else:
                                _aplicar_conciliacao_manual(
                                    db, empresa_id,
                                    st.session_state["_pendente_sel_r"],
                                    st.session_state["_pendente_sel_e"],
                                    df_nao, justificativa=just.strip()
                                )
                    with cj2:
                        if st.button("❌ Cancelar", key="btn_cancel_div"):
                            st.session_state.pop("_pendente_divergente", None)
                            st.session_state.pop("_pendente_sel_r", None)
                            st.session_state.pop("_pendente_sel_e", None)
                            st.rerun()
                    st.markdown("</div>", unsafe_allow_html=True)

            # ── Razão não conciliado (do resultado_df)
            df_razao_nao = df_nao[
                df_nao["historico_razao"].notna() &
                (df_nao["historico_razao"].astype(str).str.strip() != "") &
                (df_nao["historico_razao"].astype(str) != "-")
            ].copy()

            # ── Extrato pendente — busca direto do banco excluindo já conciliados
            df_ext_banco = db.consultar_extrato_pendente(cliente_id=empresa_id)

            # ── Calcula valor de referência (soma dos Razão selecionados)
            valor_ref = None
            if st.session_state["sel_razao_idxs"]:
                rows_sel = df_razao_nao[df_razao_nao["index"].isin(st.session_state["sel_razao_idxs"])]
                if not rows_sel.empty:
                    valor_ref = rows_sel["valor_razao"].apply(
                        lambda x: float(x) if str(x) not in ("", "nan") else 0
                    ).sum()

            # ── Ordena extrato por proximidade ao valor de referência
            if valor_ref is not None and not df_ext_banco.empty:
                df_ext_banco = df_ext_banco.copy()
                df_ext_banco["_dist"] = (df_ext_banco["valor"].apply(
                    lambda x: abs(abs(float(x or 0)) - abs(valor_ref))
                ))
                df_ext_banco = df_ext_banco.sort_values("_dist").drop(columns=["_dist"])
            elif not df_ext_banco.empty:
                df_ext_banco = df_ext_banco.sort_values("data_lancamento")

            # ── Dois painéis lado a lado com checkboxes
            st.markdown("**Marque os lançamentos de cada lado que deseja conciliar:**")
            if valor_ref is not None:
                st.caption(f"🎯 Extrato ordenado por proximidade ao valor selecionado: **{fmt_moeda(valor_ref)}**")

            col_razao, col_ext = st.columns(2)

            with col_razao:
                st.markdown(f"**📒 Razão — {len(df_razao_nao)} lançamentos**")
                for _, row in df_razao_nao.iterrows():
                    orig_idx = row["index"]
                    chave = f"ck_r_{orig_idx}"
                    marcado = orig_idx in st.session_state["sel_razao_idxs"]
                    label = f"{fmt_data(row.get('data_razao'))} | {str(row.get('historico_razao',''))[:36]} | **{fmt_moeda(row.get('valor_razao'))}**"
                    novo = st.checkbox(label, value=marcado, key=chave)
                    if novo and orig_idx not in st.session_state["sel_razao_idxs"]:
                        st.session_state["sel_razao_idxs"].append(orig_idx)
                    elif not novo and orig_idx in st.session_state["sel_razao_idxs"]:
                        st.session_state["sel_razao_idxs"].remove(orig_idx)

            with col_ext:
                n_ext = len(df_ext_banco) if not df_ext_banco.empty else 0
                st.markdown(f"**🏦 Extrato — {n_ext} lançamentos pendentes**")
                if df_ext_banco.empty:
                    st.info("Nenhum lançamento de extrato pendente.")
                else:
                    for _, row in df_ext_banco.head(300).iterrows():
                        ext_id = int(row["id"])
                        chave = f"ck_e_{ext_id}"
                        marcado = ext_id in st.session_state["sel_ext_idxs"]
                        desc = str(row.get("descricao") or row.get("descricao_extrato") or "")
                        val  = row.get("valor") or row.get("valor_extrato") or 0
                        data = row.get("data_lancamento") or row.get("data_extrato") or ""
                        label = f"{fmt_data(data)} | {desc[:36] or '—'} | **{fmt_moeda(val)}**"
                        novo = st.checkbox(label, value=marcado, key=chave)
                        if novo and ext_id not in st.session_state["sel_ext_idxs"]:
                            st.session_state["sel_ext_idxs"].append(ext_id)
                        elif not novo and ext_id in st.session_state["sel_ext_idxs"]:
                            st.session_state["sel_ext_idxs"].remove(ext_id)

    # ─────────────────────────────────────────────────────────────
    # ABA: EXPORTAR
    # ─────────────────────────────────────────────────────────────
    with aba_div:
        st.markdown("**⬇️ Exportar resultado completo**")
        df_show = preparar_df_exibicao(df_filtrado)
        excel_bytes = gerar_excel(df_show)
        st.download_button(
            "⬇️ Baixar Excel",
            data=excel_bytes,
            file_name=f"conciliacao_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )
        st.caption(f"Exportando {len(df_filtrado)} registros do filtro atual.")
        st.dataframe(preparar_df_exibicao(df_filtrado), use_container_width=True, hide_index=True, height=400)


def _aplicar_conciliacao_manual(db, empresa_id, sel_r, sel_e, df_nao, justificativa=None):
    """
    Aplica conciliação manual.
    sel_r = lista de índices do resultado_df (Razão)
    sel_e = lista de IDs do banco extrato
    """
    divergente = justificativa is not None
    novo_status = "MANUAL_DIVERGENTE" if divergente else "MANUAL_CONCILIADO"
    obs = f"DIVERGENTE — Justificativa: {justificativa}" if divergente else "Conciliado manualmente"

    # Atualiza resultado_df para o Razão selecionado
    for idx in sel_r:
        try:
            st.session_state["resultado_df"].at[idx, "status"]      = novo_status
            st.session_state["resultado_df"].at[idx, "tipo_match"]  = "MANUAL"
            st.session_state["resultado_df"].at[idx, "observacoes"] = obs
        except Exception:
            pass

    # Grava diretamente no banco: razao_id × extrato_id
    try:
        import hashlib as _hl
        df_res = st.session_state["resultado_df"]

        # Resolve razao_id para cada índice sel_r
        razao_ids = []
        for idx in sel_r:
            row = df_res.loc[idx]
            data_r  = str(row.get("data_razao", "") or "")
            hist_r  = str(row.get("historico_razao", "") or "")
            valor_r = float(row.get("valor_razao", 0) or 0)
            h = _hl.md5(f"{data_r}|{hist_r}|{valor_r}".encode()).hexdigest()
            with db._conn() as _c:
                r = _c.execute("SELECT id FROM razao WHERE hash_linha=? LIMIT 1", (h,)).fetchone()
                if not r:
                    r = _c.execute(
                        "SELECT id FROM razao WHERE data_lancamento=? AND historico=? AND ABS(valor-?)<0.01 LIMIT 1",
                        (data_r, hist_r, valor_r)
                    ).fetchone()
                if r:
                    razao_ids.append(r["id"])

        # Para cada par razao_id × extrato_id, insere conciliação
        with db._conn() as _c:
            for rid in razao_ids:
                for eid in sel_e:
                    existe = _c.execute(
                        "SELECT id FROM conciliacoes WHERE razao_id=? AND extrato_id=?", (rid, eid)
                    ).fetchone()
                    if not existe:
                        _c.execute(
                            """INSERT INTO conciliacoes
                               (razao_id, extrato_id, status, tipo_match, confidence,
                                observacoes, cliente_id, conciliado_por)
                               VALUES (?,?,?,?,?,?,?,?)""",
                            (rid, eid, novo_status, "MANUAL", 100.0, obs, empresa_id, "MANUAL")
                        )

        # Aprende pares histórico × descrição
        for idx in sel_r:
            hist = str(df_res.at[idx, "historico_razao"] or "")
            for eid in sel_e:
                with db._conn() as _c:
                    erow = _c.execute("SELECT descricao FROM extrato WHERE id=?", (eid,)).fetchone()
                desc = str(erow["descricao"] if erow else "")
                if hist and desc:
                    db.registrar_padrao(hist, desc, "MANUAL", "MANUAL")

    except Exception:
        pass

    # Limpa estado
    st.session_state["sel_razao_idxs"] = []
    st.session_state["sel_ext_idxs"]   = []
    st.session_state.pop("_pendente_divergente", None)
    st.session_state.pop("_pendente_sel_r", None)
    st.session_state.pop("_pendente_sel_e", None)
    st.session_state.pop("resultado_df", None)
    st.session_state.pop("resultado_empresa", None)

    # Limpa seleções do box novo para evitar painel e checkboxes "grudados"
    for k in list(st.session_state.keys()):
        if (
            str(k).startswith("sel_razao_")
            or str(k).startswith("sel_ext_")
            or str(k).startswith("sel_sig_")
            or str(k).startswith("ck_r_")
            or str(k).startswith("ck_e_")
            or str(k).startswith("just_top_")
        ):
            st.session_state.pop(k, None)

    if divergente:
        st.warning(f"⚠️ Conciliado com divergência registrada. Justificativa salva.")
    else:
        st.success(f"✅ {len(sel_r) + len(sel_e)} lançamentos conciliados manualmente!")
    st.rerun()


# =============================================================================
# TELA INICIAL — HUB DE ENTRADA
# =============================================================================
def pagina_hub():
    """
    Tela inicial exibida quando nenhuma empresa está ativa na sessão.
    Opção A: selecionar empresa existente.
    Opção B: importar arquivos (Razão + Extrato) para detectar/criar empresa e rodar conciliação.
    """
    db = get_db()

    # ── Cabeçalho
    st.markdown("""
    <div style="text-align:center;padding:48px 0 8px 0;">
        <div style="font-size:2.2rem;font-weight:800;color:#0f172a;letter-spacing:-1px;">
            Sistema de Conciliação
        </div>
        <div style="font-size:1rem;color:#64748b;margin-top:6px;">Gestão Financeira Inteligente</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # Filtra clientes por contabilidade (exceto IGP)
    usuario = st.session_state.get("usuario_logado", {})
    if usuario.get("perfil") == "igp":
        clientes = db.listar_clientes()
    else:
        empresas_permitidas = st.session_state.get("empresas_permitidas", [])
        clientes = empresas_permitidas if empresas_permitidas else db.listar_clientes()

    col_a, col_sep, col_b = st.columns([5, 1, 5])

    # ─────────────────────────────────────────────────────
    # CARD A — Selecionar empresa existente
    # ─────────────────────────────────────────────────────
    with col_a:
        st.markdown("""
        <div style="background:white;border-radius:16px;padding:28px 28px 20px 28px;
                    box-shadow:0 1px 6px rgba(0,0,0,0.08);border-top:4px solid #3b82f6;
                    min-height:340px;">
            <div style="font-size:1.2rem;font-weight:700;color:#0f172a;margin-bottom:4px;">
                🏢 Selecionar Empresa
            </div>
            <div style="font-size:0.85rem;color:#64748b;margin-bottom:20px;">
                Escolha uma empresa já cadastrada para acessar o dashboard.
            </div>
        </div>
        """, unsafe_allow_html=True)

        if not clientes:
            st.info("Nenhuma empresa cadastrada ainda. Use a opção ao lado para importar.")
        else:
            opcoes = {c["nome"]: c["id"] for c in clientes}
            # Começa com seleção vazia — usuário deve escolher explicitamente
            opcoes_com_placeholder = {"": None, **opcoes}
            nome_sel = st.selectbox(
                "Empresa",
                options=list(opcoes_com_placeholder.keys()),
                index=0,
                key="hub_empresa_sel",
                label_visibility="collapsed",
                format_func=lambda x: "— Selecione uma empresa —" if x == "" else x,
            )
            cli_sel = next((c for c in clientes if c["nome"] == nome_sel), None) if nome_sel else None
            if cli_sel:
                cnpj_txt = cli_sel.get("cnpj") or "—"
                resp_txt = cli_sel.get("responsavel") or ""
                st.markdown(f"""
                <div style="background:#f8fafc;border-radius:10px;padding:12px 16px;
                            margin:8px 0 16px 0;font-size:0.85rem;color:#334155;">
                    <b>CNPJ:</b> {cnpj_txt}<br>
                    {"<b>Resp.:</b> " + resp_txt if resp_txt else ""}
                </div>
                """, unsafe_allow_html=True)

            btn_disabled = not bool(nome_sel)
            if st.button("▶️ Entrar no Dashboard", type="primary",
                         use_container_width=True, key="hub_btn_entrar",
                         disabled=btn_disabled):
                st.session_state["cliente_id_ativo"]  = opcoes[nome_sel]
                st.session_state["cliente_nome_sel"]  = nome_sel
                st.session_state["hub_ativo"]         = False
                st.rerun()

    # separador visual
    with col_sep:
        st.markdown("""
        <div style="display:flex;align-items:center;justify-content:center;
                    height:340px;color:#cbd5e1;font-size:1.4rem;font-weight:300;">ou</div>
        """, unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────
    # CARD B — Importar arquivos
    # ─────────────────────────────────────────────────────
    with col_b:
        st.markdown("""
        <div style="background:white;border-radius:16px;padding:28px 28px 20px 28px;
                    box-shadow:0 1px 6px rgba(0,0,0,0.08);border-top:4px solid #10b981;
                    min-height:340px;">
            <div style="font-size:1.2rem;font-weight:700;color:#0f172a;margin-bottom:4px;">
                📂 Importar Arquivos
            </div>
            <div style="font-size:0.85rem;color:#64748b;margin-bottom:20px;">
                Envie o Razão e o Extrato. A empresa será detectada automaticamente.
            </div>
        </div>
        """, unsafe_allow_html=True)

        arqs_razao   = st.file_uploader(
            "📒 Razão (xlsx / xls / csv / pdf)",
            type=["xlsx", "xls", "csv", "pdf"],
            accept_multiple_files=True,
            key="hub_razao",
        )
        arqs_extrato = st.file_uploader(
            "🏦 Extrato (xlsx / xls / csv / pdf)",
            type=["xlsx", "xls", "csv", "pdf"],
            accept_multiple_files=True,
            key="hub_extrato",
        )

        pode_importar = bool(arqs_razao and arqs_extrato)
        if st.button("🚀 Importar e Conciliar", type="primary",
                     use_container_width=True, key="hub_btn_importar",
                     disabled=not pode_importar):
            # Recupera parâmetros padrão (sem sidebar aberta ainda)
            score_min = 80
            usar_sim  = True
            tol_dias  = 3
            _executar_conciliacao(
                arqs_razao, arqs_extrato,
                empresa_id=None, empresa_ativa=None,
                score_min=score_min, usar_sim=usar_sim, tol_dias=tol_dias,
                db=db,
            )
            # _executar_conciliacao já seta cliente_id_ativo e faz st.rerun()

        if not pode_importar:
            st.caption("Envie ao menos um arquivo de Razão e um de Extrato para continuar.")

    # ── Rodapé
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("""
    <div class="section-header">📘 Novo Módulo</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div style="background:white;border-radius:16px;padding:26px 28px 22px 28px;
                box-shadow:0 1px 6px rgba(0,0,0,0.08);border-top:4px solid #10b981;">
        <div style="font-size:1.18rem;font-weight:700;color:#0f172a;margin-bottom:6px;">
            📗 Conciliação Razão
        </div>
        <div style="font-size:0.92rem;color:#475569;margin-bottom:8px;">
            Novo módulo para interpretar o arquivo Razão e trabalhar conciliações de contas contábeis específicas.
        </div>
        <div style="font-size:0.84rem;color:#64748b;">
            Nesta primeira etapa, deixei a entrada do módulo pronta para receber um arquivo de Razão e abrir uma tela dedicada.
        </div>
    </div>
    """, unsafe_allow_html=True)

    arq_razao_mod = st.file_uploader(
        "📒 Arquivo do Razão para o novo módulo",
        type=["xlsx", "xls", "csv", "pdf"],
        accept_multiple_files=False,
        key="hub_razao_modulo",
    )

    if arq_razao_mod is not None:
        st.markdown(f"""
        <div style="background:#f8fafc;border-radius:10px;padding:12px 16px;
                    margin:8px 0 12px 0;font-size:0.85rem;color:#334155;">
            <b>Arquivo selecionado:</b> {arq_razao_mod.name}<br>
            <b>Tamanho:</b> {len(arq_razao_mod.getvalue()):,} bytes
        </div>
        """, unsafe_allow_html=True)

    if st.button(
        "▶️ Abrir módulo Conciliação Razão",
        type="primary",
        use_container_width=True,
        key="hub_btn_conc_razao",
    ):
        st.session_state["hub_modulo"] = "conciliacao_razao"
        st.rerun()

    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown(
        "<div style='text-align:center;color:#94a3b8;font-size:0.78rem;'>"
        "v2.0 · Python + Streamlit</div>",
        unsafe_allow_html=True
    )


def _fmt_dc(debito: float, credito: float) -> str:
    """Retorna 'D' se débito, 'C' se crédito, '' se ambos zero."""
    if debito and debito > 0:
        return "D"
    if credito and credito > 0:
        return "C"
    return ""


def _fmt_valor_razao(debito: float, credito: float) -> str:
    """Formata o valor com sinal + letra D/C."""
    v = debito if debito else credito
    if not v:
        return "—"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def pagina_conciliacao_razao():
    db = get_db()

    st.markdown('<p class="page-title">📗 Razão Contábil</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="page-sub">Visualização por Conta Contábil → Mês → Lançamentos</p>',
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Barra superior: Voltar | Limpar
    col_voltar, col_info, col_limpar = st.columns([1, 4, 1])
    with col_voltar:
        if st.button("⬅️ Voltar ao Hub", use_container_width=True, key="btn_voltar_hub_razao"):
            st.session_state["hub_modulo"] = None
            st.session_state["hub_ativo"] = True
            st.session_state.pop("razao_contabil_df", None)
            st.session_state.pop("razao_contabil_meta", None)
            st.session_state.pop("razao_contabil_cliente_id", None)
            st.rerun()

    # Mostra empresa carregada (se houver)
    meta_atual = st.session_state.get("razao_contabil_meta", {})
    with col_info:
        if meta_atual.get("empresa"):
            st.markdown(
                f"<div style='padding:8px 12px;background:#f1f5f9;border-radius:8px;"
                f"font-size:0.9rem;color:#334155;'>"
                f"<b>{meta_atual['empresa']}</b>"
                f"{'  ·  CNPJ: ' + meta_atual['cnpj'] if meta_atual.get('cnpj') else ''}"
                f"</div>",
                unsafe_allow_html=True,
            )

    with col_limpar:
        # Resolve cliente_id mesmo que não esteja na sessão (ex: após recarregamento da página)
        cliente_id_atual = st.session_state.get("razao_contabil_cliente_id")
        if not cliente_id_atual:
            with db._conn() as _c:
                _row = _c.execute(
                    "SELECT cliente_id FROM importacoes WHERE tipo='RAZAO_CONTABIL' ORDER BY importado_em DESC LIMIT 1"
                ).fetchone()
            if _row:
                cliente_id_atual = _row["cliente_id"]

        if st.button("🗑️ Limpar dados", use_container_width=True,
                     key="btn_limpar_razao_contabil", type="secondary"):
            if cliente_id_atual:
                db.limpar_razao_contabil(cliente_id_atual)
            st.session_state.pop("razao_contabil_df", None)
            st.session_state.pop("razao_contabil_meta", None)
            st.session_state.pop("razao_contabil_cliente_id", None)
            st.session_state.pop("razao_contabil_nome", None)
            st.success("Dados limpos.")
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Seletor de empresa com dados de Razão Contábil no banco
    # Filtra por contabilidade (exceto IGP)
    usuario = st.session_state.get("usuario_logado", {})
    empresas_permitidas = st.session_state.get("empresas_permitidas", [])
    ids_permitidos = [e["id"] for e in empresas_permitidas] if empresas_permitidas else None
    
    with db._conn() as _c:
        if usuario.get("perfil") == "igp" or not ids_permitidos:
            # IGP vê todas
            _empresas_db = _c.execute(
                """SELECT DISTINCT i.cliente_id, i.empresa, i.cnpj,
                          MAX(i.importado_em) as ultima_imp
                   FROM importacoes i
                   WHERE i.tipo='RAZAO_CONTABIL'
                   GROUP BY i.cliente_id
                   ORDER BY i.empresa"""
            ).fetchall()
        else:
            # Outros usuários veem apenas suas empresas
            placeholders = ','.join('?' * len(ids_permitidos))
            _empresas_db = _c.execute(
                f"""SELECT DISTINCT i.cliente_id, i.empresa, i.cnpj,
                          MAX(i.importado_em) as ultima_imp
                   FROM importacoes i
                   WHERE i.tipo='RAZAO_CONTABIL'
                     AND i.cliente_id IN ({placeholders})
                   GROUP BY i.cliente_id
                   ORDER BY i.empresa""",
                ids_permitidos
            ).fetchall()

    if _empresas_db:
        _emp_opts = {f"{r['empresa']} ({r['cnpj'] or 'sem CNPJ'})": r["cliente_id"]
                     for r in _empresas_db}
        _cid_atual = st.session_state.get("razao_contabil_cliente_id")
        # Determina índice padrão baseado no cliente atual da sessão
        _idx_default = 0
        if _cid_atual:
            _ids_lista = list(_emp_opts.values())
            if _cid_atual in _ids_lista:
                _idx_default = _ids_lista.index(_cid_atual)

        col_emp_sel, _ = st.columns([3, 7])
        with col_emp_sel:
            _sel_emp_label = st.selectbox(
                "🏢 Empresa",
                options=list(_emp_opts.keys()),
                index=_idx_default,
                key="razao_empresa_sel",
            )
        _sel_cid = _emp_opts[_sel_emp_label]

        # Se empresa mudou → recarrega dados do banco
        if _sel_cid != _cid_atual:
            df_banco = db.carregar_razao_contabil(_sel_cid)
            _row_meta = next((r for r in _empresas_db if r["cliente_id"] == _sel_cid), None)
            st.session_state["razao_contabil_df"]         = df_banco
            st.session_state["razao_contabil_meta"]       = {
                "empresa": _row_meta["empresa"] if _row_meta else "",
                "cnpj":    _row_meta["cnpj"] if _row_meta else "",
                "periodo": "",
            }
            st.session_state["razao_contabil_cliente_id"] = _sel_cid
            st.session_state.pop("razao_contabil_nome", None)
            st.rerun()

    # ── Carrega do banco na primeira abertura (sessão sem dados em memória)
    if "razao_contabil_df" not in st.session_state:
        with db._conn() as _c:
            row = _c.execute(
                """SELECT i.cliente_id, i.empresa, i.cnpj
                   FROM importacoes i
                   WHERE i.tipo='RAZAO_CONTABIL'
                   ORDER BY i.importado_em DESC LIMIT 1"""
            ).fetchone()
        if row:
            cid = row["cliente_id"]
            df_banco = db.carregar_razao_contabil(cid)
            if not df_banco.empty:
                st.session_state["razao_contabil_df"]        = df_banco
                st.session_state["razao_contabil_meta"]      = {
                    "empresa": row["empresa"] or "",
                    "cnpj":    row["cnpj"] or "",
                    "periodo": "",
                }
                st.session_state["razao_contabil_cliente_id"] = cid

    # ── Upload
    arq = st.file_uploader(
        "📒 Envie o arquivo Razão para importar (CSV, XLSX ou PDF)",
        type=["xlsx", "xls", "csv", "pdf"],
        accept_multiple_files=False,
        key="pagina_razao_upload",
    )

    if arq is not None:
        if st.session_state.get("razao_contabil_nome") != arq.name:
            with st.spinner("Processando arquivo..."):
                import tempfile as _tf
                suffix = Path(arq.name).suffix
                with _tf.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(arq.read())
                    tmp_path = tmp.name
                try:
                    parser = RazaoParser()
                    df_raw = parser.carregar(tmp_path)
                    meta_raw = parser.obter_metadados()
                    # Resolve cliente pelo nome/CNPJ do arquivo
                    emp = meta_raw.get("empresa", "") or "SEM NOME"
                    cnpj = meta_raw.get("cnpj", "") or ""
                    cliente_id_sel = db.get_or_create_cliente(emp, cnpj)
                    # Grava no banco
                    n = db.salvar_razao_contabil(
                        cliente_id=cliente_id_sel,
                        df=df_raw,
                        empresa=emp,
                        cnpj=cnpj,
                        periodo=meta_raw.get("periodo", ""),
                        nome_arquivo=arq.name,
                    )
                    st.session_state["razao_contabil_df"]         = df_raw
                    st.session_state["razao_contabil_meta"]       = meta_raw
                    st.session_state["razao_contabil_nome"]       = arq.name
                    st.session_state["razao_contabil_cliente_id"] = cliente_id_sel
                    st.success(f"✅ {n:,} lançamentos importados e salvos para **{emp}**.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Erro ao processar arquivo: {exc}")
                    return

    df_razao = st.session_state.get("razao_contabil_df")
    meta     = st.session_state.get("razao_contabil_meta", {})

    if df_razao is None or df_razao.empty:
        st.info("Envie um arquivo acima para visualizar os lançamentos.")
        return

    # ── Carrega conciliações manuais + bancárias do banco (uma vez por cliente/sessão)
    _cid_ativo = st.session_state.get("razao_contabil_cliente_id")
    _manuais_load_key = f"razao_manuais_loaded_{_cid_ativo}"
    if _cid_ativo and _manuais_load_key not in st.session_state:
        try:
            _db_tmp = DatabaseManager()
            # Conciliações manuais do módulo Razão
            _manuais_db = _db_tmp.carregar_conciliacoes_manuais(_cid_ativo)
            _manuais_sess = st.session_state.get("razao_manuais", {})
            for _k, _v in _manuais_db.items():
                _manuais_sess.setdefault(_k, set()).update(_v)
            st.session_state["razao_manuais"] = _manuais_sess
            # IDs conciliados no módulo bancário
            _banc_ids = _db_tmp.carregar_razao_ids_conciliados_bancario(_cid_ativo)
            st.session_state["razao_bancario_concil_ids"] = _banc_ids
        except Exception:
            st.session_state["razao_bancario_concil_ids"] = set()
        st.session_state[_manuais_load_key] = True

    # ── Placeholder para o cabeçalho (será preenchido após calcular _contas_info)
    _hdr_placeholder = st.empty()

    # ── Filtro de grupo de contas
    grupos_disponiveis = sorted(
        df_razao["conta_codigo"].dropna()
        .astype(str)
        .str.split(".")
        .str[0]
        .unique()
        .tolist()
    )
    col_f1, col_f2 = st.columns([2, 5])
    with col_f1:
        grupo_sel = st.multiselect(
            "Filtrar por grupo contábil",
            options=grupos_disponiveis,
            default=[],
            placeholder="Todos os grupos",
            key="razao_grupo_sel",
        )
    with col_f2:
        busca_conta = st.text_input(
            "Buscar conta (código ou nome)",
            value="",
            placeholder="Ex: 2.1 ou Salários",
            key="razao_busca_conta",
        )

    # ── Botões de filtro por status de conciliação
    _filtro_status = st.session_state.get("razao_filtro_status", None)

    # Estilos dos botões
    _btn_base = "padding:8px 18px;border-radius:8px;font-weight:700;font-size:0.88rem;cursor:pointer;border:none;box-shadow:0 2px 6px rgba(0,0,0,0.18);transition:opacity 0.15s;"
    _btn_todos  = _btn_base + ("background:#3b82f6;color:#fff;opacity:1;" if _filtro_status is None else "background:#e2e8f0;color:#475569;opacity:0.85;")
    _btn_parc   = _btn_base + ("background:#f59e0b;color:#fff;opacity:1;" if _filtro_status == "parcial" else "background:#fef3c7;color:#92400e;opacity:0.85;")
    _btn_nao    = _btn_base + ("background:#ef4444;color:#fff;opacity:1;" if _filtro_status == "nao" else "background:#fee2e2;color:#991b1b;opacity:0.85;")

    _col_esp, col_btn0, col_btn1, col_btn2 = st.columns([5, 1.2, 2.2, 2])
    with col_btn0:
        if st.button("🔵 Todos", key="btn_filtro_todos", use_container_width=True):
            st.session_state["razao_filtro_status"] = None
            st.rerun()
    with col_btn1:
        if st.button("⚠️ Parcialmente conciliado", key="btn_filtro_parcial", use_container_width=True):
            st.session_state["razao_filtro_status"] = None if _filtro_status == "parcial" else "parcial"
            st.rerun()
    with col_btn2:
        if st.button("❌ Não conciliados", key="btn_filtro_nao", use_container_width=True):
            st.session_state["razao_filtro_status"] = None if _filtro_status == "nao" else "nao"
            st.rerun()

    # CSS para colorir os botões corretamente via nth-child não disponível no Streamlit,
    # então usamos markdown injetado com IDs únicos
    _active_todos = "background-color:#3b82f6!important;color:#fff!important;" if _filtro_status is None else "background-color:#e2e8f0!important;color:#475569!important;"
    _active_parc  = "background-color:#f59e0b!important;color:#fff!important;" if _filtro_status == "parcial" else "background-color:#fef3c7!important;color:#92400e!important;"
    _active_nao   = "background-color:#ef4444!important;color:#fff!important;" if _filtro_status == "nao" else "background-color:#fee2e2!important;color:#991b1b!important;"
    st.markdown(f"""
    <style>
    div[data-testid="stButton"] button[kind="secondary"] {{
        box-shadow: 0 2px 6px rgba(0,0,0,0.15);
        border-radius: 8px;
        font-weight: 700;
    }}
    div[data-testid="column"]:nth-child(1) button {{
        {_active_todos}
        box-shadow: 0 2px 8px rgba(59,130,246,0.3);
    }}
    div[data-testid="column"]:nth-child(2) button {{
        {_active_parc}
        box-shadow: 0 2px 8px rgba(245,158,11,0.3);
    }}
    div[data-testid="column"]:nth-child(3) button {{
        {_active_nao}
        box-shadow: 0 2px 8px rgba(239,68,68,0.3);
    }}
    </style>
    """, unsafe_allow_html=True)

    # Aplica filtros
    df_view = df_razao.copy()
    if grupo_sel:
        df_view = df_view[
            df_view["conta_codigo"].astype(str).str.split(".").str[0].isin(grupo_sel)
        ]
    if busca_conta.strip():
        q = busca_conta.strip().upper()
        df_view = df_view[
            df_view["conta_codigo"].astype(str).str.upper().str.contains(q, na=False)
            | df_view["conta_nome"].astype(str).str.upper().str.contains(q, na=False)
        ]

    if df_view.empty:
        st.warning("Nenhuma conta encontrada com os filtros aplicados.")
        return

    # ── Hierarquia: Conta → Lançamentos (sem separação por mês)
    # Chave: conta_codigo + conta_reduzida para evitar triplicação
    _cols_chave = ["conta_codigo", "conta_nome"]
    if "conta_reduzida" in df_view.columns:
        _cols_chave = ["conta_codigo", "conta_reduzida", "conta_nome"]
    contas = (
        df_view[_cols_chave]
        .drop_duplicates()
        .sort_values(_cols_chave)
    )

    def _fmt_brl(v: float, dc: str = "") -> str:
        s = f"R$ {abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        return s + (f" {dc}" if dc else "")

    def _dc(v: float) -> str:
        """Retorna 'd' para devedor (negativo) ou 'c' para credor (positivo/zero)."""
        return "d" if v < 0 else "c"

    TOLERANCIA = 0.02

    # ── Passagem 1: calcula status de cada conta (para filtro global)
    def _calcular_status_conta(df_c):
        """Retorna (conta_concil, algum_concil, matched_idx, match_info, df_reset)"""
        df_r = df_c.reset_index(drop=True)
        _sa_raw = df_c["saldo_anterior"].iloc[0] if "saldo_anterior" in df_c.columns else None
        _sa_known = _sa_raw is not None and not (isinstance(_sa_raw, float) and pd.isna(_sa_raw))
        _sa = float(_sa_raw) if _sa_known else 0.0
        _mi: dict = {}
        lncs = []
        for _i, _r in df_r.iterrows():
            _deb = float(_r.get("debito", 0) or 0)
            _cre = float(_r.get("credito", 0) or 0)
            _hist = str(_r.get("historico", "") or "")
            _data = _r["data_razao"].strftime("%d/%m/%Y") if pd.notna(_r.get("data_razao")) else "?"
            lncs.append({"idx": _i, "num": _i + 1, "deb": _deb, "cre": _cre,
                         "matched": False, "hist": _hist, "data": _data})
        if abs(_sa) > TOLERANCIA:
            _sa_desc = f"Saldo Anterior {_fmt_brl(abs(_sa), _dc(_sa))}"
            if _sa > 0:
                lncs.insert(0, {"idx": -1, "num": 0, "deb": 0.0, "cre": abs(_sa),
                                 "matched": False, "hist": _sa_desc, "data": "SA"})
            else:
                lncs.insert(0, {"idx": -1, "num": 0, "deb": abs(_sa), "cre": 0.0,
                                 "matched": False, "hist": _sa_desc, "data": "SA"})
        creds = [l for l in lncs if l["cre"] > TOLERANCIA]
        debs  = [l for l in lncs if l["deb"] > TOLERANCIA]
        for c in sorted(creds, key=lambda x: x["cre"]):
            for d in sorted(debs, key=lambda x: abs(x["deb"] - c["cre"])):
                if d["matched"]: continue
                if abs(d["deb"] - c["cre"]) <= TOLERANCIA:
                    c["matched"] = True; d["matched"] = True
                    if c["idx"] >= 0:
                        _pd = f"#{d['num']} — {d['hist'][:70]}" if d["idx"] >= 0 else d["hist"]
                        _mi[c["idx"]] = f"Conciliado com {_pd}  |  Débito {_fmt_brl(d['deb'])}"
                    if d["idx"] >= 0:
                        _pc = f"#{c['num']} — {c['hist'][:70]}" if c["idx"] >= 0 else c["hist"]
                        _mi[d["idx"]] = f"Conciliado com {_pc}  |  Crédito {_fmt_brl(c['cre'])}"
                    break
        mx = {l["idx"] for l in lncs if l["matched"]}
        cc = all(l["matched"] for l in lncs if l["idx"] >= 0)
        ac = any(l["matched"] for l in lncs if l["idx"] >= 0)
        # Regra B
        if not cc and "saldo_exercicio" in df_r.columns:
            _ux = df_r["saldo_exercicio"].dropna()
            if not _ux.empty and abs(float(_ux.iloc[-1])) <= TOLERANCIA:
                cc = True; ac = True; mx = set(range(len(df_r)))
                _ud = df_r["data_razao"].iloc[-1]
                _uds = _ud.strftime("%d/%m/%Y") if pd.notna(_ud) else "?"
                _tb = f"Saldo Exercício zerou no último lançamento ({_uds})"
                for _gi in range(len(df_r)):
                    if _gi not in _mi: _mi[_gi] = _tb
        # Regra C
        if not cc and "saldo_exercicio" in df_r.columns:
            _sxv = df_r["saldo_exercicio"].reset_index(drop=True)
            _gi0 = 0; _nm = set(mx)
            for _ki in range(len(_sxv)):
                _sv = _sxv.iloc[_ki]
                if pd.isna(_sv): continue
                if abs(float(_sv)) <= TOLERANCIA:
                    _zd = df_r["data_razao"].iloc[_ki]
                    _zds = _zd.strftime("%d/%m/%Y") if pd.notna(_zd) else "?"
                    _tc = f"Grupo zerado em {_zds} (Saldo Exerc. = 0)"
                    for _gi in range(_gi0, _ki + 1):
                        _nm.add(_gi)
                        if _gi not in _mi: _mi[_gi] = _tc
                    _gi0 = _ki + 1
            if len(_nm) > len(mx):
                mx = _nm; ac = True
                cc = all(i in mx for i in range(len(df_r)))
        return cc, ac, mx, _mi, df_r

    # Monta lista de contas com status pré-calculado
    _contas_info = []
    for _, conta_row in contas.iterrows():
        cod      = str(conta_row["conta_codigo"] or "")
        nome     = str(conta_row["conta_nome"] or "")
        reduzida = str(conta_row.get("conta_reduzida", "") or "")
        mask = df_view["conta_codigo"] == cod
        if reduzida and "conta_reduzida" in df_view.columns:
            mask = mask & (df_view["conta_reduzida"] == reduzida)
        df_conta = df_view[mask].copy().sort_values("data_razao")
        cc, ac, mx, mi, df_r = _calcular_status_conta(df_conta)
        # ── Aplica conciliações manuais já salvas
        _mk = f"razao_manual_{cod}_{reduzida}"
        _manuais_agora = st.session_state.get("razao_manuais", {}).get(_mk, set())
        if _manuais_agora:
            mx = mx | _manuais_agora
            for _m in _manuais_agora:
                if _m not in mi:
                    mi[_m] = "Conciliado manualmente"
            ac = True
            cc = all(i in mx for i in range(len(df_r)))

        # ── Aplica conciliações do módulo bancário (cruzamento por data+valor+histórico)
        _banc_chaves = st.session_state.get("razao_bancario_concil_ids", set())
        if _banc_chaves:
            for _bi, _brow in df_r.iterrows():
                if _bi in mx:
                    continue
                _bdata = _brow["data_razao"].strftime("%Y-%m-%d") if pd.notna(_brow.get("data_razao")) else ""
                _bval  = round(float(_brow.get("valor_razao", 0) or 0), 2)
                _bhist = str(_brow.get("historico", "") or "").lower().strip()
                if (_bdata, _bval, _bhist) in _banc_chaves:
                    mx.add(_bi)
                    mi[_bi] = "Conciliado no módulo bancário"
            ac = len(mx) > 0
            cc = all(i in mx for i in range(len(df_r)))
        _status = "concil" if cc else ("parcial" if ac else "nao")
        _contas_info.append((cod, nome, reduzida, df_conta, df_r, cc, ac, mx, mi, _status))

    # ── Calcula totais reais (inclui manuais) e preenche o cabeçalho
    _t_concil = sum(1 for c in _contas_info if c[9] == "concil")
    _t_parc   = sum(1 for c in _contas_info if c[9] == "parcial")
    _t_nao    = sum(1 for c in _contas_info if c[9] == "nao")
    _t_total_contas = len(_contas_info)

    # Totais de lançamentos (linhas) por status
    _lanc_concil = sum(len(c[4]) for c in _contas_info if c[9] == "concil")
    _lanc_parc   = sum(len(c[4]) for c in _contas_info if c[9] == "parcial")
    _lanc_nao    = sum(len(c[4]) for c in _contas_info if c[9] == "nao")
    _lanc_total  = _lanc_concil + _lanc_parc + _lanc_nao

    _badge = (
        "display:inline-block;padding:7px 18px;border-radius:22px;"
        "font-size:0.95rem;font-weight:800;box-shadow:0 2px 8px rgba(0,0,0,0.12);"
        "margin:2px 4px;"
    )
    _hdr_placeholder.markdown(
        f"""
        <div style="background:white;border-radius:14px;padding:20px 28px;
                    box-shadow:0 1px 6px rgba(0,0,0,0.09);margin-bottom:20px;
                    border-left:6px solid #0f766e;">
            <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:16px;">
                <div>
                    <div style="font-size:1.15rem;font-weight:700;color:#0f172a;">
                        {meta.get('empresa','—')}
                    </div>
                    <div style="font-size:0.82rem;color:#64748b;margin-top:4px;">
                        CNPJ: {meta.get('cnpj','—')} &nbsp;·&nbsp; Período: {meta.get('periodo','—')}
                        &nbsp;·&nbsp; <b>{len(df_razao):,}</b> lançamentos totais ·
                        <b>{df_razao['conta_codigo'].nunique()}</b> contas contábeis
                    </div>
                </div>
            </div>
            <div style="margin-top:16px;border-top:1px solid #f1f5f9;padding-top:14px;
                        display:flex;flex-wrap:wrap;gap:20px;align-items:flex-start;">
                <div style="min-width:280px;">
                    <div style="font-size:0.75rem;font-weight:700;color:#94a3b8;letter-spacing:0.08em;
                                text-transform:uppercase;margin-bottom:8px;">Resumo por Conta Contábil</div>
                    <div>
                        <span style="{_badge}background:#dcfce7;color:#166534;">✅ {_t_concil} conciliadas</span>
                        <span style="{_badge}background:#fef3c7;color:#92400e;">⚠️ {_t_parc} parciais</span>
                        <span style="{_badge}background:#fee2e2;color:#991b1b;">❌ {_t_nao} não conciliadas</span>
                        <span style="{_badge}background:#f1f5f9;color:#475569;">📋 {_t_total_contas} total</span>
                    </div>
                </div>
                <div style="min-width:280px;">
                    <div style="font-size:0.75rem;font-weight:700;color:#94a3b8;letter-spacing:0.08em;
                                text-transform:uppercase;margin-bottom:8px;">Resumo por Lançamento</div>
                    <div>
                        <span style="{_badge}background:#dcfce7;color:#166534;">✅ {_lanc_concil:,} conciliados</span>
                        <span style="{_badge}background:#fef3c7;color:#92400e;">⚠️ {_lanc_parc:,} parciais</span>
                        <span style="{_badge}background:#fee2e2;color:#991b1b;">❌ {_lanc_nao:,} não conciliados</span>
                        <span style="{_badge}background:#f1f5f9;color:#475569;">📋 {_lanc_total:,} total</span>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Aplica filtro global de status
    _filtro_status = st.session_state.get("razao_filtro_status", None)
    if _filtro_status == "parcial":
        _contas_info = [c for c in _contas_info if c[9] == "parcial"]
    elif _filtro_status == "nao":
        _contas_info = [c for c in _contas_info if c[9] == "nao"]

    st.markdown(
        f"<small style='color:#64748b'>Exibindo {len(_contas_info)} contas"
        + (f" · filtro: <b>{'Parcialmente conciliado' if _filtro_status=='parcial' else 'Não conciliado'}</b>" if _filtro_status else "")
        + "</small>",
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Passagem 2: renderiza cada conta
    for cod, nome, reduzida, df_conta, df_conta_reset, conta_concil, algum_concil, matched_idx, match_info, _ in _contas_info:
        n_lanc = len(df_conta)

        # ── Saldo Anterior (para exibição no badge)
        _sa_raw = df_conta_reset["saldo_anterior"].iloc[0] if "saldo_anterior" in df_conta_reset.columns else None
        sa_known = _sa_raw is not None and not (isinstance(_sa_raw, float) and pd.isna(_sa_raw))
        saldo_ant = float(_sa_raw) if sa_known else 0.0

        # ── Cor do cabeçalho da conta
        if conta_concil:
            hdr_bg  = "#dcfce7"; hdr_txt = "#166534"; hdr_icone = "✅"
        elif algum_concil:
            hdr_bg  = "#fef9c3"; hdr_txt = "#854d0e"; hdr_icone = "⚠️"
        else:
            hdr_bg  = "#fee2e2"; hdr_txt = "#991b1b"; hdr_icone = "❌"

        sufixo_conta = " — Conciliado" if conta_concil else (" — Parcialmente conciliado" if algum_concil else " — Não conciliado")
        red_label = f" [{reduzida}]" if reduzida else ""

        expander_label = f"{hdr_icone}{red_label} {cod}  —  {nome}  ({n_lanc} lançamentos){sufixo_conta}"
        with st.expander(expander_label, expanded=False):
            # ── Botão "Filtro Não conciliados" + saldo anterior em destaque
            _toggle_key = f"toggle_nc_{cod}_{reduzida}"
            _nc_primeiro = st.session_state.get(_toggle_key, False)
            _col_esp, _col_sa, _col_btn = st.columns([4, 3, 2])
            with _col_sa:
                if sa_known:
                    _sa_display = "R$ 0,00" if saldo_ant == 0.0 else _fmt_brl(abs(saldo_ant), _dc(saldo_ant))
                    st.markdown(
                        f"<div style='padding:8px 0 4px 0;text-align:right;font-size:1rem;font-weight:600;color:#334155;'>"
                        f"Saldo Anterior:&nbsp;<span style='color:{hdr_txt};font-size:1.25rem;font-weight:800;'>{_sa_display}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
            with _col_btn:
                # Só mostra botão se há lançamentos não conciliados
                _n_nc = sum(1 for _i in range(len(df_conta_reset)) if _i not in matched_idx)
                if _n_nc > 0 and not conta_concil:
                    _btn_label = "🔴 Filtro Não conciliados" if not _nc_primeiro else "↩️ Ordem original"
                    _btn_style = (
                        "background-color:#ef4444;color:#fff;font-weight:700;"
                        "border-radius:8px;box-shadow:0 2px 8px rgba(239,68,68,0.35);border:none;"
                        if not _nc_primeiro else
                        "background-color:#e2e8f0;color:#475569;font-weight:600;"
                        "border-radius:8px;box-shadow:0 2px 4px rgba(0,0,0,0.1);border:none;"
                    )
                    st.markdown(f"<style>div[data-testid='stButton'] {{margin-top:4px}}</style>", unsafe_allow_html=True)
                    if st.button(_btn_label, key=_toggle_key + "_btn",
                                 use_container_width=True, type="secondary"):
                        st.session_state[_toggle_key] = not _nc_primeiro
                        st.rerun()

            _manual_key = f"razao_manual_{cod}_{reduzida}"

            # Ordena linhas conforme toggle: não conciliados primeiro (mantendo # original)
            _indices_render = list(range(len(df_conta_reset)))
            if _nc_primeiro:
                _nao_concil = [i for i in _indices_render if i not in matched_idx]
                _concil     = [i for i in _indices_render if i in matched_idx]
                _indices_render = _nao_concil + _concil

            # ── Tabela plana de todos os lançamentos da conta
            html = (
                "<div style='overflow-x:auto;margin-top:8px;'>"
                "<table style='width:100%;border-collapse:collapse;font-size:0.82rem;'>"
                "<thead><tr style='background:#f1f5f9;color:#475569;'>"
                "<th style='padding:6px 8px;text-align:center;width:36px;color:#94a3b8'>#</th>"
                "<th style='padding:6px 8px;text-align:left;white-space:nowrap;width:90px'>Data</th>"
                "<th style='padding:6px 8px;text-align:left;min-width:300px'>Histórico</th>"
                "<th style='padding:6px 8px;text-align:left;width:70px'>Cta. C/P</th>"
                "<th style='padding:6px 8px;text-align:right;width:110px'>Débito</th>"
                "<th style='padding:6px 8px;text-align:right;width:110px'>Crédito</th>"
                "<th style='padding:6px 8px;text-align:right;width:120px'>Saldo</th>"
                "<th style='padding:6px 8px;text-align:right;width:120px'>Saldo Exerc.</th>"
                "</tr></thead><tbody>"
            )

            for _i in _indices_render:
                r = df_conta_reset.iloc[_i]
                is_matched = _i in matched_idx
                bg_row = "#dbeafe" if is_matched else "#ef4444"
                bg_alt = "#eff6ff" if is_matched else "#ef4444"
                txt    = "#1e3a5f" if is_matched else "#ffffff"
                bg     = bg_alt if _i % 2 == 0 else bg_row

                data_str  = r["data_razao"].strftime("%d/%m/%Y") if pd.notna(r["data_razao"]) else "—"
                deb       = float(r.get("debito", 0) or 0)
                cre       = float(r.get("credito", 0) or 0)
                saldo_val = float(r.get("saldo", 0) or 0)
                saldo_ex  = float(r.get("saldo_exercicio", 0) or 0)
                hist      = str(r.get("historico", "") or "")
                cta       = str(r.get("cta_contrapartida", "") or "")

                # Tooltip: explica com o que foi conciliado (ou que não foi)
                if is_matched and _i in match_info:
                    _title = match_info[_i]
                elif is_matched:
                    _title = "Conciliado"
                else:
                    _title = "Não conciliado"

                _num_linha = _i + 1
                html += (
                    f"<tr style='background:{bg};color:{txt};' title='{_title}'>"
                    f"<td style='padding:5px 8px;text-align:center;font-size:0.75rem;color:#94a3b8;white-space:nowrap'>{_num_linha}</td>"
                    f"<td style='padding:5px 8px;white-space:nowrap'>{data_str}</td>"
                    f"<td style='padding:5px 8px'>{hist}</td>"
                    f"<td style='padding:5px 8px;white-space:nowrap'>{cta}</td>"
                    f"<td style='padding:5px 8px;text-align:right;white-space:nowrap'>{_fmt_brl(deb)}</td>"
                    f"<td style='padding:5px 8px;text-align:right;white-space:nowrap'>{_fmt_brl(cre)}</td>"
                    f"<td style='padding:5px 8px;text-align:right;white-space:nowrap'>{_fmt_brl(abs(saldo_val), _dc(saldo_val))}</td>"
                    f"<td style='padding:5px 8px;text-align:right;white-space:nowrap'>{_fmt_brl(abs(saldo_ex), _dc(saldo_ex))}</td>"
                    "</tr>"
                )

            html += "</tbody></table></div>"
            st.markdown(html, unsafe_allow_html=True)
            # ── Conciliação Manual: estrutura pronta, UI desabilitada temporariamente
            # TODO: reativar quando necessário chamando _render_conciliacao_manual()


# =============================================================================
# AUTENTICAÇÃO — LOGIN
# =============================================================================
def tela_login():
    """Tela de login — fundo azul #4da6e8, card branco centralizado, logo IGP."""
    import base64 as _b64
    from pathlib import Path as _Path

    # Carrega logo como base64
    _logo_path = _Path(__file__).parent / "logo_igp.png"
    _logo_b64 = ""
    if _logo_path.exists():
        _logo_b64 = _b64.b64encode(_logo_path.read_bytes()).decode()
    _logo_tag = (
        f'<img src="data:image/png;base64,{_logo_b64}" '
        f'style="width:72px;height:72px;object-fit:contain;border-radius:50%;" />'
        if _logo_b64 else
        '<div style="width:72px;height:72px;border-radius:50%;background:#dbeafe;'
        'display:flex;align-items:center;justify-content:center;'
        'font-weight:900;color:#1d4ed8;font-size:1rem;">igp</div>'
    )

    st.markdown("""
    <style>
    /* ── Fundo azul em toda a viewport ── */
    html, body,
    [data-testid="stAppViewContainer"],
    [data-testid="stAppViewContainer"] > section,
    [data-testid="stMain"],
    [data-testid="stMain"] > div,
    .main .block-container { background-color: #4da6e8 !important; }

    /* Esconde header, toolbar e sidebar */
    [data-testid="stHeader"],
    [data-testid="stToolbar"],
    [data-testid="stSidebar"],
    #MainMenu { display: none !important; }

    /* Padding do container */
    .main .block-container {
        padding-top: 8vh !important;
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }

    /* ── Card branco ── */
    div[data-testid="stForm"] {
        background: white !important;
        border-radius: 18px !important;
        padding: 44px 44px 32px !important;
        box-shadow: 0 8px 40px rgba(0,0,0,0.13) !important;
        border: none !important;
    }

    /* Labels negrito */
    div[data-testid="stForm"] label p {
        font-weight: 700 !important;
        color: #1e293b !important;
        font-size: 0.88rem !important;
    }

    /* Inputs */
    div[data-testid="stForm"] input {
        border: 1.5px solid #cbd5e1 !important;
        border-radius: 7px !important;
        background: #f0f7ff !important;
        font-size: 0.95rem !important;
    }
    div[data-testid="stForm"] input:focus {
        border-color: #2563eb !important;
        box-shadow: 0 0 0 3px rgba(37,99,235,0.12) !important;
        outline: none !important;
    }

    /* Botão Entrar */
    div[data-testid="stForm"] button[kind="primaryFormSubmit"] {
        background: #2563eb !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-size: 1rem !important;
        font-weight: 700 !important;
        height: 46px !important;
        margin-top: 8px !important;
    }
    div[data-testid="stForm"] button[kind="primaryFormSubmit"]:hover {
        background: #1d4ed8 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    _lc, _mc, _rc = st.columns([1.2, 1, 1.2])
    with _mc:
        with st.form("form_login", clear_on_submit=False):
            st.markdown(
                f"<div style='text-align:center;margin-bottom:28px;'>"
                f"<div style='display:flex;justify-content:center;margin-bottom:14px;'>{_logo_tag}</div>"
                f"<div style='font-size:1.8rem;font-weight:800;color:#0f172a;line-height:1.35;'>"
                f"Conciliador</div>"
                f"<div style='font-size:0.82rem;color:#94a3b8;margin-top:6px;'>"
                f"Faça login para continuar</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
            email = st.text_input("Email", placeholder="usuario@contabilidade.com.br")
            senha = st.text_input("Senha", type="password", placeholder="••••••••")
            entrar = st.form_submit_button("Entrar", use_container_width=True, type="primary")

        st.markdown(
            "<div style='text-align:center;margin-top:16px;font-size:0.85rem;color:white;'>"
            "Não tem uma conta? "
            "<span style='font-weight:700;text-decoration:underline;color:#dbeafe;cursor:pointer;'>"
            "Cadastre-se aqui</span></div>",
            unsafe_allow_html=True,
        )

    if entrar:
        if not email or not senha:
            st.error("Preencha e-mail e senha.")
            return False
        from database.db_manager import DatabaseManager as _DM
        db = _DM()
        usuario = db.autenticar(email, senha)
        if usuario:
            st.session_state["usuario_logado"] = usuario
            st.session_state["empresas_permitidas"] = db.empresas_do_usuario(
                usuario["id"], usuario["perfil"], usuario.get("escritorio_id")
            )
            st.rerun()
        else:
            st.error("E-mail ou senha incorretos.")
    return False


# =============================================================================
# PAINEL ADMIN — IGP
# =============================================================================
def painel_admin_igp():
    """Painel exclusivo para o perfil IGP: gerenciar contabilidades, usuários e vínculos."""
    db = DatabaseManager()

    st.markdown("""
    <div style="background:linear-gradient(90deg,#1e3a5f,#0f766e);color:white;
                padding:20px 28px;border-radius:14px;margin-bottom:24px;">
        <span style="font-size:1.3rem;font-weight:800;">⚙️ Painel Administrativo IGP</span>
        <span style="font-size:0.85rem;opacity:0.8;margin-left:12px;">Gestão de contabilidades e usuários</span>
    </div>
    """, unsafe_allow_html=True)

    aba = st.tabs(["👥 Usuários", "🏢 Contabilidades", "🔗 Vínculo Usuário-Contabilidade", "🏭 Vínculo Empresa-Contabilidade"])

    # ── helpers
    def _listar_contabilidades():
        with db._conn() as _c:
            return [dict(r) for r in _c.execute(
                "SELECT id, nome, cnpj, responsavel FROM escritorio ORDER BY nome"
            ).fetchall()]

    def _empresas_da_contabilidade(escritorio_id):
        """Retorna clientes cujos usuários pertencem a essa contabilidade — via escritório."""
        with db._conn() as _c:
            return [dict(r) for r in _c.execute(
                """SELECT DISTINCT c.id, c.nome FROM clientes c
                   JOIN importacoes i ON i.cliente_id = c.id
                   WHERE c.ativo=1 ORDER BY c.nome"""
            ).fetchall()]

    # ── Aba Usuários
    with aba[0]:
        st.subheader("Usuários do sistema")
        _contabs = _listar_contabilidades()
        _cont_nome = {e["id"]: e["nome"] for e in _contabs}
        usuarios = db.listar_usuarios()
        if usuarios:
            for u in usuarios:
                _cor = {"igp": "#6366f1", "gerente": "#0891b2", "usuario": "#64748b"}.get(u["perfil"], "#94a3b8")
                _icone = {"igp": "👑", "gerente": "🏢", "usuario": "👤"}.get(u["perfil"], "👤")
                _ativo = "✅" if u["ativo"] else "❌"
                _contab_label = u.get("escritorio_nome") or "—"
                with st.expander(f"{_icone} {u['nome']} — {u['email']} {_ativo}  ·  {_contab_label}"):
                    c1, c2, c3 = st.columns(3)
                    c1.markdown(f"**Perfil:** <span style='color:{_cor};font-weight:700'>{u['perfil'].upper()}</span>", unsafe_allow_html=True)
                    c2.markdown(f"**Contabilidade:** {_contab_label}")
                    c3.markdown(f"**Último acesso:** {u.get('ultimo_acesso') or 'nunca'}")
                    _c1, _c2, _c3 = st.columns([2, 1, 1])
                    nova_senha = _c1.text_input("Nova senha", key=f"np_{u['id']}", placeholder="deixe vazio para não alterar")
                    if _c2.button("💾 Salvar senha", key=f"ss_{u['id']}", use_container_width=True):
                        if nova_senha:
                            db.alterar_senha(u["id"], nova_senha)
                            st.success("Senha alterada.")
                    _label_ativo = "🔴 Desativar" if u["ativo"] else "🟢 Ativar"
                    if _c3.button(_label_ativo, key=f"at_{u['id']}", use_container_width=True):
                        db.ativar_desativar_usuario(u["id"], not u["ativo"])
                        st.rerun()

        st.divider()
        st.subheader("Criar novo usuário")
        with st.form("form_novo_usuario"):
            _nc1, _nc2 = st.columns(2)
            _nu_email  = _nc1.text_input("E-mail")
            _nu_nome   = _nc2.text_input("Nome")
            _nc3, _nc4 = st.columns(2)
            _nu_senha  = _nc3.text_input("Senha", type="password")
            _nu_perfil = _nc4.selectbox("Perfil", ["igp", "gerente", "usuario"])
            _cont_opts = {"— nenhuma —": None} | {e["nome"]: e["id"] for e in _contabs}
            _nu_cont = st.selectbox("Contabilidade", list(_cont_opts.keys()))
            _esc_id = _cont_opts[_nu_cont]
            if st.form_submit_button("✅ Criar usuário", type="primary"):
                if _nu_email and _nu_nome and _nu_senha:
                    try:
                        db.criar_usuario(_nu_email, _nu_nome, _nu_senha, _nu_perfil, _esc_id)
                        st.success(f"Usuário {_nu_email} criado.")
                        st.rerun()
                    except Exception as _e:
                        st.error(f"Erro: {_e}")
                else:
                    st.warning("Preencha e-mail, nome e senha.")

    # ── Aba Contabilidades
    with aba[1]:
        st.subheader("Contabilidades cadastradas")
        _contabs_now = _listar_contabilidades()
        if _contabs_now:
            for _e in _contabs_now:
                with st.expander(f"🏢 {_e['nome']}"):
                    _ea, _eb = st.columns(2)
                    _ea.markdown(f"**CNPJ:** {_e.get('cnpj') or '—'}")
                    _eb.markdown(f"**Responsável:** {_e.get('responsavel') or '—'}")
                    # Lista usuários desta contabilidade
                    _us = [u for u in db.listar_usuarios(escritorio_id=_e["id"])]
                    if _us:
                        st.markdown("**Usuários:** " + ", ".join(f"{u['nome']} ({u['perfil']})" for u in _us))
        else:
            st.info("Nenhuma contabilidade cadastrada.")

        st.divider()
        st.subheader("Criar contabilidade")
        with st.form("form_novo_escritorio"):
            _ec1, _ec2 = st.columns(2)
            _e_nome = _ec1.text_input("Nome da contabilidade")
            _e_cnpj = _ec2.text_input("CNPJ (opcional)")
            _e_resp = st.text_input("Responsável (opcional)")
            if st.form_submit_button("✅ Criar", type="primary"):
                if _e_nome:
                    with db._conn() as _c:
                        _c.execute(
                            "INSERT INTO escritorio (nome, cnpj, responsavel) VALUES (?,?,?)",
                            (_e_nome, _e_cnpj or None, _e_resp or None)
                        )
                    st.success(f"Contabilidade '{_e_nome}' criada.")
                    st.rerun()
                else:
                    st.warning("Informe o nome da contabilidade.")

    # ── Aba Vínculos
    with aba[2]:
        st.subheader("Vincular usuário a uma Contabilidade")
        st.caption("O usuário herda acesso a todas as empresas da contabilidade selecionada.")

        _todos_usuarios = db.listar_usuarios()
        _u_opts = {
            f"{u['nome']} ({u['email']})": u
            for u in _todos_usuarios if u["perfil"] in ("gerente", "usuario")
        }
        if not _u_opts:
            st.info("Nenhum usuário do tipo gerente ou usuário cadastrado.")
        else:
            _sel_u = st.selectbox("Selecione o usuário", list(_u_opts.keys()))
            _u_sel = _u_opts[_sel_u]

            _contabs_v = _listar_contabilidades()
            _cont_opts_v = {e["nome"]: e["id"] for e in _contabs_v}
            _atual = _u_sel.get("escritorio_nome") or "— nenhuma —"
            _default_idx = list(_cont_opts_v.keys()).index(_atual) if _atual in _cont_opts_v else 0

            _nova_cont = st.selectbox(
                "Contabilidade",
                list(_cont_opts_v.keys()),
                index=_default_idx,
            )
            st.markdown(
                f"<div style='padding:10px;background:#f0f7ff;border-radius:8px;"
                f"border-left:4px solid #2563eb;margin:8px 0;font-size:0.85rem;'>"
                f"O usuário <b>{_u_sel['nome']}</b> terá acesso a <b>todas as empresas</b> "
                f"da contabilidade <b>{_nova_cont}</b>.</div>",
                unsafe_allow_html=True,
            )
            if st.button("💾 Salvar vínculo", type="primary"):
                _novo_esc_id = _cont_opts_v[_nova_cont]
                with db._conn() as _c:
                    _c.execute(
                        "UPDATE usuarios SET escritorio_id=? WHERE id=?",
                        (_novo_esc_id, _u_sel["id"])
                    )
                st.success(f"Vínculo salvo: {_u_sel['nome']} → {_nova_cont}")
                st.rerun()

    # ── Aba Vínculo Empresa-Contabilidade
    with aba[3]:
        st.subheader("Vincular Empresas a uma Contabilidade")
        st.caption("Selecione a contabilidade e marque quais empresas ela atende.")

        _contabs_e = _listar_contabilidades()
        # Filtra IGP da lista (IGP não é uma contabilidade de clientes)
        _cont_e_opts = {e["nome"]: e["id"] for e in _contabs_e if e["nome"] != "IGP"}
        if not _cont_e_opts:
            st.info("Nenhuma contabilidade cadastrada.")
        else:
            _sel_cont_e = st.selectbox("Contabilidade", list(_cont_e_opts.keys()), key="sel_cont_emp")
            _esc_e_id = _cont_e_opts[_sel_cont_e]

            with db._conn() as _ce:
                _all_cli = [dict(r) for r in _ce.execute(
                    "SELECT id, nome FROM clientes WHERE ativo=1 ORDER BY nome"
                ).fetchall()]
                _ja_vinc = {r["cliente_id"] for r in _ce.execute(
                    "SELECT cliente_id FROM escritorio_clientes WHERE escritorio_id=?", (_esc_e_id,)
                ).fetchall()}

            _cli_e_opts = {c["nome"]: c["id"] for c in _all_cli}
            _default_e = [c["nome"] for c in _all_cli if c["id"] in _ja_vinc]

            # Key inclui o id da contabilidade para resetar ao trocar
            _sel_cli = st.multiselect(
                "Empresas desta contabilidade",
                list(_cli_e_opts.keys()),
                default=_default_e,
                key=f"multisel_emp_cont_{_esc_e_id}"
            )
            if st.button("💾 Salvar vínculos de empresas", type="primary",
                         key=f"btn_salvar_emp_cont_{_esc_e_id}"):
                _ids_cli = [_cli_e_opts[n] for n in _sel_cli]
                db.vincular_empresa_contabilidade(_esc_e_id, _ids_cli)
                _u_log = st.session_state.get("usuario_logado", {})
                if _u_log.get("escritorio_id") == _esc_e_id:
                    st.session_state["empresas_permitidas"] = db.empresas_do_usuario(
                        _u_log["id"], _u_log["perfil"], _esc_e_id
                    )
                st.success(f"✅ {len(_ids_cli)} empresa(s) vinculada(s) a **{_sel_cont_e}**.")
                st.rerun()

        # ── Resumo geral de vínculos
        st.divider()
        st.subheader("📋 Resumo de vínculos por Contabilidade")
        with db._conn() as _cr:
            _resumo = [dict(r) for r in _cr.execute(
                """SELECT e.nome AS contabilidade, c.nome AS empresa
                   FROM escritorio_clientes ec
                   JOIN escritorio e ON e.id = ec.escritorio_id
                   JOIN clientes c ON c.id = ec.cliente_id
                   ORDER BY e.nome, c.nome"""
            ).fetchall()]

        if not _resumo:
            st.info("Nenhum vínculo configurado ainda.")
        else:
            from itertools import groupby as _groupby
            _sorted = sorted(_resumo, key=lambda x: x["contabilidade"])
            for _cont_name, _group in _groupby(_sorted, key=lambda x: x["contabilidade"]):
                _emps = [g["empresa"] for g in _group]
                st.markdown(
                    f"<div style='background:#f8fafc;border-radius:10px;padding:12px 16px;"
                    f"margin-bottom:10px;border-left:4px solid #2563eb;'>"
                    f"<div style='font-weight:700;color:#1e3a5f;font-size:0.95rem;margin-bottom:6px;'>"
                    f"🏢 {_cont_name}</div>"
                    f"<div style='font-size:0.85rem;color:#475569;'>"
                    + "".join(
                        f"<span style='display:inline-block;background:#dbeafe;color:#1d4ed8;"
                        f"border-radius:20px;padding:2px 10px;margin:2px 4px 2px 0;"
                        f"font-size:0.8rem;font-weight:600;'>{e}</span>"
                        for e in _emps
                    )
                    + "</div></div>",
                    unsafe_allow_html=True,
                )


# =============================================================================
# MAIN
# =============================================================================
def main():
    import unicodedata

    # ── Portão de autenticação ─────────────────────────────────────────────────
    usuario = st.session_state.get("usuario_logado")
    if not usuario:
        tela_login()
        st.stop()

    # Botão de logout na sidebar
    with st.sidebar:
        _u = usuario
        _perfil_cor = {"igp": "#6366f1", "gerente": "#0891b2", "usuario": "#64748b"}.get(_u["perfil"], "#94a3b8")
        st.markdown(
            f"<div style='padding:10px 4px 4px;'>"
            f"<div style='font-size:0.78rem;color:#94a3b8;'>Logado como</div>"
            f"<div style='font-weight:700;color:#0f172a;font-size:0.9rem;'>{_u['nome']}</div>"
            f"<div style='font-size:0.75rem;color:{_perfil_cor};font-weight:600;text-transform:uppercase;'>"
            f"{_u['perfil']}"
            f"{'  ·  ' + _u.get('escritorio_nome','') if _u.get('escritorio_nome') else ''}"
            f"</div></div>",
            unsafe_allow_html=True,
        )
        if st.button("🚪 Sair", key="btn_logout", use_container_width=True):
            for _k in list(st.session_state.keys()):
                del st.session_state[_k]
            st.rerun()
        st.divider()

    # ── Roteamento por perfil ──────────────────────────────────────────────────
    # IGP → painel admin
    if usuario["perfil"] == "igp":
        _igp_pg = st.sidebar.radio(
            "Área IGP", ["🖥️ Sistema Normal", "⚙️ Administração"],
            key="igp_area"
        )
        if _igp_pg == "⚙️ Administração":
            painel_admin_igp()
            return

    # ── Decide se mostra hub ou dashboard
    #    Hub aparece quando: nunca houve seleção OU usuário pediu troca explícita
    hub_ativo = st.session_state.get("hub_ativo", None)
    hub_modulo = st.session_state.get("hub_modulo")
    tem_empresa = bool(st.session_state.get("cliente_id_ativo"))

    # Primeira vez na sessão (hub_ativo não existe) e sem empresa → hub
    if hub_ativo is None and not tem_empresa:
        st.session_state["hub_ativo"] = True
        hub_ativo = True

    if hub_modulo == "conciliacao_razao":
        pagina_conciliacao_razao()
        return

    # Se hub explicitamente ativado, mostra hub sem sidebar
    if hub_ativo:
        pagina_hub()
        return

    # ── Dashboard normal
    pagina, cliente_id, cliente_ativo, score_min, usar_sim, tol_dias = renderizar_sidebar()

    # Botão "Trocar Empresa" na sidebar para voltar ao hub
    with st.sidebar:
        st.divider()
        if st.button("🔄 Trocar Empresa", use_container_width=True, key="btn_trocar_empresa"):
            st.session_state["hub_ativo"] = True
            st.session_state.pop("cliente_id_ativo", None)
            st.session_state.pop("cliente_nome_sel", None)
            st.rerun()

    pagina_limpa = pagina.strip()
    pg = "".join(c for c in pagina_limpa if not unicodedata.category(c).startswith("So")).strip()

    if "Visão Geral" in pg:
        if not cliente_ativo:
            st.warning("Selecione ou cadastre um cliente na sidebar.")
        else:
            pagina_visao_geral(cliente_id, cliente_ativo)

    elif "Conciliação" in pg:
        pagina_conciliacao(cliente_id, cliente_ativo, score_min, usar_sim, tol_dias)

    elif "Análise Mensal" in pg:
        if not cliente_ativo:
            st.warning("Selecione um cliente na sidebar.")
        else:
            pagina_analise_mensal(cliente_id, cliente_ativo)

    elif "Clientes" in pg:
        pagina_clientes()

    elif "Escritório" in pg:
        pagina_escritorio()


if __name__ == "__main__":
    main()
