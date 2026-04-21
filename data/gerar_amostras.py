# =============================================================================
# GERADOR DE ARQUIVOS DE AMOSTRA
# Cria arquivos Excel de exemplo (Razão e Extrato) para testes.
# Inclui cenários reais: match exato, combinado, similaridade e não conciliado.
# =============================================================================

import os
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import random

BASE = Path(__file__).parent


def gerar_arquivos_exemplo(pasta: str = None) -> dict:
    """
    Gera dois arquivos Excel de exemplo:
    - razao_exemplo.xlsx  → Lançamentos do Razão contábil no formato Domínio
    - extrato_exemplo.xlsx → Extrato bancário com correspondências variadas
    
    Returns:
        Dict com chaves 'razao' e 'extrato' apontando para os caminhos gerados
    """
    if pasta is None:
        pasta = str(BASE / "samples")
    os.makedirs(pasta, exist_ok=True)

    hoje = datetime.now()
    dia = lambda d: (hoje - timedelta(days=d)).strftime("%d/%m/%Y")
    ts  = lambda d: hoje - timedelta(days=d)

    # =========================================================================
    # ARQUIVO 1 - RAZÃO CONTÁBIL
    # Simula o layout exportado pelo Domínio Web com cabeçalho de conta e
    # lançamentos com: data, histórico, documento, débito, crédito, saldo
    # =========================================================================
    razao_linhas = [
        # Cabeçalho da conta (linha de identificação — será ignorada pelo parser)
        ["CONTA: 1.1.1.01 - BANCO DO BRASIL CONTA CORRENTE", "", "", "", "", ""],
        ["DATA",       "HISTÓRICO",                     "DOCUMENTO",  "DÉBITO",    "CRÉDITO",   "SALDO"],

        # Lançamentos reais (serão capturados)
        [dia(30), "PIX RECEBIDO CLIENTE ABC COMERCIO",    "TXID001",   "5.000,00",  "",          "5.000,00"],
        [dia(28), "TED ENVIADA FORNECEDOR SILVA LTDA",    "TED002",    "",          "12.500,00", "-7.500,00"],
        [dia(25), "PAGAMENTO BOLETO ENERGIA ELETRICA",    "BOL003",    "",          "847,32",    "-8.347,32"],
        [dia(20), "DEPOSITO CHEQUE CLIENTE BETA",         "CHQ004",    "3.200,00",  "",          "-5.147,32"],
        [dia(15), "VENDAS DO DIA 15 SOMA GERAL",          "",          "1.000,00",  "",          "-4.147,32"],
        [dia(10), "FOLHA PAGAMENTO JANEIRO 2024",         "FP001",     "",          "25.000,00", "-29.147,32"],
        [dia(8),  "PIX JOAO SILVA SANTOS",                "",          "750,00",    "",          "-28.397,32"],
        [dia(5),  "PAGAMENTO FORNECEDOR MAQUINAS",        "NF789",     "",          "4.300,00",  "-32.697,32"],
        [dia(3),  "TRANSFERENCIA INTERNA CC",             "TRF099",    "",          "8.000,00",  "-40.697,32"],
        [dia(1),  "ESTORNO TAXA BANCARIA",                "EST001",    "45,90",     "",          "-40.651,42"],
        [dia(0),  "RECEITA SERVICOS NOVEMBRO",            "NF1001",    "2.800,00",  "",          "-37.851,42"],

        # Linha de totalizador (será ignorada pelo parser)
        ["TOTAL:", "",  "",  "12.795,90",  "50.647,32",  ""],
    ]

    df_razao = pd.DataFrame(razao_linhas)
    caminho_razao = os.path.join(pasta, "razao_exemplo.xlsx")

    with pd.ExcelWriter(caminho_razao, engine="openpyxl") as writer:
        df_razao.to_excel(writer, index=False, header=False, sheet_name="Razão")
        # Aplica formatação visual básica
        from openpyxl.styles import Font, PatternFill
        ws = writer.sheets["Razão"]
        # Cabeçalho da conta em azul
        ws["A1"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        ws["A1"].font = Font(color="FFFFFF", bold=True)
        # Largura das colunas
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

    print(f"✅ Razão gerado:   {caminho_razao}")

    # =========================================================================
    # ARQUIVO 2 - EXTRATO BANCÁRIO
    # Simula o extrato do banco com correspondências variadas:
    # - Alguns com match exato (mesmo valor e data)
    # - Alguns divididos em 2 lançamentos (match combinado)
    # - Alguns com descrição diferente (match similaridade)
    # - Alguns sem par (não conciliado)
    # =========================================================================
    extrato_dados = {
        "Data": [
            dia(30),  # Match exato com TXID001
            dia(28),  # Match exato com TED002
            dia(25),  # Match exato com BOL003
            dia(21),  # Match exato com CHQ004 (1 dia de diferença)
            # Combinado: dois lançamentos que somam 1000,00
            dia(15),  # Parte 1 da venda do dia 15
            dia(15),  # Parte 2 da venda do dia 15
            # Combinado: dois departamentos que somam 25.000,00
            dia(10),  # Depto A da folha de pagamento
            dia(10),  # Depto B da folha de pagamento
            # Similaridade: mesmo valor, descrição diferente
            dia(8),   # "PIX J SILVA" ≈ "PIX JOAO SILVA SANTOS"
            dia(5),   # "PGTO FORNEC MAQUINAS" ≈ "PAGAMENTO FORNECEDOR MAQUINAS"
            # Não conciliado do extrato (sem par no Razão)
            dia(2),   # Taxa que não está no Razão
            # Conciliado por documento
            dia(0),   # Match exato com NF1001
        ],
        "Descrição": [
            "PIX RECEBIDO CLIENTE ABC COMERCIO",
            "TED FORNECEDOR SILVA LTDA",
            "PAGAMENTO BOLETO COELBA ENERGIA",
            "DEPOSITO CHEQUE 004",
            "PIX RECEBIDO PARTE 1 VENDAS DIA",
            "PIX RECEBIDO PARTE 2 VENDAS DIA",
            "PAGAMENTO SALARIOS DEPARTAMENTO A",
            "PAGAMENTO SALARIOS DEPARTAMENTO B",
            "PIX J SILVA",
            "PGTO FORNEC MAQUINAS",
            "TAXA MANUTENCAO CONTA CORRENTE",
            "DEPOSITO SERVICOS NOV",
        ],
        "Valor": [
            "5.000,00",
            "12.500,00",
            "847,32",
            "3.200,00",
            "600,00",
            "400,00",
            "15.000,00",
            "10.000,00",
            "750,00",
            "4.300,00",
            "35,00",
            "2.800,00",
        ],
        "Tipo": [
            "C", "D", "D", "C",
            "C", "C",
            "D", "D",
            "C",
            "D",
            "D",
            "C",
        ],
        "Documento": [
            "TXID001", "TED002", "BOL003", "CHQ004",
            "", "",
            "", "",
            "", "NF789",
            "TAX2024",
            "NF1001",
        ],
    }

    df_extrato = pd.DataFrame(extrato_dados)
    caminho_extrato = os.path.join(pasta, "extrato_exemplo.xlsx")

    with pd.ExcelWriter(caminho_extrato, engine="openpyxl") as writer:
        df_extrato.to_excel(writer, index=False, sheet_name="Extrato")
        from openpyxl.styles import Font, PatternFill
        ws = writer.sheets["Extrato"]
        # Cabeçalho
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 14

    print(f"✅ Extrato gerado: {caminho_extrato}")

    return {
        "razao": caminho_razao,
        "extrato": caminho_extrato,
    }


if __name__ == "__main__":
    caminhos = gerar_arquivos_exemplo()
    print("\n📂 Arquivos gerados com sucesso!")
    print(f"   Razão:   {caminhos['razao']}")
    print(f"   Extrato: {caminhos['extrato']}")
    print("\nPara executar a conciliação com esses arquivos:")
    print(f"   python main.py --razao {caminhos['razao']} --extrato {caminhos['extrato']}")
    print("\nPara abrir o dashboard:")
    print("   streamlit run ui/dashboard.py")
